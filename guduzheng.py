import streamlit as st
import pandas as pd
import json
import re
import time
import requests
import datetime
import plotly.express as px
import plotly.graph_objects as go
from typing import List, Dict, Any
import numpy as np
from concurrent.futures import ThreadPoolExecutor
import threading
import io
import base64
import zipfile
import tempfile

# 可选导入 - Excel功能
EXCEL_AVAILABLE = False
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.chart import BarChart, Reference
    EXCEL_AVAILABLE = True
except ImportError:
    st.warning("⚠️ 注意: 未安装openpyxl模块，Excel导出功能不可用。要启用Excel功能，请运行: pip install openpyxl")

# 页面配置
st.set_page_config(page_title="孤独症儿童AI模拟实验平台 - 医学标准版", layout="wide")

# API配置
API_KEY = "sk-DQY3QAIcPGWTMfqZN1itL0qwl3y7ejrqyQwyGLyPom6TGz2v"  # 请替换为您的API Key
API_URL = "https://api.moonshot.cn/v1/chat/completions"

# 初始化session state
if 'experiment_records' not in st.session_state:
    st.session_state.experiment_records = []
if 'current_batch_results' not in st.session_state:
    st.session_state.current_batch_results = []
if 'experiment_progress' not in st.session_state:
    st.session_state.experiment_progress = {'current': 0, 'total': 0}

# 基于DSM-5标准的场景配置
CLINICAL_SCENE_CONFIG = {
    "结构化教学环境": {
        "roles": ["孤独症儿童", "特殊教育老师", "同班同学", "教学助理", "评估师"],
        "target": "观察孤独症儿童在结构化教学环境中的社交沟通行为和适应性表现",
        "desc": "🏫 结构化教学环境（基于TEACCH教学法）",
        "activities": [
            "个别化教学任务", "视觉提示学习", "社交故事训练", 
            "轮流活动练习", "日程表使用训练", "工作系统操作"
        ],
        "triggers": [
            "日程突然改变", "新任务引入", "社交要求增加", 
            "感官刺激过载", "规则变化", "时间压力"
        ],
        "observation_points": [
            "对结构化提示的反应", "任务转换能力", "社交主动性",
            "重复行为频率", "情绪调节表现", "沟通功能性"
        ]
    },
    "自然社交情境": {
        "roles": ["孤独症儿童", "典型发育同伴", "社交技能治疗师", "观察员", "家长"],
        "target": "评估孤独症儿童在自然社交环境中的互动模式和社交认知能力",
        "desc": "👥 自然社交情境（同伴互动观察）", 
        "activities": [
            "自由游戏时间", "合作游戏任务", "角色扮演游戏",
            "分享活动", "冲突解决情境", "集体讨论"
        ],
        "triggers": [
            "同伴拒绝", "游戏规则争议", "注意力竞争",
            "身体接触要求", "情绪表达需求", "帮助请求"
        ],
        "observation_points": [
            "社交主动发起", "非语言沟通", "情绪理解",
            "游戏技能", "冲突处理", "友谊建立"
        ]
    },
    "感官调节环境": {
        "roles": ["孤独症儿童", "职业治疗师", "感统训练师", "护理人员", "评估专家"],
        "target": "观察孤独症儿童的感官处理模式和自我调节策略",
        "desc": "🌈 感官调节环境（感觉统合评估）",
        "activities": [
            "感官探索活动", "深压觉输入", "前庭刺激训练",
            "精细动作练习", "感官休息时间", "适应性行为训练"
        ],
        "triggers": [
            "噪音刺激", "光线变化", "质地变化",
            "运动要求", "拥挤环境", "多重感官输入"
        ],
        "observation_points": [
            "感官寻求行为", "感官逃避反应", "自我调节策略",
            "适应性行为", "注意力集中", "情绪稳定性"
        ]
    },
    "日常生活技能": {
        "roles": ["孤独症儿童", "生活技能训练师", "家庭成员", "康复师", "行为分析师"],
        "target": "评估孤独症儿童在日常生活技能方面的独立性和适应能力",
        "desc": "🏠 日常生活技能训练环境",
        "activities": [
            "自理技能练习", "家务参与", "购物模拟",
            "时间管理", "安全意识训练", "社区适应"
        ],
        "triggers": [
            "程序被打断", "新环境适应", "选择要求",
            "时间限制", "独立要求", "问题解决需求"
        ],
        "observation_points": [
            "独立性水平", "指令理解", "问题解决",
            "灵活性表现", "安全意识", "自我倡导"
        ]
    },
    "语言沟通评估": {
        "roles": ["孤独症儿童", "语言治疗师", "沟通伙伴", "评估师", "技术支持"],
        "target": "专门评估孤独症儿童的语言沟通能力和替代沟通使用",
        "desc": "💬 语言沟通专项评估环境",
        "activities": [
            "语言表达训练", "理解能力测试", "非语言沟通",
            "AAC设备使用", "社交语言练习", "叙事能力评估"
        ],
        "triggers": [
            "复杂指令", "抽象概念", "情绪表达要求",
            "社交语言需求", "新词汇学习", "语用技能挑战"
        ],
        "observation_points": [
            "表达性语言", "接受性语言", "语用技能",
            "非语言沟通", "沟通意图", "对话技能"
        ]
    }
}

# 基于医学标准的孤独症严重程度分级
CLINICAL_AUTISM_PROFILES = {
    "需要支持（轻度）": {
        "social_communication": 3,  # 社交沟通缺陷程度 (1-5, 5为最严重)
        "restricted_repetitive": 2,  # 刻板重复行为程度
        "sensory_processing": 3,     # 感官处理异常
        "cognitive_function": 4,     # 认知功能水平
        "adaptive_behavior": 3,      # 适应行为能力
        "language_level": 4,         # 语言发展水平
        "special_interests": "数学计算、交通工具、地图",
        "support_needs": "最小支持",
        "dsm5_severity": "需要支持"
    },
    "需要大量支持（中度）": {
        "social_communication": 4,
        "restricted_repetitive": 4,
        "sensory_processing": 4,
        "cognitive_function": 2,
        "adaptive_behavior": 2,
        "language_level": 2,
        "special_interests": "旋转物体、特定音乐、重复动作",
        "support_needs": "大量支持",
        "dsm5_severity": "需要大量支持"
    },
    "需要非常大量支持（重度）": {
        "social_communication": 5,
        "restricted_repetitive": 5,
        "sensory_processing": 5,
        "cognitive_function": 1,
        "adaptive_behavior": 1,
        "language_level": 1,
        "special_interests": "光影变化、自我刺激行为、重复发声",
        "support_needs": "非常大量支持",
        "dsm5_severity": "需要非常大量支持"
    },
    "阿斯伯格样表现": {
        "social_communication": 3,
        "restricted_repetitive": 3,
        "sensory_processing": 4,
        "cognitive_function": 5,
        "adaptive_behavior": 3,
        "language_level": 4,
        "special_interests": "科学、历史、编程、收集",
        "support_needs": "部分支持",
        "dsm5_severity": "需要支持（高功能）"
    },
    "共患ADHD": {
        "social_communication": 3,
        "restricted_repetitive": 3,
        "sensory_processing": 4,
        "cognitive_function": 3,
        "adaptive_behavior": 2,
        "language_level": 3,
        "special_interests": "运动、游戏、动态活动",
        "support_needs": "中等支持",
        "dsm5_severity": "需要支持+注意缺陷多动障碍"
    }
}

# 基于CARS、ABC、SCQ等量表的综合评估指标
CLINICAL_EVALUATION_METRICS = {
    "社交互动质量": {
        "description": "社会情感互惠性缺陷的程度",
        "subscales": {
            "社交发起": "主动发起社交互动的频率和质量",
            "社交反应": "对他人社交信号的反应性和适当性", 
            "社交维持": "维持社交互动的能力和持续性",
            "情感分享": "分享情感和兴趣的能力"
        },
        "scoring_criteria": {
            5: "严重缺陷 - 极少主动社交，不回应社交信号",
            4: "明显缺陷 - 社交发起有限，反应不当",
            3: "中度缺陷 - 可以社交但质量差，需要支持", 
            2: "轻度缺陷 - 基本社交能力，偶有困难",
            1: "无明显缺陷 - 社交互动基本正常"
        }
    },
    "沟通交流能力": {
        "description": "语言和非语言沟通的缺陷程度",
        "subscales": {
            "表达性沟通": "语言或替代方式表达需求和想法的能力",
            "接受性沟通": "理解他人语言和非语言信息的能力",
            "社交性沟通": "在社交情境中使用沟通的语用技能",
            "非语言沟通": "眼神、手势、表情等非语言沟通的使用"
        },
        "scoring_criteria": {
            5: "严重缺陷 - 无功能性沟通或仅有少量单词",
            4: "明显缺陷 - 有限的功能性语言，语用严重受损",
            3: "中度缺陷 - 基本沟通能力但语用技能差",
            2: "轻度缺陷 - 沟通基本流畅但有社交语用困难",
            1: "无明显缺陷 - 沟通能力年龄适宜"
        }
    },
    "刻板重复行为": {
        "description": "限制性重复行为模式的严重程度",
        "subscales": {
            "刻板动作": "重复的运动模式或动作的频率和强度",
            "仪式化行为": "坚持固定程序和仪式的程度",
            "狭隘兴趣": "异常强烈或限制性兴趣的程度",
            "感官行为": "异常感官兴趣或感官寻求/逃避行为"
        },
        "scoring_criteria": {
            5: "严重程度 - 重复行为严重干扰功能",
            4: "明显程度 - 重复行为明显且影响日常活动",
            3: "中度程度 - 有明显重复行为但可以中断",
            2: "轻度程度 - 偶有重复行为，影响轻微",
            1: "无或极轻微 - 很少重复行为"
        }
    },
    "感官处理能力": {
        "description": "感官信息处理的异常程度",
        "subscales": {
            "感官过敏": "对感官刺激过度敏感的程度",
            "感官寻求": "主动寻求感官刺激的行为",
            "感官调节": "调节感官输入的能力",
            "感官整合": "整合多种感官信息的能力"
        },
        "scoring_criteria": {
            5: "严重异常 - 感官问题严重影响日常功能",
            4: "明显异常 - 感官敏感性明显且经常出现",
            3: "中度异常 - 有感官处理困难需要策略支持",
            2: "轻度异常 - 偶有感官敏感但可以适应",
            1: "正常范围 - 感官处理基本正常"
        }
    },
    "情绪行为调节": {
        "description": "情绪识别、表达和调节的能力",
        "subscales": {
            "情绪识别": "识别自己和他人情绪的能力",
            "情绪表达": "适当表达情绪的能力",
            "情绪调节": "管理和调节情绪的策略和能力",
            "行为控制": "控制冲动和不当行为的能力"
        },
        "scoring_criteria": {
            5: "严重困难 - 情绪失调频繁，自伤或攻击行为",
            4: "明显困难 - 经常情绪爆发，调节能力差",
            3: "中度困难 - 情绪调节有困难但有改善空间",
            2: "轻度困难 - 大多数时候能调节情绪",
            1: "良好 - 情绪调节年龄适宜"
        }
    },
    "认知适应功能": {
        "description": "认知功能和适应性行为的水平",
        "subscales": {
            "学习能力": "学习新技能和概念的能力",
            "问题解决": "解决日常问题的能力",
            "执行功能": "计划、组织和灵活思维的能力",
            "适应性行为": "在日常环境中独立功能的能力"
        },
        "scoring_criteria": {
            5: "重度缺陷 - 明显智力障碍，适应能力极差",
            4: "中度缺陷 - 学习困难，需要大量支持",
            3: "轻度缺陷 - 学习能力有限但可训练",
            2: "边缘水平 - 认知能力基本正常但有弱项",
            1: "正常范围 - 认知和适应功能年龄适宜"
        }
    }
}

def call_kimi_with_profile(prompt, autism_profile, max_retries=3):
    """调用AI API生成对话，带重试机制，基于临床特征描述"""
    headers = {
        "Authorization": f"Bearer {API_KEY}",
        "Content-Type": "application/json"
    }
    
    profile_description = f"""
    孤独症儿童临床特征配置（基于DSM-5标准）：
    - DSM-5严重程度: {autism_profile.get('dsm5_severity', '未指定')}
    - 社交沟通缺陷程度: {autism_profile['social_communication']}/5 (5为最严重)
    - 刻板重复行为程度: {autism_profile['restricted_repetitive']}/5
    - 感官处理异常: {autism_profile['sensory_processing']}/5
    - 认知功能水平: {autism_profile['cognitive_function']}/5 (1为重度障碍，5为正常)
    - 适应行为能力: {autism_profile['adaptive_behavior']}/5
    - 语言发展水平: {autism_profile['language_level']}/5
    - 特殊兴趣领域: {autism_profile['special_interests']}
    - 所需支持水平: {autism_profile.get('support_needs', '未指定')}
    """
    
    system_prompt = (
        "你是一个专业的孤独症临床行为专家，请严格按照以下DSM-5诊断标准和临床特征来模拟孤独症儿童的行为：\n"
        + profile_description +
        "\n核心症状表现要求："
        "\n1. 社交沟通缺陷：根据严重程度展现眼神回避、社交发起困难、情感分享受限等"
        "\n2. 刻板重复行为：体现重复动作、仪式化行为、狭隘兴趣、感官异常等"
        "\n3. 感官处理：根据敏感度显示过敏、寻求或逃避特定感官刺激"
        "\n4. 语言特点：根据语言水平展现回声式语言、字面理解、语用困难等"
        "\n5. 情绪调节：根据调节能力显示情绪爆发、自我安抚行为等"
        "\n严格格式：\"角色名:发言内容\"。每句换行，语言真实自然，行为符合临床观察特点。"
    )
    
    payload = {
        "model": "moonshot-v1-8k",
        "messages": [
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": prompt}
        ],
        "temperature": 0.6,
        "max_tokens": 2048
    }
    
    for attempt in range(max_retries):
        try:
            response = requests.post(API_URL, headers=headers, json=payload, timeout=30)
            
            if response.status_code == 200:
                return response.json()['choices'][0]['message']['content']
            elif response.status_code == 429:
                wait_time = (attempt + 1) * 5
                print(f"API速率限制，等待{wait_time}秒后重试...")
                if 'st' in globals():
                    st.warning(f"⏱️ API速率限制，等待{wait_time}秒后重试... (尝试 {attempt + 1}/{max_retries})")
                time.sleep(wait_time)
                continue
            else:
                return f"API调用失败: {response.status_code} - {response.text}"
                
        except requests.exceptions.Timeout:
            if attempt == max_retries - 1:
                return "API调用超时，请稍后重试"
            time.sleep(2)
            continue
        except Exception as e:
            if attempt == max_retries - 1:
                return f"网络错误: {str(e)}"
            time.sleep(2)
            continue
    
    return "API调用失败：超过最大重试次数"

def clinical_evaluate_dialogue(dialogue, autism_profile, scene_info):
    """基于临床标准评估对话质量"""
    lines = dialogue.split('\n')
    autism_child_lines = [line for line in lines if '孤独症儿童' in line]
    
    if not autism_child_lines:
        return {metric: 0.0 for metric in CLINICAL_EVALUATION_METRICS.keys()}
    
    evaluation_scores = {}
    
    # 社交互动质量评估
    social_base = autism_profile['social_communication']
    interaction_indicators = 0
    
    # 检查社交发起行为
    proactive_social = len([line for line in autism_child_lines if any(word in line for word in ['我想', '我们一起', '可以吗', '你好'])])
    if proactive_social > 0:
        interaction_indicators += 1
    
    # 检查社交反应
    responsive_social = len([line for line in autism_child_lines if any(word in line for word in ['好的', '是的', '不要', '谢谢'])])
    if responsive_social > len(autism_child_lines) * 0.3:
        interaction_indicators += 1
    
    social_score = social_base - (interaction_indicators * 0.5)
    evaluation_scores["社交互动质量"] = max(1, min(5, social_score))
    
    # 沟通交流能力评估
    comm_base = autism_profile['social_communication']
    communication_quality = 0
    
    # 检查语言功能性
    functional_language = len([line for line in autism_child_lines if any(word in line for word in ['我要', '帮助', '不懂', '为什么'])])
    if functional_language > 0:
        communication_quality += 1
    
    # 检查回声式语言（重复他人话语）
    echolalia_signs = len([line for line in autism_child_lines if '?' in line and len(line.split()) < 5])
    if echolalia_signs > 0:
        communication_quality -= 0.5
    
    # 根据语言水平调整
    language_modifier = (autism_profile['language_level'] - 3) * 0.3
    comm_score = comm_base - communication_quality + language_modifier
    evaluation_scores["沟通交流能力"] = max(1, min(5, comm_score))
    
    # 刻板重复行为评估
    repetitive_base = autism_profile['restricted_repetitive']
    repetitive_indicators = 0
    
    # 检查重复表达
    repeated_phrases = len(set(autism_child_lines)) / len(autism_child_lines) if autism_child_lines else 1
    if repeated_phrases < 0.7:  # 如果重复率高
        repetitive_indicators += 1
    
    # 检查特殊兴趣相关表达
    special_interest_mentions = len([line for line in autism_child_lines 
                                   if any(interest.lower() in line.lower() 
                                         for interest in autism_profile['special_interests'].split('、'))])
    if special_interest_mentions > len(autism_child_lines) * 0.3:
        repetitive_indicators += 1
    
    repetitive_score = repetitive_base + (repetitive_indicators * 0.5)
    evaluation_scores["刻板重复行为"] = max(1, min(5, repetitive_score))
    
    # 感官处理能力评估  
    sensory_base = autism_profile['sensory_processing']
    sensory_responses = 0
    
    # 检查感官相关反应
    sensory_words = ['太吵', '太亮', '不喜欢', '害怕', '疼', '舒服']
    sensory_mentions = len([line for line in autism_child_lines 
                           if any(word in line for word in sensory_words)])
    if sensory_mentions > 0:
        sensory_responses = min(sensory_mentions * 0.3, 1.0)
    
    sensory_score = sensory_base - (sensory_responses * 0.5)
    evaluation_scores["感官处理能力"] = max(1, min(5, sensory_score))
    
    # 情绪行为调节评估
    emotion_base = autism_profile['social_communication']  # 基于社交沟通缺陷推断情绪调节
    emotion_regulation = 0
    
    # 检查情绪表达
    emotion_words = ['开心', '难过', '生气', '害怕', '着急', '不高兴']
    emotion_expressions = len([line for line in autism_child_lines 
                              if any(word in line for word in emotion_words)])
    if emotion_expressions > 0:
        emotion_regulation += 0.5
    
    # 检查调节策略
    regulation_words = ['我需要', '休息', '安静', '停止']
    regulation_attempts = len([line for line in autism_child_lines 
                              if any(word in line for word in regulation_words)])
    if regulation_attempts > 0:
        emotion_regulation += 0.5
    
    emotion_score = emotion_base - emotion_regulation
    evaluation_scores["情绪行为调节"] = max(1, min(5, emotion_score))
    
    # 认知适应功能评估
    cognitive_base = 6 - autism_profile['cognitive_function']  # 转换评分方向
    adaptation_quality = 0
    
    # 检查问题解决尝试
    problem_solving = len([line for line in autism_child_lines 
                          if any(word in line for word in ['怎么办', '试试', '想想', '办法'])])
    if problem_solving > 0:
        adaptation_quality += 0.5
    
    # 检查学习表现
    learning_indicators = len([line for line in autism_child_lines 
                              if any(word in line for word in ['学会', '知道了', '明白', '记住'])])
    if learning_indicators > 0:
        adaptation_quality += 0.5
    
    cognitive_score = cognitive_base - adaptation_quality
    evaluation_scores["认知适应功能"] = max(1, min(5, cognitive_score))
    
    # 添加随机变异模拟真实评估的不确定性
    for metric in evaluation_scores:
        variation = np.random.normal(0, 0.2)  # 小幅随机变化
        evaluation_scores[metric] = max(1, min(5, evaluation_scores[metric] + variation))
        evaluation_scores[metric] = round(evaluation_scores[metric], 2)
    
    return evaluation_scores

def generate_experiment_batch(templates, scenes, num_experiments_per_combo=3):
    """生成批量实验配置"""
    experiments = []
    experiment_counter = 0
    
    for template_name, profile in templates.items():
        for scene_name, scene_data in scenes.items():
            for activity in scene_data['activities'][:2]:
                for trigger in scene_data['triggers'][:2]:
                    for i in range(num_experiments_per_combo):
                        experiment_counter += 1
                        
                        # 添加轻微临床变异
                        varied_profile = profile.copy()
                        for key in ['social_communication', 'restricted_repetitive', 'sensory_processing']:
                            if key in varied_profile:
                                variation = np.random.randint(-1, 2)
                                varied_profile[key] = max(1, min(5, varied_profile[key] + variation))
                        
                        timestamp = datetime.datetime.now().strftime('%Y%m%d_%H%M%S_%f')[:-3]
                        unique_id = f"CLIN_{experiment_counter:03d}_{template_name[:4]}_{scene_name[:4]}_{timestamp}"
                        
                        experiments.append({
                            'template': template_name,
                            'scene': scene_name,
                            'activity': activity,
                            'trigger': trigger,
                            'autism_profile': varied_profile,
                            'experiment_id': unique_id,
                            'batch_index': experiment_counter
                        })
    
    return experiments

def run_single_experiment(experiment_config):
    """运行单个实验"""
    try:
        scene_data = CLINICAL_SCENE_CONFIG[experiment_config['scene']]
        
        # 构建基于临床观察的prompt
        prompt = (
            f"临床观察情境：{experiment_config['scene']} - {experiment_config['activity']}\n"
            f"观察要点：{', '.join(scene_data['observation_points'][:3])}\n"
            f"触发因素：{experiment_config['trigger']}\n"
            f"参与角色：孤独症儿童、{scene_data['roles'][1]}、{scene_data['roles'][2]}\n"
            f"请基于DSM-5孤独症诊断标准，模拟该儿童在此情境下的真实行为表现。\n"
            f"要求：15-20轮对话，体现核心症状（社交沟通缺陷、刻板重复行为），"
            f"语言和行为符合该严重程度的临床特征。\n"
            f"格式：'角色名:内容'，每句换行。"
        )
        
        dialogue = call_kimi_with_profile(prompt, experiment_config['autism_profile'])
        
        # 使用临床标准评估
        evaluation_scores = clinical_evaluate_dialogue(
            dialogue, 
            experiment_config['autism_profile'], 
            CLINICAL_SCENE_CONFIG[experiment_config['scene']]
        )
        
        record = {
            'experiment_id': experiment_config['experiment_id'],
            'timestamp': datetime.datetime.now(),
            'template': experiment_config['template'],
            'scene': experiment_config['scene'],
            'activity': experiment_config['activity'],
            'trigger': experiment_config['trigger'],
            'autism_profile': experiment_config['autism_profile'],
            'dialogue': dialogue,
            'evaluation_scores': evaluation_scores,
            'clinical_observations': extract_clinical_observations(dialogue),
            'notes': f"临床标准评估 - {experiment_config['template']}"
        }
        
        return record
        
    except Exception as e:
        return {
            'experiment_id': experiment_config['experiment_id'],
            'timestamp': datetime.datetime.now(),
            'error': str(e),
            'template': experiment_config.get('template', 'unknown'),
            'scene': experiment_config.get('scene', 'unknown')
        }

def extract_clinical_observations(dialogue):
    """从对话中提取临床观察要点"""
    lines = dialogue.split('\n')
    autism_child_lines = [line for line in lines if '孤独症儿童' in line]
    
    observations = {
        "社交行为观察": [],
        "语言沟通特点": [],
        "重复行为表现": [],
        "感官反应": [],
        "情绪调节": []
    }
    
    for line in autism_child_lines:
        # 社交行为识别
        if any(word in line for word in ['你好', '再见', '一起', '朋友']):
            observations["社交行为观察"].append("主动社交尝试")
        elif any(word in line for word in ['不要', '不喜欢', '走开']):
            observations["社交行为观察"].append("社交回避行为")
        
        # 语言特点识别  
        if line.count('是') > 2 or line.count('不') > 2:
            observations["语言沟通特点"].append("回声式语言特征")
        if any(word in line for word in ['为什么', '什么时候', '在哪里']):
            observations["语言沟通特点"].append("疑问句使用")
        
        # 重复行为识别
        if any(word in line for word in ['又', '还要', '一直', '再']):
            observations["重复行为表现"].append("重复需求表达")
        
        # 感官反应识别
        if any(word in line for word in ['太吵', '太亮', '疼', '痒']):
            observations["感官反应"].append("感官过敏反应")
        
        # 情绪识别
        if any(word in line for word in ['生气', '难过', '害怕', '开心']):
            observations["情绪调节"].append("情绪表达尝试")
    
    # 清理空列表
    observations = {k: v for k, v in observations.items() if v}
    
    return observations

def run_batch_experiments(experiments, progress_callback=None):
    """运行批量实验"""
    results = []
    delay_between_requests = 25
    
    for i, experiment in enumerate(experiments):
        if progress_callback:
            progress_callback(i + 1, len(experiments))
        
        if 'st' in globals():
            remaining_experiments = len(experiments) - i - 1
            estimated_time = remaining_experiments * delay_between_requests / 60
            st.info(f"⏳ 正在处理第 {i+1}/{len(experiments)} 个临床实验，预计还需 {estimated_time:.1f} 分钟")
        
        result = run_single_experiment(experiment)
        results.append(result)
        
        if i < len(experiments) - 1:
            print(f"等待{delay_between_requests}秒避免API限制...")
            if 'st' in globals():
                progress_bar = st.progress(0)
                for wait_second in range(delay_between_requests):
                    progress_bar.progress((wait_second + 1) / delay_between_requests)
                    time.sleep(1)
                progress_bar.empty()
            else:
                time.sleep(delay_between_requests)
    
    return results

def generate_clinical_analysis(records):
    """生成临床标准的统计分析报告"""
    if not records:
        return {}
    
    analysis = {}
    
    # 基础临床统计
    analysis['临床评估概况'] = {
        '评估总数': len(records),
        '评估时间跨度': f"{min(r['timestamp'] for r in records).strftime('%Y-%m-%d')} 至 {max(r['timestamp'] for r in records).strftime('%Y-%m-%d')}",
        '涉及情境数': len(set(r['scene'] for r in records)),
        '涉及严重程度数': len(set(r.get('template', '自定义') for r in records))
    }
    
    # 按严重程度分析
    severity_stats = {}
    for record in records:
        severity = record.get('template', '自定义')
        if severity not in severity_stats:
            severity_stats[severity] = {
                '评估次数': 0,
                '社交互动得分': [],
                '沟通交流得分': [],
                '刻板行为得分': [],
                '感官处理得分': [],
                '情绪调节得分': [],
                '认知适应得分': []
            }
        severity_stats[severity]['评估次数'] += 1
        severity_stats[severity]['社交互动得分'].append(record['evaluation_scores']['社交互动质量'])
        severity_stats[severity]['沟通交流得分'].append(record['evaluation_scores']['沟通交流能力'])
        severity_stats[severity]['刻板行为得分'].append(record['evaluation_scores']['刻板重复行为'])
        severity_stats[severity]['感官处理得分'].append(record['evaluation_scores']['感官处理能力'])
        severity_stats[severity]['情绪调节得分'].append(record['evaluation_scores']['情绪行为调节'])
        severity_stats[severity]['认知适应得分'].append(record['evaluation_scores']['认知适应功能'])
    
    # 计算统计值
    for severity, stats in severity_stats.items():
        for metric in ['社交互动得分', '沟通交流得分', '刻板行为得分', '感官处理得分', '情绪调节得分', '认知适应得分']:
            scores = stats[metric]
            stats[f'{metric}_均值'] = np.mean(scores)
            stats[f'{metric}_标准差'] = np.std(scores)
            stats[f'{metric}_范围'] = f"{np.min(scores):.1f}-{np.max(scores):.1f}"
            del stats[metric]
    
    analysis['严重程度分析'] = severity_stats
    
    # 按评估情境分析
    context_stats = {}
    for record in records:
        context = record['scene']
        if context not in context_stats:
            context_stats[context] = {
                '评估次数': 0,
                '社交表现': [],
                '沟通表现': [],
                '适应表现': []
            }
        context_stats[context]['评估次数'] += 1
        context_stats[context]['社交表现'].append(record['evaluation_scores']['社交互动质量'])
        context_stats[context]['沟通表现'].append(record['evaluation_scores']['沟通交流能力'])
        context_stats[context]['适应表现'].append(record['evaluation_scores']['认知适应功能'])
    
    for context, stats in context_stats.items():
        for metric in ['社交表现', '沟通表现', '适应表现']:
            scores = stats[metric]
            stats[f'{metric}_均值'] = np.mean(scores)
            del stats[metric]
    
    analysis['情境分析'] = context_stats
    
    # 整体临床表现
    all_social = [r['evaluation_scores']['社交互动质量'] for r in records]
    all_comm = [r['evaluation_scores']['沟通交流能力'] for r in records]
    all_repetitive = [r['evaluation_scores']['刻板重复行为'] for r in records]
    all_sensory = [r['evaluation_scores']['感官处理能力'] for r in records]
    all_emotion = [r['evaluation_scores']['情绪行为调节'] for r in records]
    all_cognitive = [r['evaluation_scores']['认知适应功能'] for r in records]
    
    analysis['整体临床表现'] = {
        '社交互动缺陷程度': f"{np.mean(all_social):.2f} ± {np.std(all_social):.2f}",
        '沟通交流缺陷程度': f"{np.mean(all_comm):.2f} ± {np.std(all_comm):.2f}",
        '刻板重复行为程度': f"{np.mean(all_repetitive):.2f} ± {np.std(all_repetitive):.2f}",
        '感官处理异常程度': f"{np.mean(all_sensory):.2f} ± {np.std(all_sensory):.2f}",
        '情绪调节困难程度': f"{np.mean(all_emotion):.2f} ± {np.std(all_emotion):.2f}",
        '认知适应缺陷程度': f"{np.mean(all_cognitive):.2f} ± {np.std(all_cognitive):.2f}",
        '核心症状综合严重度': f"{(np.mean(all_social) + np.mean(all_comm) + np.mean(all_repetitive))/3:.2f}"
    }
    
    # 临床发现和建议
    findings = []
    
    # 分析核心症状
    core_symptom_avg = (np.mean(all_social) + np.mean(all_comm) + np.mean(all_repetitive)) / 3
    if core_symptom_avg >= 4.0:
        findings.append("核心症状严重，建议密集型干预治疗")
    elif core_symptom_avg >= 3.0:
        findings.append("核心症状中等，建议结构化教学和行为干预")
    else:
        findings.append("核心症状相对较轻，建议社交技能训练")
    
    # 分析共患情况
    if np.mean(all_sensory) >= 4.0:
        findings.append("存在明显感官处理异常，建议感觉统合治疗")
    
    if np.mean(all_emotion) >= 4.0:
        findings.append("情绪调节困难显著，建议心理行为干预")
    
    # 分析最优情境
    if context_stats:
        best_context = min(context_stats.keys(), 
                          key=lambda x: (context_stats[x]['社交表现_均值'] + 
                                       context_stats[x]['沟通表现_均值']) / 2)
        findings.append(f"在{best_context}中表现相对较好，可作为干预起点")
    
    analysis['临床发现与建议'] = findings
    
    return analysis

def create_clinical_excel_report(records, analysis):
    """创建临床标准的Excel报告"""
    if not EXCEL_AVAILABLE:
        return None
    
    output = io.BytesIO()
    workbook = Workbook()
    workbook.remove(workbook.active)
    
    # 1. 临床评估概览
    overview_sheet = workbook.create_sheet("临床评估概览")
    overview_sheet.append(["孤独症儿童临床行为评估报告（基于DSM-5标准）"])
    overview_sheet.append([])
    overview_sheet.append(["报告生成时间", datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')])
    overview_sheet.append(["评估标准", "DSM-5孤独症诊断标准 + CARS/ABC/SCQ量表"])
    overview_sheet.append([])
    
    overview_sheet.append(["评估概况"])
    for key, value in analysis.get('临床评估概况', {}).items():
        overview_sheet.append([key, value])
    
    overview_sheet.append([])
    overview_sheet.append(["整体临床表现"])
    for key, value in analysis.get('整体临床表现', {}).items():
        overview_sheet.append([key, value])
    
    overview_sheet.append([])
    overview_sheet.append(["临床发现与建议"])
    for finding in analysis.get('临床发现与建议', []):
        overview_sheet.append([finding])
    
    # 2. 详细评估数据
    data_sheet = workbook.create_sheet("详细评估数据")
    headers = ["评估ID", "时间", "严重程度", "评估情境", "观察活动", "触发因素",
              "社交互动缺陷", "沟通交流缺陷", "刻板重复行为", "感官处理异常", 
              "情绪调节困难", "认知适应缺陷", "核心症状严重度",
              "DSM-5严重程度", "所需支持水平", "特殊兴趣", "备注"]
    data_sheet.append(headers)
    
    for record in records:
        profile = record.get('autism_profile', {})
        scores = record['evaluation_scores']
        core_symptom_severity = (scores['社交互动质量'] + scores['沟通交流能力'] + scores['刻板重复行为']) / 3
        
        row = [
            record['experiment_id'],
            record['timestamp'].strftime('%Y-%m-%d %H:%M:%S'),
            record.get('template', '自定义'),
            record['scene'],
            record.get('activity', ''),
            record.get('trigger', ''),
            scores['社交互动质量'],
            scores['沟通交流能力'],
            scores['刻板重复行为'],
            scores['感官处理能力'],
            scores['情绪行为调节'],
            scores['认知适应功能'],
            f"{core_symptom_severity:.2f}",
            profile.get('dsm5_severity', ''),
            profile.get('support_needs', ''),
            profile.get('special_interests', ''),
            record.get('notes', '')
        ]
        data_sheet.append(row)
    
    # 3. 严重程度对比分析
    if analysis.get('严重程度分析'):
        severity_sheet = workbook.create_sheet("严重程度分析")
        severity_headers = ["严重程度", "评估次数", "社交缺陷均值", "沟通缺陷均值", 
                          "刻板行为均值", "感官异常均值", "情绪困难均值", "认知缺陷均值",
                          "核心症状综合"]
        severity_sheet.append(severity_headers)
        
        for severity, stats in analysis['严重程度分析'].items():
            core_avg = (stats['社交互动得分_均值'] + stats['沟通交流得分_均值'] + stats['刻板行为得分_均值']) / 3
            row = [
                severity,
                stats['评估次数'],
                f"{stats['社交互动得分_均值']:.2f}",
                f"{stats['沟通交流得分_均值']:.2f}",
                f"{stats['刻板行为得分_均值']:.2f}",
                f"{stats['感官处理得分_均值']:.2f}",
                f"{stats['情绪调节得分_均值']:.2f}",
                f"{stats['认知适应得分_均值']:.2f}",
                f"{core_avg:.2f}"
            ]
            severity_sheet.append(row)
    
    # 4. 临床观察记录
    if any('clinical_observations' in record for record in records):
        obs_sheet = workbook.create_sheet("临床观察记录")
        obs_sheet.append(["评估ID", "社交行为观察", "语言沟通特点", "重复行为表现", "感官反应", "情绪调节"])
        
        for record in records:
            if 'clinical_observations' in record:
                obs = record['clinical_observations']
                row = [
                    record['experiment_id'],
                    '; '.join(obs.get('社交行为观察', [])),
                    '; '.join(obs.get('语言沟通特点', [])),
                    '; '.join(obs.get('重复行为表现', [])),
                    '; '.join(obs.get('感官反应', [])),
                    '; '.join(obs.get('情绪调节', []))
                ]
                obs_sheet.append(row)
    
    # 5. 对话记录（用于质性分析）
    dialogue_sheet = workbook.create_sheet("对话记录")
    dialogue_sheet.append(["评估ID", "严重程度", "评估情境", "对话内容"])
    
    for record in records:
        dialogue_sheet.append([
            record['experiment_id'],
            record.get('template', '自定义'),
            record['scene'],
            record['dialogue']
        ])
    
    # 应用专业样式
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if cell.row == 1:
                    cell.font = Font(bold=True, size=12)
                    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    cell.font = Font(bold=True, color="FFFFFF")
                elif any(keyword in str(cell.value) for keyword in ['严重', '明显', '需要支持']) if cell.value else False:
                    cell.fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
                cell.alignment = Alignment(wrap_text=True, vertical="top")
    
    workbook.save(output)
    output.seek(0)
    return output.getvalue()

# 主页面
st.title("🏥 孤独症儿童AI模拟实验平台 - 医学标准版")
st.markdown("**基于DSM-5诊断标准和权威评估量表（CARS、ABC、SCQ、M-CHAT等）**")

# 侧边栏导航
st.sidebar.title("🔍 导航")
page = st.sidebar.selectbox("选择功能页面", [
    "临床快速评估", "批量临床研究", "个性化评估设计", 
    "临床数据分析", "评估记录管理", "📊 临床报告中心"
])

# 页面路由
if page == "临床快速评估":
    st.header("🩺 临床快速评估")
    st.markdown("使用标准化严重程度分级进行快速临床行为评估")
    
    col1, col2 = st.columns(2)
    
    with col1:
        st.subheader("📋 选择评估对象")
        selected_severity = st.selectbox("严重程度分级", list(CLINICAL_AUTISM_PROFILES.keys()))
        
        profile = CLINICAL_AUTISM_PROFILES[selected_severity]
        
        # 显示临床特征
        with st.expander("查看DSM-5特征配置", expanded=True):
            st.write(f"**DSM-5严重程度**: {profile['dsm5_severity']}")
            st.write(f"**社交沟通缺陷**: {profile['social_communication']}/5")
            st.write(f"**刻板重复行为**: {profile['restricted_repetitive']}/5")
            st.write(f"**感官处理异常**: {profile['sensory_processing']}/5")
            st.write(f"**认知功能水平**: {profile['cognitive_function']}/5")
            st.write(f"**适应行为能力**: {profile['adaptive_behavior']}/5")
            st.write(f"**语言发展水平**: {profile['language_level']}/5")
            st.write(f"**特殊兴趣**: {profile['special_interests']}")
            st.write(f"**所需支持**: {profile['support_needs']}")
        
        selected_scene = st.selectbox("选择评估情境", list(CLINICAL_SCENE_CONFIG.keys()))
        
        scene_data = CLINICAL_SCENE_CONFIG[selected_scene]
        
        # 显示场景信息
        with st.expander("评估情境详情"):
            st.write(f"**目标**: {scene_data['target']}")
            st.write(f"**观察要点**: {', '.join(scene_data['observation_points'])}")
        
        selected_activity = st.selectbox("选择观察活动", scene_data['activities'])
        selected_trigger = st.selectbox("选择触发因素", scene_data['triggers'])
    
    with col2:
        st.subheader("🔬 执行评估")
        
        if st.button("🩺 开始临床评估", type="primary", use_container_width=True):
            timestamp = datetime.datetime.now().strftime('%Y%m%d_%H%M%S_%f')[:-3]
            experiment_config = {
                'template': selected_severity,
                'scene': selected_scene,
                'activity': selected_activity,
                'trigger': selected_trigger,
                'autism_profile': profile.copy(),
                'experiment_id': f"CLIN_{selected_severity[:4]}_{timestamp}"
            }
            
            with st.spinner("🤖 正在生成临床评估对话..."):
                result = run_single_experiment(experiment_config)
            
            if 'error' not in result:
                st.session_state.experiment_records.append(result)
                
                st.success(f"✅ 临床评估完成！ID: {result['experiment_id']}")
                
                # 显示评估结果
                st.subheader("📊 临床评估结果")
                
                col_result1, col_result2 = st.columns(2)
                
                with col_result1:
                    st.write("**核心症状评估得分** (5分为最严重):")
                    for metric, score in result['evaluation_scores'].items():
                        # 根据得分显示不同颜色
                        if score >= 4.0:
                            st.error(f"{metric}: {score}/5.0 (严重)")
                        elif score >= 3.0:
                            st.warning(f"{metric}: {score}/5.0 (中度)")
                        else:
                            st.success(f"{metric}: {score}/5.0 (轻度)")
                
                with col_result2:
                    st.write("**临床观察要点**:")
                    if 'clinical_observations' in result:
                        for category, observations in result['clinical_observations'].items():
                            if observations:
                                st.write(f"**{category}**: {', '.join(observations)}")
                    
                    st.write("**对话预览**:")
                    dialogue_lines = result['dialogue'].split('\n')[:8]
                    for line in dialogue_lines:
                        if ':' in line and line.strip():
                            if '孤独症儿童' in line:
                                st.markdown(f"🧒 {line}")
                            else:
                                st.markdown(f"👤 {line}")
                    
                    if len(result['dialogue'].split('\n')) > 8:
                        st.markdown("*...查看完整记录请前往'评估记录管理'*")
                
                # 显示临床建议
                st.subheader("💡 临床建议")
                core_avg = (result['evaluation_scores']['社交互动质量'] + 
                           result['evaluation_scores']['沟通交流能力'] + 
                           result['evaluation_scores']['刻板重复行为']) / 3
                
                if core_avg >= 4.0:
                    st.error("🚨 建议：核心症状严重，需要密集型干预和专业支持")
                elif core_avg >= 3.0:
                    st.warning("⚠️ 建议：核心症状中等，建议结构化教学和行为干预")
                else:
                    st.success("✅ 建议：症状相对较轻，可重点进行社交技能训练")
                    
            else:
                st.error(f"❌ 评估失败: {result['error']}")

elif page == "批量临床研究":
    st.header("🔬 批量临床研究")
    st.markdown("进行多组临床对照研究，获取统计学有效的评估数据")
    
    col1, col2 = st.columns([2, 1])
    
    with col1:
        st.subheader("🎯 研究设计")
        
        research_scale = st.radio(
            "选择研究规模",
            ["试点研究（推荐）", "标准研究", "大样本研究"],
            help="根据研究需要选择合适的样本规模"
        )
        
        if research_scale == "试点研究（推荐）":
            default_severities = list(CLINICAL_AUTISM_PROFILES.keys())[:2]
            default_contexts = list(CLINICAL_SCENE_CONFIG.keys())[:2]
            default_repeats = 1
            st.info("🚀 试点研究：验证评估工具效果，约需5-8分钟")
        elif research_scale == "标准研究":
            default_severities = list(CLINICAL_AUTISM_PROFILES.keys())[:3]
            default_contexts = list(CLINICAL_SCENE_CONFIG.keys())[:3]
            default_repeats = 2
            st.info("⏳ 标准研究：获得可靠统计数据，约需20-30分钟")
        else:
            default_severities = list(CLINICAL_AUTISM_PROFILES.keys())
            default_contexts = list(CLINICAL_SCENE_CONFIG.keys())
            default_repeats = 2
            st.warning("⏰ 大样本研究：完整临床研究数据，约需60-90分钟")
        
        selected_severities = st.multiselect(
            "选择严重程度组", 
            list(CLINICAL_AUTISM_PROFILES.keys()),
            default=default_severities
        )
        
        selected_contexts = st.multiselect(
            "选择评估情境",
            list(CLINICAL_SCENE_CONFIG.keys()),
            default=default_contexts
        )
        
        repeats_per_combo = st.slider(
            "每组合重复次数", 
            1, 3, 
            default_repeats,
            help="增加重复次数提高统计可靠性"
        )
        
        if selected_severities and selected_contexts:
            severity_dict = {k: CLINICAL_AUTISM_PROFILES[k] for k in selected_severities}
            context_dict = {k: CLINICAL_SCENE_CONFIG[k] for k in selected_contexts}
            
            experiments = generate_experiment_batch(
                severity_dict, 
                context_dict, 
                repeats_per_combo
            )
            
            st.info(f"📊 将生成 {len(experiments)} 个临床评估")
            
            # 研究设计预览
            with st.expander("研究设计预览", expanded=False):
                preview_df = pd.DataFrame([
                    {
                        '严重程度': exp['template'],
                        '评估情境': exp['scene'],
                        '观察活动': exp['activity'],
                        '触发因素': exp['trigger']
                    } for exp in experiments[:10]
                ])
                st.dataframe(preview_df)
                if len(experiments) > 10:
                    st.write(f"*...还有 {len(experiments) - 10} 个评估*")
    
    with col2:
        st.subheader("🚀 执行研究")
        
        if 'clinical_batch_ready' not in st.session_state:
            st.session_state.clinical_batch_ready = False
        if 'clinical_batch_running' not in st.session_state:
            st.session_state.clinical_batch_running = False
        
        if selected_severities and selected_contexts:
            estimated_minutes = len(experiments) * 25 / 60
            st.info(f"📊 评估数量: {len(experiments)}")
            st.info(f"⏰ 预计时间: {estimated_minutes:.1f} 分钟")
            
            if not st.session_state.clinical_batch_ready and not st.session_state.clinical_batch_running:
                if st.button("⚡ 准备开始研究", use_container_width=True):
                    st.session_state.clinical_batch_ready = True
                    st.rerun()
            
            elif st.session_state.clinical_batch_ready and not st.session_state.clinical_batch_running:
                st.warning("⏰ **重要**: 由于API限制，批量研究需要较长时间")
                st.info("💡 确认后将立即开始，请保持网络稳定")
                
                col_btn1, col_btn2 = st.columns(2)
                
                with col_btn1:
                    if st.button("❌ 取消", use_container_width=True):
                        st.session_state.clinical_batch_ready = False
                        st.rerun()
                
                with col_btn2:
                    if st.button("✅ 开始研究", type="primary", use_container_width=True):
                        st.session_state.clinical_batch_running = True
                        st.session_state.clinical_batch_ready = False
                        st.rerun()
            
            elif st.session_state.clinical_batch_running:
                st.success("🔬 临床研究进行中...")
                
                progress_container = st.container()
                with progress_container:
                    progress_bar = st.progress(0)
                    status_text = st.empty()
                    result_container = st.empty()
                
                def update_progress(current, total):
                    progress = current / total
                    progress_bar.progress(progress)
                    remaining_time = (total - current) * 25 / 60
                    status_text.text(f"研究进度: {current}/{total} ({progress:.1%}) - 预计还需 {remaining_time:.1f} 分钟")
                
                try:
                    results = run_batch_experiments(experiments, update_progress)
                    
                    successful_results = [r for r in results if 'error' not in r]
                    failed_results = [r for r in results if 'error' in r]
                    
                    st.session_state.experiment_records.extend(successful_results)
                    st.session_state.current_batch_results = successful_results
                    
                    with result_container:
                        st.success(f"✅ 临床研究完成！")
                        st.write(f"**成功评估**: {len(successful_results)} 个")
                        
                        if failed_results:
                            st.error(f"**失败评估**: {len(failed_results)} 个")
                        
                        if successful_results:
                            st.subheader("📈 研究结果概览")
                            
                            # 按严重程度统计
                            severity_stats = {}
                            for result in successful_results:
                                severity = result['template']
                                if severity not in severity_stats:
                                    severity_stats[severity] = []
                                
                                # 计算核心症状综合得分
                                core_score = (result['evaluation_scores']['社交互动质量'] + 
                                            result['evaluation_scores']['沟通交流能力'] + 
                                            result['evaluation_scores']['刻板重复行为']) / 3
                                severity_stats[severity].append(core_score)
                            
                            stats_df = pd.DataFrame([
                                {
                                    '严重程度': severity,
                                    '样本数': len(scores),
                                    '核心症状均值': f"{np.mean(scores):.2f}",
                                    '标准差': f"{np.std(scores):.2f}",
                                    '95%置信区间': f"{np.mean(scores) - 1.96*np.std(scores)/np.sqrt(len(scores)):.2f}-{np.mean(scores) + 1.96*np.std(scores)/np.sqrt(len(scores)):.2f}"
                                } for severity, scores in severity_stats.items()
                            ])
                            
                            st.dataframe(stats_df, use_container_width=True)
                    
                    st.session_state.clinical_batch_running = False
                    
                    if st.button("🔄 进行新研究", use_container_width=True):
                        st.session_state.clinical_batch_ready = False
                        st.session_state.clinical_batch_running = False
                        st.rerun()
                        
                except Exception as e:
                    st.error(f"❌ 研究出错: {str(e)}")
                    st.session_state.clinical_batch_running = False
                    if st.button("🔄 重新尝试", use_container_width=True):
                        st.rerun()
        
        else:
            st.error("请先选择严重程度和评估情境")

elif page == "个性化评估设计":
    st.header("⚙️ 个性化评估设计")
    st.markdown("基于DSM-5标准自定义个体化临床评估参数")
    
    col1, col2 = st.columns(2)
    
    with col1:
        st.subheader("🎭 评估情境设置")
        selected_scene = st.selectbox("选择评估情境", list(CLINICAL_SCENE_CONFIG.keys()))
        scene_data = CLINICAL_SCENE_CONFIG[selected_scene]
        
        st.info(f"**评估目标**: {scene_data['target']}")
        
        # 显示观察要点
        with st.expander("临床观察要点"):
            for point in scene_data['observation_points']:
                st.write(f"• {point}")
        
        selected_activity = st.selectbox("选择观察活动", scene_data['activities'])
        selected_trigger = st.selectbox("选择触发因素", scene_data['triggers'])
    
    with col2:
        st.subheader("👤 DSM-5特征配置")
        
        template_base = st.selectbox("基于标准分级", ["自定义"] + list(CLINICAL_AUTISM_PROFILES.keys()))
        
        if template_base != "自定义":
            base_profile = CLINICAL_AUTISM_PROFILES[template_base]
            st.info(f"基于: {base_profile['dsm5_severity']}")
        else:
            base_profile = {
                'social_communication': 3,
                'restricted_repetitive': 3,
                'sensory_processing': 3,
                'cognitive_function': 3,
                'adaptive_behavior': 3,
                'language_level': 3,
                'special_interests': "特定物体、活动或主题",
                'support_needs': "中等支持",
                'dsm5_severity': "自定义配置"
            }
        
        st.write("**核心症状配置** (基于DSM-5诊断标准A、B条目)")
        
        social_comm = st.slider(
            "A. 社交沟通缺陷程度", 1, 5, base_profile['social_communication'],
            help="1=无明显缺陷，5=严重缺陷（社会情感互惠性、非语言沟通、关系发展困难）"
        )
        
        repetitive_behavior = st.slider(
            "B. 刻板重复行为程度", 1, 5, base_profile['restricted_repetitive'],
            help="1=很少重复行为，5=严重重复行为（刻板动作、仪式、狭隘兴趣、感官异常）"
        )
        
        st.write("**相关功能配置**")
        
        sensory_processing = st.slider(
            "感官处理异常程度", 1, 5, base_profile['sensory_processing'],
            help="1=正常处理，5=严重异常（过敏、寻求、逃避）"
        )
        
        cognitive_function = st.slider(
            "认知功能水平", 1, 5, base_profile['cognitive_function'],
            help="1=重度障碍，5=正常范围"
        )
        
        adaptive_behavior = st.slider(
            "适应行为能力", 1, 5, base_profile['adaptive_behavior'],
            help="1=严重困难，5=年龄适宜"
        )
        
        language_level = st.slider(
            "语言发展水平", 1, 5, base_profile['language_level'],
            help="1=无功能性语言，5=年龄适宜"
        )
        
        special_interests = st.text_input(
            "特殊兴趣/限制性兴趣", 
            base_profile['special_interests'],
            help="描述具体的特殊兴趣或重复行为"
        )
        
        # 根据配置自动推断支持需求
        total_severity = social_comm + repetitive_behavior
        if total_severity >= 8:
            support_level = "需要非常大量支持"
            dsm5_level = "需要非常大量支持"
        elif total_severity >= 6:
            support_level = "需要大量支持"
            dsm5_level = "需要大量支持"
        else:
            support_level = "需要支持"
            dsm5_level = "需要支持"
        
        st.info(f"**推断的DSM-5严重程度**: {dsm5_level}")
        st.info(f"**推断的支持需求**: {support_level}")
        
        autism_profile = {
            'social_communication': social_comm,
            'restricted_repetitive': repetitive_behavior,
            'sensory_processing': sensory_processing,
            'cognitive_function': cognitive_function,
            'adaptive_behavior': adaptive_behavior,
            'language_level': language_level,
            'special_interests': special_interests,
            'support_needs': support_level,
            'dsm5_severity': dsm5_level
        }
    
    # 执行个性化评估
    st.subheader("🔬 执行个性化评估")
    
    if st.button("🩺 开始个性化评估", type="primary", use_container_width=True):
        timestamp = datetime.datetime.now().strftime('%Y%m%d_%H%M%S_%f')[:-3]
        experiment_config = {
            'template': template_base if template_base != "自定义" else "个性化配置",
            'scene': selected_scene,
            'activity': selected_activity,
            'trigger': selected_trigger,
            'autism_profile': autism_profile,
            'experiment_id': f"CUSTOM_{timestamp}"
        }
        
        with st.spinner("🤖 正在生成个性化评估..."):
            result = run_single_experiment(experiment_config)
        
        if 'error' not in result:
            st.session_state.experiment_records.append(result)
            
            st.success(f"✅ 个性化评估完成！ID: {result['experiment_id']}")
            
            # 显示详细评估结果
            st.subheader("📊 个性化评估结果")
            
            col_res1, col_res2, col_res3 = st.columns(3)
            
            with col_res1:
                st.write("**核心症状评估**:")
                st.metric("社交沟通缺陷", f"{result['evaluation_scores']['社交互动质量']:.2f}/5")
                st.metric("刻板重复行为", f"{result['evaluation_scores']['刻板重复行为']:.2f}/5")
                
                core_avg = (result['evaluation_scores']['社交互动质量'] + 
                           result['evaluation_scores']['刻板重复行为']) / 2
                st.metric("核心症状综合", f"{core_avg:.2f}/5")
            
            with col_res2:
                st.write("**相关功能评估**:")
                st.metric("沟通交流能力", f"{result['evaluation_scores']['沟通交流能力']:.2f}/5")
                st.metric("感官处理能力", f"{result['evaluation_scores']['感官处理能力']:.2f}/5")
                st.metric("情绪行为调节", f"{result['evaluation_scores']['情绪行为调节']:.2f}/5")
                st.metric("认知适应功能", f"{result['evaluation_scores']['认知适应功能']:.2f}/5")
            
            with col_res3:
                st.write("**临床观察**:")
                if 'clinical_observations' in result:
                    for category, observations in result['clinical_observations'].items():
                        if observations:
                            st.write(f"**{category}**:")
                            for obs in observations:
                                st.write(f"• {obs}")
                
            # 个性化建议
            st.subheader("💡 个性化干预建议")
            
            suggestions = []
            
            if result['evaluation_scores']['社交互动质量'] >= 4:
                suggestions.append("🎯 优先目标：社交技能训练（眼神接触、轮流交替、情感分享）")
            
            if result['evaluation_scores']['沟通交流能力'] >= 4:
                suggestions.append("🗣️ 沟通干预：语言治疗、AAC辅助沟通、社交语用训练")
            
            if result['evaluation_scores']['刻板重复行为'] >= 4:
                suggestions.append("🔄 行为管理：功能性行为分析、替代行为训练、环境结构化")
            
            if result['evaluation_scores']['感官处理能力'] >= 4:
                suggestions.append("🌈 感官支持：感觉统合治疗、环境调适、自我调节策略")
            
            if result['evaluation_scores']['情绪行为调节'] >= 4:
                suggestions.append("😌 情绪支持：情绪识别训练、应对策略教学、行为干预")
            
            if not suggestions:
                suggestions.append("✅ 整体表现良好，建议维持现有支持并监测发展")
            
            for suggestion in suggestions:
                st.success(suggestion)
                
        else:
            st.error(f"❌ 评估失败: {result['error']}")
    
    # 保存配置
    if st.button("💾 保存评估配置", use_container_width=True):
        st.session_state.custom_autism_profile = autism_profile
        st.session_state.custom_scene_config = {
            'scene': selected_scene,
            'activity': selected_activity,
            'trigger': selected_trigger
        }
        st.success("✅ 个性化配置已保存！")

elif page == "临床数据分析":
    st.header("📈 临床数据分析")
    
    records = st.session_state.experiment_records
    
    if not records:
        st.warning("📊 暂无评估数据，请先进行临床评估")
        st.stop()
    
    # 生成临床分析
    analysis = generate_clinical_analysis(records)
    
    # 临床概况
    st.subheader("🏥 临床评估概况")
    
    col1, col2, col3, col4 = st.columns(4)
    with col1:
        st.metric("评估总数", len(records))
    with col2:
        severities = [r.get('template', '自定义') for r in records]
        most_common = max(set(severities), key=severities.count) if severities else "无"
        st.metric("主要严重程度", most_common.split('（')[0])
    with col3:
        contexts = [r['scene'] for r in records]
        most_used_context = max(set(contexts), key=contexts.count)
        st.metric("主要评估情境", most_used_context.replace('结构化', '结构'))
    with col4:
        all_core_scores = []
        for r in records:
            core_score = (r['evaluation_scores']['社交互动质量'] + 
                         r['evaluation_scores']['沟通交流能力'] + 
                         r['evaluation_scores']['刻板重复行为']) / 3
            all_core_scores.append(core_score)
        avg_core = np.mean(all_core_scores)
        st.metric("平均核心症状", f"{avg_core:.2f}/5")
    
    # DSM-5核心症状分析
    st.subheader("🧠 DSM-5核心症状分析")
    
    # 核心症状雷达图
    avg_scores = {}
    for metric in CLINICAL_EVALUATION_METRICS.keys():
        scores = [r['evaluation_scores'][metric] for r in records]
        avg_scores[metric] = np.mean(scores)
    
    fig_radar = go.Figure()
    fig_radar.add_trace(go.Scatterpolar(
        r=list(avg_scores.values()),
        theta=list(avg_scores.keys()),
        fill='toself',
        name='平均缺陷程度',
        line_color='rgb(255, 100, 100)'
    ))
    fig_radar.update_layout(
        polar=dict(
            radialaxis=dict(
                visible=True,
                range=[1, 5],
                tickvals=[1, 2, 3, 4, 5],
                ticktext=['正常', '轻度', '中度', '明显', '严重']
            )),
        showlegend=True,
        title="DSM-5核心症状及相关功能评估雷达图",
        height=500
    )
    st.plotly_chart(fig_radar, use_container_width=True)
    
    # 严重程度对比分析
    st.subheader("📊 严重程度组间对比")
    
    if len(set([r.get('template', '自定义') for r in records])) > 1:
        severity_data = {}
        for record in records:
            severity = record.get('template', '自定义')
            if severity not in severity_data:
                severity_data[severity] = {
                    '社交沟通缺陷': [],
                    '刻板重复行为': [],
                    '感官处理异常': [],
                    '认知适应缺陷': []
                }
            
            severity_data[severity]['社交沟通缺陷'].append(
                (record['evaluation_scores']['社交互动质量'] + 
                 record['evaluation_scores']['沟通交流能力']) / 2
            )
            severity_data[severity]['刻板重复行为'].append(
                record['evaluation_scores']['刻板重复行为']
            )
            severity_data[severity]['感官处理异常'].append(
                record['evaluation_scores']['感官处理能力']
            )
            severity_data[severity]['认知适应缺陷'].append(
                record['evaluation_scores']['认知适应功能']
            )
        
        # 创建对比图表
        comparison_data = []
        for severity, metrics in severity_data.items():
            for metric, scores in metrics.items():
                comparison_data.append({
                    '严重程度': severity,
                    '症状域': metric,
                    '平均得分': np.mean(scores),
                    '标准差': np.std(scores)
                })
        
        df_comparison = pd.DataFrame(comparison_data)
        
        fig_comparison = px.bar(
            df_comparison, 
            x='严重程度', 
            y='平均得分', 
            color='症状域',
            title="不同严重程度组的症状域对比",
            labels={'平均得分': '缺陷程度 (1-5分)'},
            height=400
        )
        fig_comparison.update_layout(yaxis_range=[1, 5])
        st.plotly_chart(fig_comparison, use_container_width=True)
    
    # 评估情境效应分析
    st.subheader("🎭 评估情境效应分析")
    
    context_data = {}
    for record in records:
        context = record['scene']
        if context not in context_data:
            context_data[context] = []
        
        # 计算综合表现得分（得分越低表现越好）
        comprehensive_score = np.mean(list(record['evaluation_scores'].values()))
        context_data[context].append(comprehensive_score)
    
    if len(context_data) > 1:
        context_comparison = []
        for context, scores in context_data.items():
            context_comparison.append({
                '评估情境': context,
                '样本数': len(scores),
                '平均表现': np.mean(scores),
                '标准差': np.std(scores),
                '表现水平': '较好' if np.mean(scores) < 3.0 else '中等' if np.mean(scores) < 4.0 else '困难'
            })
        
        df_context = pd.DataFrame(context_comparison)
        
        fig_context = px.bar(
            df_context,
            x='评估情境',
            y='平均表现',
            color='表现水平',
            title="不同评估情境下的表现对比",
            labels={'平均表现': '平均困难程度 (1-5分)'},
            height=400
        )
        st.plotly_chart(fig_context, use_container_width=True)
        
        # 显示情境分析表格
        st.dataframe(df_context.drop('表现水平', axis=1), use_container_width=True)
    
    # 临床发现和建议
    st.subheader("🔍 临床发现与干预建议")
    
    if analysis.get('临床发现与建议'):
        col_finding1, col_finding2 = st.columns(2)
        
        with col_finding1:
            st.write("### 📋 主要临床发现")
            for i, finding in enumerate(analysis['临床发现与建议'], 1):
                if '建议' in finding:
                    st.success(f"{i}. {finding}")
                elif '严重' in finding:
                    st.error(f"{i}. {finding}")
                else:
                    st.info(f"{i}. {finding}")
        
        with col_finding2:
            st.write("### 💡 循证干预建议")
            
            # 基于评估结果提供具体建议
            social_avg = np.mean([r['evaluation_scores']['社交互动质量'] for r in records])
            comm_avg = np.mean([r['evaluation_scores']['沟通交流能力'] for r in records])
            repetitive_avg = np.mean([r['evaluation_scores']['刻板重复行为'] for r in records])
            
            st.write("**基于循证实践的干预建议**:")
            
            if social_avg >= 4.0:
                st.write("• 🎯 **社交技能训练** (SST)")
                st.write("  - 结构化社交技能教学")
                st.write("  - 同伴中介干预")
                st.write("  - 视频建模技术")
            
            if comm_avg >= 4.0:
                st.write("• 🗣️ **沟通干预**")
                st.write("  - 功能性沟通训练")
                st.write("  - 图片交换沟通系统(PECS)")
                st.write("  - 语言行为干预")
            
            if repetitive_avg >= 4.0:
                st.write("• 🔄 **行为干预**")
                st.write("  - 应用行为分析(ABA)")
                st.write("  - 功能性行为评估")
                st.write("  - 正向行为支持")
    
    # 统计显著性检验（如果有多组数据）
    severities = [r.get('template', '自定义') for r in records]
    if len(set(severities)) > 1:
        st.subheader("📐 统计学分析")
        
        try:
            from scipy import stats
            
            # 进行方差分析
            groups = {}
            for record in records:
                severity = record.get('template', '自定义')
                if severity not in groups:
                    groups[severity] = []
                
                core_score = (record['evaluation_scores']['社交互动质量'] + 
                             record['evaluation_scores']['沟通交流能力'] + 
                             record['evaluation_scores']['刻板重复行为']) / 3
                groups[severity].append(core_score)
            
            if len(groups) >= 2:
                group_values = list(groups.values())
                f_stat, p_value = stats.f_oneway(*group_values)
                
                st.write(f"**单因素方差分析结果**:")
                st.write(f"- F统计量: {f_stat:.3f}")
                st.write(f"- p值: {p_value:.3f}")
                
                if p_value < 0.05:
                    st.success("✅ 不同严重程度组间差异具有统计学意义 (p < 0.05)")
                else:
                    st.info("ℹ️ 不同严重程度组间差异无统计学意义 (p ≥ 0.05)")
        
        except ImportError:
            st.info("💡 安装scipy模块可启用统计学分析功能")

elif page == "评估记录管理":
    st.header("📚 评估记录管理")
    
    records = st.session_state.experiment_records
    
    if not records:
        st.info("📝 暂无评估记录")
        st.stop()
    
    st.subheader(f"📊 共有 {len(records)} 条临床评估记录")
    
    # 高级筛选选项
    col_filter1, col_filter2, col_filter3, col_filter4 = st.columns(4)
    
    with col_filter1:
        severity_filter = st.selectbox(
            "按严重程度筛选", 
            ["全部"] + list(set([r.get('template', '自定义') for r in records]))
        )
    
    with col_filter2:
        context_filter = st.selectbox(
            "按评估情境筛选",
            ["全部"] + list(set([r['scene'] for r in records]))
        )
    
    with col_filter3:
        score_filter = st.selectbox(
            "按严重程度筛选",
            ["全部", "轻度 (1-2分)", "中度 (2-3分)", "重度 (3-4分)", "极重度 (4-5分)"]
        )
    
    with col_filter4:
        sort_by = st.selectbox(
            "排序方式", 
            ["时间倒序", "时间正序", "核心症状严重度", "社交缺陷程度", "沟通缺陷程度"]
        )
    
    # 应用筛选
    filtered_records = records
    
    if severity_filter != "全部":
        filtered_records = [r for r in filtered_records if r.get('template', '自定义') == severity_filter]
    
    if context_filter != "全部":
        filtered_records = [r for r in filtered_records if r['scene'] == context_filter]
    
    if score_filter != "全部":
        def get_core_score(record):
            return (record['evaluation_scores']['社交互动质量'] + 
                   record['evaluation_scores']['沟通交流能力'] + 
                   record['evaluation_scores']['刻板重复行为']) / 3
        
        if score_filter == "轻度 (1-2分)":
            filtered_records = [r for r in filtered_records if get_core_score(r) <= 2.0]
        elif score_filter == "中度 (2-3分)":
            filtered_records = [r for r in filtered_records if 2.0 < get_core_score(r) <= 3.0]
        elif score_filter == "重度 (3-4分)":
            filtered_records = [r for r in filtered_records if 3.0 < get_core_score(r) <= 4.0]
        elif score_filter == "极重度 (4-5分)":
            filtered_records = [r for r in filtered_records if get_core_score(r) > 4.0]
    
    # 应用排序
    if sort_by == "时间正序":
        filtered_records = sorted(filtered_records, key=lambda x: x['timestamp'])
    elif sort_by == "核心症状严重度":
        filtered_records = sorted(filtered_records, 
            key=lambda x: (x['evaluation_scores']['社交互动质量'] + 
                          x['evaluation_scores']['沟通交流能力'] + 
                          x['evaluation_scores']['刻板重复行为']) / 3, 
            reverse=True)
    elif sort_by == "社交缺陷程度":
        filtered_records = sorted(filtered_records, 
            key=lambda x: x['evaluation_scores']['社交互动质量'], reverse=True)
    elif sort_by == "沟通缺陷程度":
        filtered_records = sorted(filtered_records, 
            key=lambda x: x['evaluation_scores']['沟通交流能力'], reverse=True)
    else:  # 时间倒序
        filtered_records = sorted(filtered_records, key=lambda x: x['timestamp'], reverse=True)
    
    st.write(f"筛选后记录数: {len(filtered_records)}")
    
    # 记录列表显示
    for i, record in enumerate(filtered_records):
        
        # 计算核心症状严重度
        core_severity = (record['evaluation_scores']['社交互动质量'] + 
                        record['evaluation_scores']['沟通交流能力'] + 
                        record['evaluation_scores']['刻板重复行为']) / 3
        
        severity_label = ""
        if core_severity >= 4.0:
            severity_label = "🔴 极重度"
        elif core_severity >= 3.0:
            severity_label = "🟠 重度"
        elif core_severity >= 2.0:
            severity_label = "🟡 中度"
        else:
            severity_label = "🟢 轻度"
        
        template_info = f" - {record.get('template', '自定义')}" if record.get('template') else ""
        
        with st.expander(f"🩺 {record['experiment_id']}{template_info} - {record['scene']} - {severity_label} ({record['timestamp'].strftime('%Y-%m-%d %H:%M')})"):
            
            col1, col2, col3 = st.columns(3)
            
            with col1:
                st.write("**📋 评估基本信息:**")
                if record.get('template'):
                    st.write(f"• 严重程度分级: {record['template']}")
                st.write(f"• 评估情境: {record['scene']}")
                st.write(f"• 观察活动: {record.get('activity', '未指定')}")
                st.write(f"• 触发因素: {record.get('trigger', '未指定')}")
                st.write(f"• 评估时间: {record['timestamp'].strftime('%Y-%m-%d %H:%M:%S')}")
                
                if record.get('autism_profile'):
                    st.write("**👤 DSM-5特征配置:**")
                    profile = record['autism_profile']
                    st.write(f"• DSM-5严重程度: {profile.get('dsm5_severity', 'N/A')}")
                    st.write(f"• 社交沟通缺陷: {profile.get('social_communication', 'N/A')}/5")
                    st.write(f"• 刻板重复行为: {profile.get('restricted_repetitive', 'N/A')}/5")
                    st.write(f"• 认知功能水平: {profile.get('cognitive_function', 'N/A')}/5")
                    st.write(f"• 特殊兴趣: {profile.get('special_interests', 'N/A')}")
            
            with col2:
                st.write("**📊 临床评估得分:**")
                
                scores = record['evaluation_scores']
                
                # 核心症状
                st.write("*DSM-5核心症状:*")
                social_score = scores['社交互动质量']
                comm_score = scores['沟通交流能力']
                repetitive_score = scores['刻板重复行为']
                
                if social_score >= 4.0:
                    st.error(f"• 社交互动质量: {social_score}/5 (严重缺陷)")
                elif social_score >= 3.0:
                    st.warning(f"• 社交互动质量: {social_score}/5 (明显缺陷)")
                else:
                    st.success(f"• 社交互动质量: {social_score}/5 (轻度缺陷)")
                
                if comm_score >= 4.0:
                    st.error(f"• 沟通交流能力: {comm_score}/5 (严重缺陷)")
                elif comm_score >= 3.0:
                    st.warning(f"• 沟通交流能力: {comm_score}/5 (明显缺陷)")
                else:
                    st.success(f"• 沟通交流能力: {comm_score}/5 (轻度缺陷)")
                
                if repetitive_score >= 4.0:
                    st.error(f"• 刻板重复行为: {repetitive_score}/5 (严重程度)")
                elif repetitive_score >= 3.0:
                    st.warning(f"• 刻板重复行为: {repetitive_score}/5 (明显程度)")
                else:
                    st.success(f"• 刻板重复行为: {repetitive_score}/5 (轻度程度)")
                
                # 相关功能
                st.write("*相关功能:*")
                st.write(f"• 感官处理能力: {scores['感官处理能力']}/5")
                st.write(f"• 情绪行为调节: {scores['情绪行为调节']}/5")
                st.write(f"• 认知适应功能: {scores['认知适应功能']}/5")
                
                st.write(f"**核心症状综合严重度: {core_severity:.2f}/5**")
            
            with col3:
                st.write("**🔍 临床观察记录:**")
                if 'clinical_observations' in record and record['clinical_observations']:
                    for category, observations in record['clinical_observations'].items():
                        if observations:
                            st.write(f"*{category}:*")
                            for obs in observations:
                                st.write(f"• {obs}")
                else:
                    st.write("暂无特殊临床观察记录")
                
                if record.get('notes'):
                    st.write(f"**📝 备注:** {record['notes']}")
            
            # 对话记录
            st.write("**💬 行为观察对话记录:**")
            dialogue_lines = record['dialogue'].split('\n')
            dialogue_text = '\n'.join([line for line in dialogue_lines if line.strip() and ':' in line])
            
            unique_key = f"clinical_dialogue_{i}_{record['experiment_id']}_{record['timestamp'].strftime('%Y%m%d_%H%M%S')}"
            st.text_area("", dialogue_text, height=200, key=unique_key)
            
            # 快速操作按钮
            col_btn1, col_btn2, col_btn3 = st.columns(3)
            
            with col_btn1:
                if st.button(f"📋 生成个案报告", key=f"report_{record['experiment_id']}"):
                    st.info("个案报告生成功能开发中...")
            
            with col_btn2:
                if st.button(f"📈 趋势分析", key=f"trend_{record['experiment_id']}"):
                    st.info("个体趋势分析功能开发中...")
            
            with col_btn3:
                if st.button(f"🔄 重复评估", key=f"repeat_{record['experiment_id']}"):
                    st.info("重复评估功能开发中...")

elif page == "📊 临床报告中心":
    st.header("📊 临床报告中心")
    st.markdown("基于循证实践生成专业的临床评估报告和研究数据")
    
    records = st.session_state.experiment_records
    
    if not records:
        st.warning("📊 暂无评估数据，请先进行临床评估")
        st.stop()
    
    st.success(f"📊 当前共有 {len(records)} 条临床评估记录可生成报告")
    
    # 报告类型选择
    st.subheader("📋 选择报告类型")
    
    col1, col2 = st.columns(2)
    
    with col1:
        st.write("### 📄 标准临床报告")
        
        # 基础CSV报告
        if st.button("📊 下载基础评估数据 (CSV)", use_container_width=True):
            df_export = []
            for record in records:
                export_row = {
                    '评估ID': record['experiment_id'],
                    '评估时间': record['timestamp'].strftime('%Y-%m-%d %H:%M:%S'),
                    '严重程度分级': record.get('template', '自定义'),
                    '评估情境': record['scene'],
                    '观察活动': record.get('activity', ''),
                    '触发因素': record.get('trigger', ''),
                    '社交互动缺陷程度': record['evaluation_scores']['社交互动质量'],
                    '沟通交流缺陷程度': record['evaluation_scores']['沟通交流能力'],
                    '刻板重复行为程度': record['evaluation_scores']['刻板重复行为'],
                    '感官处理异常程度': record['evaluation_scores']['感官处理能力'],
                    '情绪调节困难程度': record['evaluation_scores']['情绪行为调节'],
                    '认知适应缺陷程度': record['evaluation_scores']['认知适应功能'],
                    '备注': record.get('notes', '')
                }
                
                # 添加DSM-5特征
                if record.get('autism_profile'):
                    profile = record['autism_profile']
                    export_row.update({
                        'DSM5严重程度': profile.get('dsm5_severity', ''),
                        '所需支持水平': profile.get('support_needs', ''),
                        '社交沟通缺陷设置': profile.get('social_communication', ''),
                        '刻板重复行为设置': profile.get('restricted_repetitive', ''),
                        '感官处理异常设置': profile.get('sensory_processing', ''),
                        '认知功能水平设置': profile.get('cognitive_function', ''),
                        '适应行为能力设置': profile.get('adaptive_behavior', ''),
                        '语言发展水平设置': profile.get('language_level', ''),
                        '特殊兴趣描述': profile.get('special_interests', '')
                    })
                
                # 计算核心症状综合严重度
                core_severity = (record['evaluation_scores']['社交互动质量'] + 
                               record['evaluation_scores']['沟通交流能力'] + 
                               record['evaluation_scores']['刻板重复行为']) / 3
                export_row['核心症状综合严重度'] = round(core_severity, 2)
                
                df_export.append(export_row)
            
            df = pd.DataFrame(df_export)
            csv = df.to_csv(index=False, encoding='utf-8-sig')
            
            st.download_button(
                label="⬇️ 下载临床评估数据",
                data=csv,
                file_name=f"autism_clinical_assessment_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
                mime='text/csv'
            )
        
        # 对话记录下载
        if st.button("💬 下载行为观察记录 (TXT)", use_container_width=True):
            observation_content = []
            observation_content.append("=" * 70)
            observation_content.append("孤独症儿童临床行为观察记录")
            observation_content.append("基于DSM-5诊断标准 | 循证评估工具")
            observation_content.append("=" * 70)
            observation_content.append(f"生成时间: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
            observation_content.append(f"评估记录总数: {len(records)}")
            observation_content.append("=" * 70)
            observation_content.append("")
            
            for i, record in enumerate(records, 1):
                core_severity = (record['evaluation_scores']['社交互动质量'] + 
                               record['evaluation_scores']['沟通交流能力'] + 
                               record['evaluation_scores']['刻板重复行为']) / 3
                
                observation_content.append(f"\n【临床评估 {i}】")
                observation_content.append(f"评估ID: {record['experiment_id']}")
                observation_content.append(f"评估时间: {record['timestamp'].strftime('%Y-%m-%d %H:%M:%S')}")
                observation_content.append(f"严重程度分级: {record.get('template', '自定义')}")
                observation_content.append(f"评估情境: {record['scene']}")
                observation_content.append(f"观察活动: {record.get('activity', '未指定')}")
                observation_content.append(f"触发因素: {record.get('trigger', '未指定')}")
                
                if record.get('autism_profile'):
                    profile = record['autism_profile']
                    observation_content.append(f"DSM-5严重程度: {profile.get('dsm5_severity', '')}")
                    observation_content.append(f"所需支持水平: {profile.get('support_needs', '')}")
                
                observation_content.append(f"核心症状综合严重度: {core_severity:.2f}/5.0")
                observation_content.append("-" * 50)
                
                observation_content.append("临床评估得分:")
                for metric, score in record['evaluation_scores'].items():
                    observation_content.append(f"  • {metric}: {score}/5.0")
                
                if 'clinical_observations' in record and record['clinical_observations']:
                    observation_content.append("临床观察要点:")
                    for category, observations in record['clinical_observations'].items():
                        if observations:
                            observation_content.append(f"  {category}: {', '.join(observations)}")
                
                observation_content.append("行为观察对话:")
                observation_content.append(record['dialogue'])
                observation_content.append("-" * 50)
                observation_content.append("")
            
            observation_text = '\n'.join(observation_content)
            
            st.download_button(
                label="⬇️ 下载行为观察记录",
                data=observation_text.encode('utf-8'),
                file_name=f"autism_clinical_observations_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.txt",
                mime='text/plain'
            )
        
        # JSON完整数据
        if st.button("🔧 下载完整临床数据 (JSON)", use_container_width=True):
            json_data = {
                'clinical_assessment_report': {
                    'report_metadata': {
                        'generation_time': datetime.datetime.now().isoformat(),
                        'assessment_standard': 'DSM-5诊断标准',
                        'evaluation_tools': 'CARS, ABC, SCQ, M-CHAT参考',
                        'total_assessments': len(records),
                        'platform_version': '医学标准版 v1.0'
                    },
                    'assessment_summary': generate_clinical_analysis(records),
                    'detailed_assessments': []
                }
            }
            
            for record in records:
                clinical_record = record.copy()
                clinical_record['timestamp'] = record['timestamp'].isoformat()
                
                # 添加计算字段
                core_severity = (record['evaluation_scores']['社交互动质量'] + 
                               record['evaluation_scores']['沟通交流能力'] + 
                               record['evaluation_scores']['刻板重复行为']) / 3
                clinical_record['core_symptom_severity'] = round(core_severity, 2)
                
                json_data['clinical_assessment_report']['detailed_assessments'].append(clinical_record)
            
            json_str = json.dumps(json_data, ensure_ascii=False, indent=2)
            
            st.download_button(
                label="⬇️ 下载完整临床数据",
                data=json_str.encode('utf-8'),
                file_name=f"autism_clinical_complete_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.json",
                mime='application/json'
            )
    
    with col2:
        st.write("### 📈 专业分析报告")
        
        # 生成临床分析报告
        if st.button("📊 生成临床统计分析", use_container_width=True):
            with st.spinner("正在生成临床分析报告..."):
                analysis = generate_clinical_analysis(records)
            
            st.success("✅ 临床分析报告生成完成！")
            
            # 显示分析预览
            with st.expander("📋 临床分析报告预览", expanded=True):
                if analysis.get('临床评估概况'):
                    st.write("**临床评估概况:**")
                    for key, value in analysis['临床评估概况'].items():
                        st.write(f"- {key}: {value}")
                
                if analysis.get('整体临床表现'):
                    st.write("**整体临床表现:**")
                    for key, value in analysis['整体临床表现'].items():
                        st.write(f"- {key}: {value}")
                
                if analysis.get('临床发现与建议'):
                    st.write("**临床发现与建议:**")
                    for finding in analysis['临床发现与建议']:
                        st.write(f"- {finding}")
            
            # 提供分析报告下载
            analysis_json = json.dumps(analysis, ensure_ascii=False, indent=2)
            st.download_button(
                label="⬇️ 下载临床分析报告 (JSON)",
                data=analysis_json.encode('utf-8'),
                file_name=f"autism_clinical_analysis_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.json",
                mime='application/json'
            )
        
        # Excel专业报告
        if EXCEL_AVAILABLE:
            if st.button("📋 生成专业Excel报告", use_container_width=True):
                with st.spinner("正在生成专业Excel报告..."):
                    analysis = generate_clinical_analysis(records)
                    excel_data = create_clinical_excel_report(records, analysis)
                
                if excel_data:
                    st.success("✅ 专业Excel报告生成完成！")
                    
                    st.download_button(
                        label="⬇️ 下载专业Excel报告",
                        data=excel_data,
                        file_name=f"autism_clinical_professional_report_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                        mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                    )
                else:
                    st.error("❌ Excel报告生成失败，请尝试其他格式")
        else:
            st.info("💡 Excel报告功能需要安装openpyxl模块")
            st.code("pip install openpyxl")
            
            # 替代详细报告
            if st.button("📊 生成详细文本报告", use_container_width=True):
                with st.spinner("正在生成详细报告..."):
                    analysis = generate_clinical_analysis(records)
                
                # 创建详细文本报告
                detailed_report = []
                detailed_report.append("孤独症儿童临床评估详细报告")
                detailed_report.append("=" * 50)
                detailed_report.append(f"报告生成时间: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
                detailed_report.append(f"评估标准: DSM-5孤独症诊断标准")
                detailed_report.append(f"参考工具: CARS, ABC, SCQ, M-CHAT量表")
                detailed_report.append("")
                
                # 添加临床评估概况
                detailed_report.append("一、临床评估概况")
                detailed_report.append("-" * 30)
                for key, value in analysis.get('临床评估概况', {}).items():
                    detailed_report.append(f"{key}: {value}")
                detailed_report.append("")
                
                # 添加整体表现
                detailed_report.append("二、整体临床表现")
                detailed_report.append("-" * 30)
                for key, value in analysis.get('整体临床表现', {}).items():
                    detailed_report.append(f"{key}: {value}")
                detailed_report.append("")
                
                # 添加严重程度分析
                if analysis.get('严重程度分析'):
                    detailed_report.append("三、严重程度组间分析")
                    detailed_report.append("-" * 30)
                    for severity, stats in analysis['严重程度分析'].items():
                        detailed_report.append(f"\n{severity} (n={stats['评估次数']}):")
                        detailed_report.append(f"  - 社交互动缺陷: {stats['社交互动得分_均值']:.2f} ± {stats['社交互动得分_标准差']:.2f}")
                        detailed_report.append(f"  - 沟通交流缺陷: {stats['沟通交流得分_均值']:.2f} ± {stats['沟通交流得分_标准差']:.2f}")
                        detailed_report.append(f"  - 刻板重复行为: {stats['刻板行为得分_均值']:.2f} ± {stats['刻板行为得分_标准差']:.2f}")
                    detailed_report.append("")
                
                # 添加临床发现
                detailed_report.append("四、临床发现与建议")
                detailed_report.append("-" * 30)
                for i, finding in enumerate(analysis.get('临床发现与建议', []), 1):
                    detailed_report.append(f"{i}. {finding}")
                detailed_report.append("")
                
                # 添加个案明细
                detailed_report.append("五、个案评估明细")
                detailed_report.append("-" * 30)
                for i, record in enumerate(records, 1):
                    core_severity = (record['evaluation_scores']['社交互动质量'] + 
                                   record['evaluation_scores']['沟通交流能力'] + 
                                   record['evaluation_scores']['刻板重复行为']) / 3
                    
                    detailed_report.append(f"\n个案 {i}: {record['experiment_id']}")
                    detailed_report.append(f"  评估时间: {record['timestamp'].strftime('%Y-%m-%d %H:%M')}")
                    detailed_report.append(f"  严重程度: {record.get('template', '自定义')}")
                    detailed_report.append(f"  评估情境: {record['scene']}")
                    detailed_report.append(f"  核心症状综合: {core_severity:.2f}/5.0")
                    
                    severity_level = "轻度" if core_severity < 2.5 else "中度" if core_severity < 3.5 else "重度"
                    detailed_report.append(f"  严重程度判断: {severity_level}")
                
                report_content = '\n'.join(detailed_report)
                
                st.success("✅ 详细文本报告生成完成！")
                
                st.download_button(
                    label="⬇️ 下载详细文本报告",
                    data=report_content.encode('utf-8'),
                    file_name=f"autism_clinical_detailed_report_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.txt",
                    mime='text/plain'
                )
        
        # 研究数据包
        if st.button("📦 生成完整研究数据包", use_container_width=True, type="primary"):
            with st.spinner("正在生成完整研究数据包..."):
                # 生成分析
                analysis = generate_clinical_analysis(records)
                
                # 创建基础CSV
                df_basic = []
                for record in records:
                    core_severity = (record['evaluation_scores']['社交互动质量'] + 
                                   record['evaluation_scores']['沟通交流能力'] + 
                                   record['evaluation_scores']['刻板重复行为']) / 3
                    
                    row = {
                        '评估ID': record['experiment_id'],
                        '评估时间': record['timestamp'].strftime('%Y-%m-%d %H:%M:%S'),
                        '严重程度分级': record.get('template', '自定义'),
                        '评估情境': record['scene'],
                        '观察活动': record.get('activity', ''),
                        '触发因素': record.get('trigger', ''),
                        '社交互动缺陷': record['evaluation_scores']['社交互动质量'],
                        '沟通交流缺陷': record['evaluation_scores']['沟通交流能力'],
                        '刻板重复行为': record['evaluation_scores']['刻板重复行为'],
                        '感官处理异常': record['evaluation_scores']['感官处理能力'],
                        '情绪调节困难': record['evaluation_scores']['情绪行为调节'],
                        '认知适应缺陷': record['evaluation_scores']['认知适应功能'],
                        '核心症状综合严重度': round(core_severity, 2),
                        '备注': record.get('notes', '')
                    }
                    
                    # 添加DSM-5配置
                    if record.get('autism_profile'):
                        profile = record['autism_profile']
                        row.update({
                            'DSM5严重程度': profile.get('dsm5_severity', ''),
                            '所需支持水平': profile.get('support_needs', ''),
                            '社交沟通缺陷设置': profile.get('social_communication', ''),
                            '刻板重复行为设置': profile.get('restricted_repetitive', ''),
                            '感官处理设置': profile.get('sensory_processing', ''),
                            '认知功能设置': profile.get('cognitive_function', ''),
                            '适应行为设置': profile.get('adaptive_behavior', ''),
                            '语言发展设置': profile.get('language_level', ''),
                            '特殊兴趣': profile.get('special_interests', '')
                        })
                    
                    df_basic.append(row)
                
                # 创建完整数据JSON
                complete_data = {
                    'research_metadata': {
                        'generation_time': datetime.datetime.now().isoformat(),
                        'assessment_standard': 'DSM-5孤独症诊断标准',
                        'evaluation_framework': '基于CARS, ABC, SCQ, M-CHAT等权威量表',
                        'total_assessments': len(records),
                        'platform_version': '医学标准版 v1.0'
                    },
                    'statistical_analysis': analysis,
                    'raw_assessment_data': []
                }
                
                for record in records:
                    research_record = record.copy()
                    research_record['timestamp'] = record['timestamp'].isoformat()
                    complete_data['raw_assessment_data'].append(research_record)
            
            st.success("✅ 完整研究数据包生成完成！")
            
            # 创建ZIP文件
            zip_buffer = io.BytesIO()
            with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
                
                # 1. 基础评估数据
                csv_data = pd.DataFrame(df_basic).to_csv(index=False, encoding='utf-8-sig')
                zip_file.writestr("01_基础评估数据.csv", csv_data)
                
                # 2. Excel专业报告（如果可用）
                excel_data = None
                if EXCEL_AVAILABLE:
                    try:
                        excel_data = create_clinical_excel_report(records, analysis)
                        if excel_data:
                            zip_file.writestr("02_专业分析报告.xlsx", excel_data)
                    except Exception as e:
                        pass  # 如果Excel生成失败，跳过
                
                if not excel_data:
                    # 生成详细文本报告作为替代
                    detailed_report = []
                    detailed_report.append("孤独症儿童临床评估专业报告")
                    detailed_report.append("基于DSM-5诊断标准")
                    detailed_report.append("=" * 60)
                    detailed_report.append(f"报告生成时间: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
                    detailed_report.append(f"评估标准: DSM-5孤独症谱系障碍诊断标准")
                    detailed_report.append(f"参考量表: CARS, ABC, SCQ, M-CHAT")
                    detailed_report.append("")
                    
                    # 临床概况
                    detailed_report.append("一、临床评估概况")
                    detailed_report.append("-" * 40)
                    for key, value in analysis.get('临床评估概况', {}).items():
                        detailed_report.append(f"{key}: {value}")
                    
                    # 整体表现
                    detailed_report.append("\n二、整体临床表现")
                    detailed_report.append("-" * 40)
                    for key, value in analysis.get('整体临床表现', {}).items():
                        detailed_report.append(f"{key}: {value}")
                    
                    # 临床发现
                    detailed_report.append("\n三、临床发现与建议")
                    detailed_report.append("-" * 40)
                    for i, finding in enumerate(analysis.get('临床发现与建议', []), 1):
                        detailed_report.append(f"{i}. {finding}")
                    
                    report_text = '\n'.join(detailed_report)
                    zip_file.writestr("02_专业分析报告.txt", report_text.encode('utf-8'))
                
                # 3. 完整研究数据
                json_str = json.dumps(complete_data, ensure_ascii=False, indent=2)
                zip_file.writestr("03_完整研究数据.json", json_str.encode('utf-8'))
                
                # 4. 行为观察记录
                observation_content = []
                observation_content.append("孤独症儿童行为观察记录集")
                observation_content.append("=" * 50)
                observation_content.append(f"观察标准: DSM-5孤独症诊断标准")
                observation_content.append(f"记录时间: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
                observation_content.append(f"观察案例数: {len(records)}")
                observation_content.append("")
                
                for i, record in enumerate(records, 1):
                    observation_content.append(f"\n【观察记录 {i}】")
                    observation_content.append(f"评估ID: {record['experiment_id']}")
                    observation_content.append(f"时间: {record['timestamp'].strftime('%Y-%m-%d %H:%M:%S')}")
                    observation_content.append(f"严重程度: {record.get('template', '自定义')}")
                    observation_content.append(f"情境: {record['scene']}")
                    observation_content.append(f"活动: {record.get('activity', '')}")
                    observation_content.append(f"触发因素: {record.get('trigger', '')}")
                    observation_content.append("-" * 30)
                    observation_content.append("行为观察对话:")
                    observation_content.append(record['dialogue'])
                    observation_content.append("-" * 30)
                    observation_content.append("")
                
                obs_text = '\n'.join(observation_content)
                zip_file.writestr("04_行为观察记录.txt", obs_text.encode('utf-8'))
                
                # 5. 研究说明文档
                readme_content = f"""
孤独症儿童AI模拟实验平台 - 医学标准版
研究数据包说明文档
======================================

生成时间: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
评估记录数: {len(records)}
评估标准: DSM-5孤独症谱系障碍诊断标准

文件说明:
--------
1. 01_基础评估数据.csv
   - 包含所有评估的核心数据
   - 适用于统计分析和数据可视化
   - 包含DSM-5特征配置和评估得分

"""
                if excel_data:
                    readme_content += """2. 02_专业分析报告.xlsx
   - 多工作表Excel专业报告
   - 包含统计分析、图表和临床解释
   - 适用于临床报告和学术研究
"""
                else:
                    readme_content += """2. 02_专业分析报告.txt
   - 详细的文本格式专业报告
   - 包含统计分析和临床解释
   - 注意: Excel功能不可用，如需Excel格式请安装openpyxl
"""

                readme_content += """
3. 03_完整研究数据.json
   - 包含所有原始数据和元数据
   - 适用于程序处理和深度分析
   - 包含完整的临床观察记录

4. 04_行为观察记录.txt
   - 所有评估的对话记录
   - 用于质性分析和行为模式研究
   - 符合临床观察记录格式

5. README.txt
   - 本说明文档

评估指标说明:
-----------
所有评估得分采用1-5分制，其中:
- 1分: 无明显问题/正常范围
- 2分: 轻度问题/需要关注
- 3分: 中度问题/需要支持
- 4分: 明显问题/需要大量支持
- 5分: 严重问题/需要非常大量支持

核心症状评估:
- 社交互动质量: 基于DSM-5诊断标准A条目
- 沟通交流能力: 基于DSM-5诊断标准A条目
- 刻板重复行为: 基于DSM-5诊断标准B条目

相关功能评估:
- 感官处理能力: 感官异常和感官寻求/逃避行为
- 情绪行为调节: 情绪识别、表达和调节能力
- 认知适应功能: 学习能力和适应性行为

使用建议:
--------
1. 临床应用:
   - 使用基础数据进行筛查和评估
   - 参考专业报告制定干预计划
   - 结合行为观察进行个案分析

2. 研究应用:
   - 使用完整数据进行统计分析
   - 对照组研究和纵向研究
   - 干预效果评估

3. 教学应用:
   - 案例教学和临床培训
   - 评估工具使用培训
   - 行为观察技能训练

技术支持:
--------
- 如需Excel功能，请安装: pip install openpyxl
- 数据分析建议使用: pandas, numpy, scipy
- 可视化建议使用: matplotlib, plotly

参考标准:
--------
- American Psychiatric Association. (2013). DSM-5
- Childhood Autism Rating Scale (CARS)
- Autism Behavior Checklist (ABC)
- Social Communication Questionnaire (SCQ)
- Modified Checklist for Autism in Toddlers (M-CHAT)

质量保证:
--------
本平台基于最新的DSM-5诊断标准和权威评估工具设计，
所有评估指标均参考国际认可的孤独症评估量表，
确保评估结果的专业性和可靠性。

版权声明:
--------
本数据包仅供学术研究和临床实践使用，
请遵循相关伦理规范和数据保护法规。
"""
                zip_file.writestr("README.txt", readme_content)
            
            zip_buffer.seek(0)
            
            st.download_button(
                label="⬇️ 下载完整研究数据包 (ZIP)",
                data=zip_buffer.getvalue(),
                file_name=f"autism_clinical_research_package_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.zip",
                mime='application/zip'
            )
    
    # 数据统计概览
    st.subheader("📈 数据统计概览")
    
    col_stat1, col_stat2, col_stat3, col_stat4 = st.columns(4)
    
    with col_stat1:
        st.metric("临床评估总数", len(records))
    
    with col_stat2:
        severities = [r.get('template', '自定义') for r in records]
        unique_severities = len(set(severities))
        st.metric("涉及严重程度类型", unique_severities)
    
    with col_stat3:
        contexts = [r['scene'] for r in records]
        unique_contexts = len(set(contexts))
        st.metric("涉及评估情境", unique_contexts)
    
    with col_stat4:
        if records:
            time_span = (max(r['timestamp'] for r in records) - min(r['timestamp'] for r in records)).days
            st.metric("评估时间跨度(天)", time_span)
    
    # 临床数据预览
    st.subheader("📊 临床数据预览")
    
    preview_data = []
    for record in records[:10]:
        # 计算核心症状严重度
        core_severity = (record['evaluation_scores']['社交互动质量'] + 
                        record['evaluation_scores']['沟通交流能力'] + 
                        record['evaluation_scores']['刻板重复行为']) / 3
        
        severity_level = "轻度" if core_severity < 2.5 else "中度" if core_severity < 3.5 else "重度"
        
        preview_data.append({
            '评估ID': record['experiment_id'][:25] + '...' if len(record['experiment_id']) > 25 else record['experiment_id'],
            '时间': record['timestamp'].strftime('%m-%d %H:%M'),
            '严重程度': record.get('template', '自定义')[:8] + '...' if len(record.get('template', '自定义')) > 8 else record.get('template', '自定义'),
            '评估情境': record['scene'].replace('结构化', '结构'),
            '核心症状': f"{core_severity:.2f}",
            '程度判断': severity_level
        })
    
    df_preview = pd.DataFrame(preview_data)
    st.dataframe(df_preview, use_container_width=True)
    
    if len(records) > 10:
        st.info(f"显示前10条记录，共{len(records)}条。完整数据请通过上方下载功能获取。")

# 侧边栏额外信息
st.sidebar.markdown("---")
st.sidebar.markdown("### 📈 评估统计")
if st.session_state.experiment_records:
    st.sidebar.metric("评估总数", len(st.session_state.experiment_records))
    recent_record = st.session_state.experiment_records[-1]
    st.sidebar.write(f"最近评估: {recent_record['timestamp'].strftime('%m-%d %H:%M')}")
    
    # 显示严重程度分布
    severities = [r.get('template', '自定义') for r in st.session_state.experiment_records]
    severity_counts = pd.Series(severities).value_counts()
    st.sidebar.write("**严重程度分布:**")
    for severity, count in severity_counts.items():
        short_name = severity.split('（')[0] if '（' in severity else severity
        st.sidebar.write(f"- {short_name}: {count}")
    
    # 核心症状统计
    all_core_scores = []
    for r in st.session_state.experiment_records:
        core_score = (r['evaluation_scores']['社交互动质量'] + 
                     r['evaluation_scores']['沟通交流能力'] + 
                     r['evaluation_scores']['刻板重复行为']) / 3
        all_core_scores.append(core_score)
    
    avg_core = np.mean(all_core_scores)
    st.sidebar.metric("平均核心症状严重度", f"{avg_core:.2f}/5")
    
    if avg_core >= 4.0:
        st.sidebar.error("整体评估显示重度症状")
    elif avg_core >= 3.0:
        st.sidebar.warning("整体评估显示中度症状")
    else:
        st.sidebar.success("整体评估显示轻度症状")
        
else:
    st.sidebar.write("暂无评估数据")

st.sidebar.markdown("---")
st.sidebar.markdown("### ℹ️ 医学标准说明")
st.sidebar.markdown("""
**评估标准**: DSM-5孤独症谱系障碍诊断标准

**核心症状**:
- A. 社交沟通缺陷
- B. 刻板重复行为模式

**参考量表**:
- CARS: 儿童孤独症评定量表
- ABC: 孤独症行为量表  
- SCQ: 社交沟通问卷
- M-CHAT: 修订版孤独症筛查量表

**严重程度分级**:
1. 需要支持（轻度）
2. 需要大量支持（中度）
3. 需要非常大量支持（重度）
""")

if not EXCEL_AVAILABLE:
    st.sidebar.markdown("---")
    st.sidebar.markdown("### 💡 功能提示")
    st.sidebar.warning("Excel功能不可用")
    st.sidebar.markdown("要启用专业Excel报告功能，请运行：")
    st.sidebar.code("pip install openpyxl")
    st.sidebar.markdown("目前可使用CSV和JSON格式导出临床数据。")

st.sidebar.markdown("---")
st.sidebar.markdown("### ⚠️ API限制说明")
st.sidebar.markdown("""
**当前API限制**: 每分钟3次请求

**对评估影响**:
- 临床快速评估: 立即完成
- 批量临床研究: 每个评估间隔25秒

**建议**:
- 批量研究选择适当规模
- 可分批次进行大样本研究
- 保持网络连接稳定
""")

st.sidebar.markdown("---")
st.sidebar.markdown("### 🏥 临床应用建议")
st.sidebar.markdown("""
**筛查应用**:
- 快速识别可能的孤独症特征
- 辅助临床诊断决策

**干预规划**:
- 基于评估结果制定个体化干预
- 监测干预效果

**研究用途**:
- 症状严重度量化
- 干预前后对比
- 群体特征分析

**注意事项**:
- 本工具仅供辅助参考
- 不能替代专业临床诊断
- 建议结合其他评估工具
""")

st.sidebar.markdown("---")
st.sidebar.markdown("### 🔄 数据管理")
if st.sidebar.button("🗑️ 清空所有评估数据"):
    if st.sidebar.checkbox("确认清空（此操作不可恢复）"):
        st.session_state.experiment_records = []
        st.sidebar.success("✅ 评估数据已清空")
        st.rerun()

# 页脚
st.markdown("---")
st.markdown("""
### 📋 平台特点

**🏥 医学标准**: 严格遵循DSM-5诊断标准，参考权威评估量表  
**🔬 科学评估**: 基于循证实践的评估指标和评分标准  
**📊 专业报告**: 生成符合临床要求的专业评估报告  
**🎯 个体化**: 支持个性化配置和定制化评估设计

**💡 使用提示**: 
- 建议先进行'临床快速评估'熟悉平台功能
- 使用'批量临床研究'获取统计学有效的数据
- 在'📊 临床报告中心'下载完整的专业报告
- 所有评估结果仅供临床参考，不能替代专业诊断

**⚠️ 重要声明**: 本平台仅供学术研究和临床辅助使用，不能替代专业医师的临床诊断。
""")

st.markdown("*基于DSM-5标准 | 参考CARS、ABC、SCQ、M-CHAT等权威量表 | 医学标准版 v1.0*")