"""孤独症平台报告中心页面"""
import streamlit as st
import pandas as pd
import datetime
import json

from common.config import EXCEL_AVAILABLE
from common.exporters import (
    export_to_csv, export_to_json, export_to_excel, 
    create_excel_workbook, apply_excel_styles, create_zip_package
)
from common.exporters.text_exporter import export_to_text
from .analyzer import generate_clinical_analysis, prepare_clinical_export_data


def page_report_center():
    """临床报告中心页面"""
    st.header("📊 临床报告中心")
    st.markdown("基于循证实践生成专业的临床评估报告和研究数据")
    
    records = st.session_state.experiment_records
    
    if not records:
        st.warning("📊 暂无评估数据，请先进行临床评估")
        st.stop()
    
    st.success(f"📊 当前共有 {len(records)} 条临床评估记录可生成报告")
    
    # 创建两个标签页
    tab1, tab2 = st.tabs(["📊 生成报告", "📥 导入数据"])

    with tab1:
        # 报告类型选择
        st.subheader("📋 选择报告类型")
        
        col1, col2 = st.columns(2)
        
        with col1:
            st.write("### 📄 标准临床报告")
            
            # 基础CSV报告
            if st.button("📊 下载基础评估数据 (CSV)", use_container_width=True):
                export_data = prepare_clinical_export_data(records)
                csv_content = export_to_csv(export_data)
                
                st.download_button(
                    label="⬇️ 下载临床评估数据",
                    data=csv_content,
                    file_name=f"autism_clinical_assessment_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
                    mime='text/csv'
                )
            
            # 对话记录下载
            if st.button("💬 下载行为观察记录 (TXT)", use_container_width=True):
                observation_content = create_observation_text(records)
                
                st.download_button(
                    label="⬇️ 下载行为观察记录",
                    data=export_to_text(observation_content),
                    file_name=f"autism_clinical_observations_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.txt",
                    mime='text/plain'
                )
            
            # JSON完整数据
            if st.button("🔧 下载完整临床数据 (JSON)", use_container_width=True):
                json_data = create_complete_json_data(records)
                json_str = export_to_json(json_data)
                
                st.download_button(
                    label="⬇️ 下载完整临床数据",
                    data=json_str.encode('utf-8'),
                    file_name=f"autism_clinical_complete_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.json",
                    mime='application/json'
                )
        
        with col2:
            st.write("### 📈 专业分析报告")
            
            # 生成临床分析报告
            if st.button("📊 生成临床统计分析", use_container_width=True):
                with st.spinner("正在生成临床分析报告..."):
                    analysis = generate_clinical_analysis(records)
                
                st.success("✅ 临床分析报告生成完成！")
                
                # 显示分析预览
                with st.expander("📋 临床分析报告预览", expanded=True):
                    if analysis.get('临床评估概况'):
                        st.write("**临床评估概况:**")
                        for key, value in analysis['临床评估概况'].items():
                            st.write(f"- {key}: {value}")
                    
                    if analysis.get('整体临床表现'):
                        st.write("**整体临床表现:**")
                        for key, value in analysis['整体临床表现'].items():
                            st.write(f"- {key}: {value}")
                    
                    if analysis.get('临床发现与建议'):
                        st.write("**临床发现与建议:**")
                        for finding in analysis['临床发现与建议']:
                            st.write(f"- {finding}")
                
                # 提供分析报告下载
                analysis_json = export_to_json(analysis)
                st.download_button(
                    label="⬇️ 下载临床分析报告 (JSON)",
                    data=analysis_json.encode('utf-8'),
                    file_name=f"autism_clinical_analysis_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.json",
                    mime='application/json'
                )
            
            # Excel专业报告
            if EXCEL_AVAILABLE:
                if st.button("📋 生成专业Excel报告", use_container_width=True):
                    with st.spinner("正在生成专业Excel报告..."):
                        analysis = generate_clinical_analysis(records)
                        excel_data = create_clinical_excel_report(records, analysis)
                    
                    if excel_data:
                        st.success("✅ 专业Excel报告生成完成！")
                        
                        st.download_button(
                            label="⬇️ 下载专业Excel报告",
                            data=excel_data,
                            file_name=f"autism_clinical_professional_report_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                            mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                        )
                    else:
                        st.error("❌ Excel报告生成失败，请尝试其他格式")
            else:
                st.info("💡 Excel报告功能需要安装openpyxl模块")
                st.code("pip install openpyxl")
                
                # 替代详细报告
                if st.button("📊 生成详细文本报告", use_container_width=True):
                    with st.spinner("正在生成详细报告..."):
                        analysis = generate_clinical_analysis(records)
                    
                    # 创建详细文本报告
                    detailed_report = create_detailed_text_report(records, analysis)
                    
                    st.success("✅ 详细文本报告生成完成！")
                    
                    st.download_button(
                        label="⬇️ 下载详细文本报告",
                        data=export_to_text(detailed_report),
                        file_name=f"autism_clinical_detailed_report_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.txt",
                        mime='text/plain'
                    )
            
            # 研究数据包
            if st.button("📦 生成完整研究数据包", use_container_width=True, type="primary"):
                with st.spinner("正在生成完整研究数据包..."):
                    zip_data = create_research_package(records)
                
                st.success("✅ 完整研究数据包生成完成！")
                
                st.download_button(
                    label="⬇️ 下载完整研究数据包 (ZIP)",
                    data=zip_data,
                    file_name=f"autism_clinical_research_package_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.zip",
                    mime='application/zip'
                )
        
        # 数据统计概览
        display_data_overview(records)
        
        # 数据预览
        display_data_preview(records)

    with tab2:
        # 添加简化的导入功能
        st.subheader("📥 快速导入历史数据")
        st.markdown("从文件导入历史儿童发展观察数据到系统中")
        
        uploaded_file = st.file_uploader(
            "选择要导入的文件",
            type=['csv', 'json', 'xlsx', 'xls', 'zip'],
            help="支持从导出的报告文件重新导入数据"
        )
        
        if uploaded_file is not None:
            col1, col2 = st.columns([3, 1])
            
            with col1:
                st.info(f"📄 已选择文件: {uploaded_file.name} ({uploaded_file.size // 1024} KB)")
            
            with col2:
                if st.button("🚀 导入", type="primary"):
                    try:
                        from common.importers import get_importer
                        from common.data_storage_manager import DataStorageManager
                        
                        # 获取导入器
                        file_ext = uploaded_file.name.split('.')[-1].lower()
                        importer_class = get_importer(file_ext)
                        importer = importer_class(assessment_type='children')
                        
                        # 导入数据
                        with st.spinner("正在导入数据..."):
                            file_content = uploaded_file.read()
                            result = importer.import_data(file_content)
                            
                            if result.status.value == 'success':
                                # 合并数据
                                storage_manager = DataStorageManager('children')
                                merged_count, _ = storage_manager.merge_imported_data(
                                    result.records, 
                                    merge_strategy='skip_duplicates'
                                )
                                
                                st.success(f"✅ 成功导入 {merged_count} 条观察记录")
                                st.rerun()
                            else:
                                st.error(f"导入失败: {result.errors[0]['message'] if result.errors else '未知错误'}")
                                
                    except Exception as e:
                        st.error(f"导入出错: {str(e)}")
        
        # 导入说明
        with st.expander("ℹ️ 导入说明"):
            st.markdown("""
            - 支持导入之前导出的所有格式文件
            - CSV/Excel文件会自动识别数据格式
            - JSON文件保留完整的发展观察数据结构
            - ZIP包可以包含多个文件批量导入
            - 系统会自动跳过重复的记录
            
            **提示**: 如需更多导入选项，请访问"数据导入"页面
            """)
        
        # 添加跳转到完整导入页面的链接
        st.markdown("---")
        if st.button("🔧 前往完整数据导入页面"):
            st.switch_page("pages/data_import_page.py")



def create_observation_text(records):
    """创建观察记录文本"""
    observation_content = []
    observation_content.append("=" * 70)
    observation_content.append("孤独症儿童临床行为观察记录")
    observation_content.append("基于DSM-5诊断标准 | 循证评估工具")
    observation_content.append("=" * 70)
    observation_content.append(f"生成时间: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    observation_content.append(f"评估记录总数: {len(records)}")
    observation_content.append("=" * 70)
    observation_content.append("")
    
    for i, record in enumerate(records, 1):
        core_severity = (record['evaluation_scores']['社交互动质量'] + 
                       record['evaluation_scores']['沟通交流能力'] + 
                       record['evaluation_scores']['刻板重复行为']) / 3
        
        observation_content.append(f"\n【临床评估 {i}】")
        observation_content.append(f"评估ID: {record['experiment_id']}")
        observation_content.append(f"评估时间: {record['timestamp'].strftime('%Y-%m-%d %H:%M:%S')}")
        observation_content.append(f"严重程度分级: {record.get('template', '自定义')}")
        observation_content.append(f"评估情境: {record['scene']}")
        observation_content.append(f"观察活动: {record.get('activity', '未指定')}")
        observation_content.append(f"触发因素: {record.get('trigger', '未指定')}")
        
        if record.get('autism_profile'):
            profile = record['autism_profile']
            observation_content.append(f"DSM-5严重程度: {profile.get('dsm5_severity', '')}")
            observation_content.append(f"所需支持水平: {profile.get('support_needs', '')}")
        
        observation_content.append(f"核心症状综合严重度: {core_severity:.2f}/5.0")
        observation_content.append("-" * 50)
        
        observation_content.append("临床评估得分:")
        for metric, score in record['evaluation_scores'].items():
            observation_content.append(f"  • {metric}: {score}/5.0")
        
        if 'clinical_observations' in record and record['clinical_observations']:
            observation_content.append("临床观察要点:")
            for category, observations in record['clinical_observations'].items():
                if observations:
                    observation_content.append(f"  {category}: {', '.join(observations)}")
        
        observation_content.append("行为观察对话:")
        observation_content.append(record['dialogue'])
        observation_content.append("-" * 50)
        observation_content.append("")
    
    return observation_content


def create_complete_json_data(records):
    """创建完整的JSON数据"""
    analysis = generate_clinical_analysis(records)
    
    json_data = {
        'clinical_assessment_report': {
            'report_metadata': {
                'generation_time': datetime.datetime.now().isoformat(),
                'assessment_standard': 'DSM-5诊断标准',
                'evaluation_tools': 'CARS, ABC, SCQ, M-CHAT参考',
                'total_assessments': len(records),
                'platform_version': '医学标准版 v1.0'
            },
            'assessment_summary': analysis,
            'detailed_assessments': []
        }
    }
    
    for record in records:
        clinical_record = record.copy()
        clinical_record['timestamp'] = record['timestamp'].isoformat()
        
        # 添加计算字段
        core_severity = (record['evaluation_scores']['社交互动质量'] + 
                       record['evaluation_scores']['沟通交流能力'] + 
                       record['evaluation_scores']['刻板重复行为']) / 3
        clinical_record['core_symptom_severity'] = round(core_severity, 2)
        
        json_data['clinical_assessment_report']['detailed_assessments'].append(clinical_record)
    
    return json_data


def create_clinical_excel_report(records, analysis):
    """创建临床标准的Excel报告"""
    if not EXCEL_AVAILABLE:
        return None
    
    workbook = create_excel_workbook()
    
    # 1. 临床评估概览
    overview_sheet = workbook.create_sheet("临床评估概览")
    overview_sheet.append(["孤独症儿童临床行为评估报告（基于DSM-5标准）"])
    overview_sheet.append([])
    overview_sheet.append(["报告生成时间", datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')])
    overview_sheet.append(["评估标准", "DSM-5孤独症诊断标准 + CARS/ABC/SCQ量表"])
    overview_sheet.append([])
    
    overview_sheet.append(["评估概况"])
    for key, value in analysis.get('临床评估概况', {}).items():
        overview_sheet.append([key, value])
    
    overview_sheet.append([])
    overview_sheet.append(["整体临床表现"])
    for key, value in analysis.get('整体临床表现', {}).items():
        overview_sheet.append([key, value])
    
    overview_sheet.append([])
    overview_sheet.append(["临床发现与建议"])
    for finding in analysis.get('临床发现与建议', []):
        overview_sheet.append([finding])
    
    # 2. 详细评估数据
    data_sheet = workbook.create_sheet("详细评估数据")
    headers = ["评估ID", "时间", "严重程度", "评估情境", "观察活动", "触发因素",
              "社交互动缺陷", "沟通交流缺陷", "刻板重复行为", "感官处理异常", 
              "情绪调节困难", "认知适应缺陷", "核心症状严重度",
              "DSM-5严重程度", "所需支持水平", "特殊兴趣", "备注"]
    data_sheet.append(headers)
    
    for record in records:
        profile = record.get('autism_profile', {})
        scores = record['evaluation_scores']
        core_symptom_severity = (scores['社交互动质量'] + scores['沟通交流能力'] + scores['刻板重复行为']) / 3
        
        row = [
            record['experiment_id'],
            record['timestamp'].strftime('%Y-%m-%d %H:%M:%S'),
            record.get('template', '自定义'),
            record['scene'],
            record.get('activity', ''),
            record.get('trigger', ''),
            scores['社交互动质量'],
            scores['沟通交流能力'],
            scores['刻板重复行为'],
            scores['感官处理能力'],
            scores['情绪行为调节'],
            scores['认知适应功能'],
            f"{core_symptom_severity:.2f}",
            profile.get('dsm5_severity', ''),
            profile.get('support_needs', ''),
            profile.get('special_interests', ''),
            record.get('notes', '')
        ]
        data_sheet.append(row)
    
    # 3. 严重程度对比分析
    if analysis.get('严重程度分析'):
        severity_sheet = workbook.create_sheet("严重程度分析")
        severity_headers = ["严重程度", "评估次数", "社交缺陷均值", "沟通缺陷均值", 
                          "刻板行为均值", "感官异常均值", "情绪困难均值", "认知缺陷均值",
                          "核心症状综合"]
        severity_sheet.append(severity_headers)
        
        for severity, stats in analysis['严重程度分析'].items():
            core_avg = (stats['社交互动得分_均值'] + stats['沟通交流得分_均值'] + stats['刻板行为得分_均值']) / 3
            row = [
                severity,
                stats['评估次数'],
                f"{stats['社交互动得分_均值']:.2f}",
                f"{stats['沟通交流得分_均值']:.2f}",
                f"{stats['刻板行为得分_均值']:.2f}",
                f"{stats['感官处理得分_均值']:.2f}",
                f"{stats['情绪调节得分_均值']:.2f}",
                f"{stats['认知适应得分_均值']:.2f}",
                f"{core_avg:.2f}"
            ]
            severity_sheet.append(row)
    
    # 4. 临床观察记录
    if any('clinical_observations' in record for record in records):
        obs_sheet = workbook.create_sheet("临床观察记录")
        obs_sheet.append(["评估ID", "社交行为观察", "语言沟通特点", "重复行为表现", "感官反应", "情绪调节"])
        
        for record in records:
            if 'clinical_observations' in record:
                obs = record['clinical_observations']
                row = [
                    record['experiment_id'],
                    '; '.join(obs.get('社交行为观察', [])),
                    '; '.join(obs.get('语言沟通特点', [])),
                    '; '.join(obs.get('重复行为表现', [])),
                    '; '.join(obs.get('感官反应', [])),
                    '; '.join(obs.get('情绪调节', []))
                ]
                obs_sheet.append(row)
    
    # 5. 对话记录（用于质性分析）
    dialogue_sheet = workbook.create_sheet("对话记录")
    dialogue_sheet.append(["评估ID", "严重程度", "评估情境", "对话内容"])
    
    for record in records:
        dialogue_sheet.append([
            record['experiment_id'],
            record.get('template', '自定义'),
            record['scene'],
            record['dialogue']
        ])
    
    # 应用专业样式
    apply_excel_styles(workbook, header_color="366092", header_font_color="FFFFFF")
    
    # 特殊样式处理
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value and any(keyword in str(cell.value) for keyword in ['严重', '明显', '需要支持']):
                    from openpyxl.styles import PatternFill
                    cell.fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
    
    return export_to_excel(workbook)


def create_detailed_text_report(records, analysis):
    """创建详细文本报告"""
    detailed_report = []
    detailed_report.append("孤独症儿童临床评估详细报告")
    detailed_report.append("=" * 50)
    detailed_report.append(f"报告生成时间: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    detailed_report.append(f"评估标准: DSM-5孤独症诊断标准")
    detailed_report.append(f"参考工具: CARS, ABC, SCQ, M-CHAT量表")
    detailed_report.append("")
    
    # 添加临床评估概况
    detailed_report.append("一、临床评估概况")
    detailed_report.append("-" * 30)
    for key, value in analysis.get('临床评估概况', {}).items():
        detailed_report.append(f"{key}: {value}")
    detailed_report.append("")
    
    # 添加整体表现
    detailed_report.append("二、整体临床表现")
    detailed_report.append("-" * 30)
    for key, value in analysis.get('整体临床表现', {}).items():
        detailed_report.append(f"{key}: {value}")
    detailed_report.append("")
    
    # 添加严重程度分析
    if analysis.get('严重程度分析'):
        detailed_report.append("三、严重程度组间分析")
        detailed_report.append("-" * 30)
        for severity, stats in analysis['严重程度分析'].items():
            detailed_report.append(f"\n{severity} (n={stats['评估次数']}):")
            detailed_report.append(f"  - 社交互动缺陷: {stats['社交互动得分_均值']:.2f} ± {stats['社交互动得分_标准差']:.2f}")
            detailed_report.append(f"  - 沟通交流缺陷: {stats['沟通交流得分_均值']:.2f} ± {stats['沟通交流得分_标准差']:.2f}")
            detailed_report.append(f"  - 刻板重复行为: {stats['刻板行为得分_均值']:.2f} ± {stats['刻板行为得分_标准差']:.2f}")
        detailed_report.append("")
    
    # 添加临床发现
    detailed_report.append("四、临床发现与建议")
    detailed_report.append("-" * 30)
    for i, finding in enumerate(analysis.get('临床发现与建议', []), 1):
        detailed_report.append(f"{i}. {finding}")
    detailed_report.append("")
    
    # 添加个案明细
    detailed_report.append("五、个案评估明细")
    detailed_report.append("-" * 30)
    for i, record in enumerate(records, 1):
        core_severity = (record['evaluation_scores']['社交互动质量'] + 
                       record['evaluation_scores']['沟通交流能力'] + 
                       record['evaluation_scores']['刻板重复行为']) / 3
        
        detailed_report.append(f"\n个案 {i}: {record['experiment_id']}")
        detailed_report.append(f"  评估时间: {record['timestamp'].strftime('%Y-%m-%d %H:%M')}")
        detailed_report.append(f"  严重程度: {record.get('template', '自定义')}")
        detailed_report.append(f"  评估情境: {record['scene']}")
        detailed_report.append(f"  核心症状综合: {core_severity:.2f}/5.0")
        
        severity_level = "轻度" if core_severity < 2.5 else "中度" if core_severity < 3.5 else "重度"
        detailed_report.append(f"  严重程度判断: {severity_level}")
    
    return detailed_report


def create_research_package(records):
    """创建研究数据包"""
    # 生成分析
    analysis = generate_clinical_analysis(records)
    
    # 准备各种格式的数据
    files_dict = {}
    
    # 1. 基础发展数据CSV
    export_data = prepare_clinical_export_data(records)
    files_dict["01_基础评估数据.csv"] = export_to_csv(export_data)
    
    # 2. 专业分析报告
    if EXCEL_AVAILABLE:
        excel_data = create_clinical_excel_report(records, analysis)
        if excel_data:
            files_dict["02_专业分析报告.xlsx"] = excel_data
    
    if "02_专业分析报告.xlsx" not in files_dict:
        # Excel不可用时的文本报告
        detailed_report = create_detailed_text_report(records, analysis)
        files_dict["02_专业分析报告.txt"] = '\n'.join(detailed_report)
    
    # 3. 完整研究数据JSON
    complete_data = create_complete_json_data(records)
    files_dict["03_完整研究数据.json"] = export_to_json(complete_data)
    
    # 4. 行为观察记录
    observation_content = create_observation_text(records)
    files_dict["04_行为观察记录.txt"] = '\n'.join(observation_content)
    
    # 5. README文件
    readme_content = create_readme_content(records, EXCEL_AVAILABLE)
    files_dict["README.txt"] = readme_content
    
    return create_zip_package(files_dict)


def create_readme_content(records, excel_available):
    """创建README内容"""
    readme_content = f"""
孤独症儿童AI模拟实验平台 - 医学标准版
研究数据包说明文档
======================================

生成时间: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
评估记录数: {len(records)}
评估标准: DSM-5孤独症谱系障碍诊断标准

文件说明:
--------
1. 01_基础评估数据.csv
   - 包含所有评估的核心数据
   - 适用于统计分析和数据可视化
   - 包含DSM-5特征配置和评估得分

"""
    if excel_available:
        readme_content += """2. 02_专业分析报告.xlsx
   - 多工作表Excel专业报告
   - 包含统计分析、图表和临床解释
   - 适用于临床报告和学术研究
"""
    else:
        readme_content += """2. 02_专业分析报告.txt
   - 详细的文本格式专业报告
   - 包含统计分析和临床解释
   - 注意: Excel功能不可用，如需Excel格式请安装openpyxl
"""

    readme_content += """
3. 03_完整研究数据.json
   - 包含所有原始数据和元数据
   - 适用于程序处理和深度分析
   - 包含完整的临床观察记录

4. 04_行为观察记录.txt
   - 所有评估的对话记录
   - 用于质性分析和行为模式研究
   - 符合临床观察记录格式

5. README.txt
   - 本说明文档

评估指标说明:
-----------
所有评估得分采用1-5分制，其中:
- 1分: 无明显问题/正常范围
- 2分: 轻度问题/需要关注
- 3分: 中度问题/需要支持
- 4分: 明显问题/需要大量支持
- 5分: 严重问题/需要非常大量支持

核心症状评估:
- 社交互动质量: 基于DSM-5诊断标准A条目
- 沟通交流能力: 基于DSM-5诊断标准A条目
- 刻板重复行为: 基于DSM-5诊断标准B条目

相关功能评估:
- 感官处理能力: 感官异常和感官寻求/逃避行为
- 情绪行为调节: 情绪识别、表达和调节能力
- 认知适应功能: 学习能力和适应性行为

使用建议:
--------
1. 临床应用:
   - 使用基础数据进行筛查和评估
   - 参考专业报告制定干预计划
   - 结合行为观察进行个案分析

2. 研究应用:
   - 使用完整数据进行统计分析
   - 对照组研究和纵向研究
   - 干预效果评估

3. 教学应用:
   - 案例教学和临床培训
   - 评估工具使用培训
   - 行为观察技能训练

技术支持:
--------
- 如需Excel功能，请安装: pip install openpyxl
- 数据分析建议使用: pandas, numpy, scipy
- 可视化建议使用: matplotlib, plotly

参考标准:
--------
- American Psychiatric Association. (2013). DSM-5
- Childhood Autism Rating Scale (CARS)
- Autism Behavior Checklist (ABC)
- Social Communication Questionnaire (SCQ)
- Modified Checklist for Autism in Toddlers (M-CHAT)

质量保证:
--------
本平台基于最新的DSM-5诊断标准和权威评估工具设计，
所有评估指标均参考国际认可的孤独症评估量表，
确保评估结果的专业性和可靠性。

版权声明:
--------
本数据包仅供学术研究和临床实践使用，
请遵循相关伦理规范和数据保护法规。
"""
    return readme_content


def display_data_overview(records):
    """显示数据统计概览"""
    st.subheader("📈 数据统计概览")
    
    col_stat1, col_stat2, col_stat3, col_stat4 = st.columns(4)
    
    with col_stat1:
        st.metric("临床评估总数", len(records))
    
    with col_stat2:
        severities = [r.get('template', '自定义') for r in records]
        unique_severities = len(set(severities))
        st.metric("涉及严重程度类型", unique_severities)
    
    with col_stat3:
        contexts = [r['scene'] for r in records]
        unique_contexts = len(set(contexts))
        st.metric("涉及评估情境", unique_contexts)
    
    with col_stat4:
        if records:
            time_span = (max(r['timestamp'] for r in records) - min(r['timestamp'] for r in records)).days
            st.metric("评估时间跨度(天)", time_span)


def display_data_preview(records):
    """显示数据预览"""
    st.subheader("📊 临床数据预览")
    
    preview_data = []
    for record in records[:10]:
        # 计算核心症状严重度
        core_severity = (record['evaluation_scores']['社交互动质量'] + 
                        record['evaluation_scores']['沟通交流能力'] + 
                        record['evaluation_scores']['刻板重复行为']) / 3
        
        severity_level = "轻度" if core_severity < 2.5 else "中度" if core_severity < 3.5 else "重度"
        
        preview_data.append({
            '评估ID': record['experiment_id'][:25] + '...' if len(record['experiment_id']) > 25 else record['experiment_id'],
            '时间': record['timestamp'].strftime('%m-%d %H:%M'),
            '严重程度': record.get('template', '自定义')[:8] + '...' if len(record.get('template', '自定义')) > 8 else record.get('template', '自定义'),
            '评估情境': record['scene'].replace('结构化', '结构'),
            '核心症状': f"{core_severity:.2f}",
            '程度判断': severity_level
        })
    
    df_preview = pd.DataFrame(preview_data)
    st.dataframe(df_preview, use_container_width=True)
    
    if len(records) > 10:
        st.info(f"显示前10条记录，共{len(records)}条。完整数据请通过上方下载功能获取。")