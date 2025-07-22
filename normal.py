import streamlit as st
import pandas as pd
import json
import re
import time
import requests
import datetime
import plotly.express as px
import plotly.graph_objects as go
from typing import List, Dict, Any
import numpy as np
from concurrent.futures import ThreadPoolExecutor
import threading
import io
import base64
import zipfile
import tempfile

# 可选导入 - Excel功能
EXCEL_AVAILABLE = False
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.chart import BarChart, Reference
    EXCEL_AVAILABLE = True
except ImportError:
    st.warning("⚠️ 注意: 未安装openpyxl模块，Excel导出功能不可用。要启用Excel功能，请运行: pip install openpyxl")

# 页面配置
st.set_page_config(page_title="正常儿童成长发育AI观察平台 - 专业版", layout="wide")

# API配置
API_KEY = "sk-DQY3QAIcPGWTMfqZN1itL0qwl3y7ejrqyQwyGLyPom6TGz2v"  # 请替换为您的API Key
API_URL = "https://api.moonshot.cn/v1/chat/completions"

# 初始化session state
if 'observation_records' not in st.session_state:
    st.session_state.observation_records = []
if 'current_batch_results' not in st.session_state:
    st.session_state.current_batch_results = []
if 'observation_progress' not in st.session_state:
    st.session_state.observation_progress = {'current': 0, 'total': 0}

# 基于儿童发育理论的场景配置
DEVELOPMENT_SCENE_CONFIG = {
    "家庭日常生活": {
        "roles": ["儿童", "妈妈", "爸爸", "兄弟姐妹", "家庭成员"],
        "target": "观察儿童在家庭环境中的自然行为和亲子互动模式",
        "desc": "🏠 家庭日常生活环境",
        "activities": [
            "早晨起床洗漱", "家庭用餐时光", "亲子游戏时间", 
            "睡前故事分享", "家务参与活动", "家庭聊天交流"
        ],
        "triggers": [
            "新玩具出现", "计划突然改变", "需要分享物品", 
            "遇到小困难", "获得表扬", "需要做选择"
        ],
        "observation_points": [
            "情感表达能力", "亲子互动质量", "自理能力发展",
            "语言沟通技巧", "社交情绪发展", "认知学习表现"
        ]
    },
    "幼儿园社交": {
        "roles": ["儿童", "老师", "同班小朋友", "园长", "保育员"],
        "target": "评估儿童在集体环境中的社交发展和适应能力",
        "desc": "🎓 幼儿园集体社交环境", 
        "activities": [
            "晨间自由活动", "集体游戏时间", "课堂学习活动",
            "午餐分享时光", "户外运动游戏", "艺术创作活动"
        ],
        "triggers": [
            "新朋友加入", "玩具争抢情况", "团队合作任务",
            "表演展示机会", "需要帮助他人", "遵守规则要求"
        ],
        "observation_points": [
            "同伴交往能力", "集体适应性", "规则理解遵守",
            "情绪自我调节", "创造想象力", "学习专注力"
        ]
    },
    "户外探索活动": {
        "roles": ["儿童", "父母", "其他孩子", "户外教练", "安全员"],
        "target": "观察儿童的探索精神、运动发展和自然认知能力",
        "desc": "🌳 户外自然探索环境",
        "activities": [
            "公园自由探索", "自然观察活动", "体感运动游戏",
            "团队探险任务", "野餐分享活动", "环保实践行动"
        ],
        "triggers": [
            "发现新事物", "遇到挑战任务", "需要勇气尝试",
            "与自然互动", "团队协作需求", "安全意识考验"
        ],
        "observation_points": [
            "探索好奇心", "运动协调能力", "自然认知发展",
            "勇气胆量表现", "团队合作精神", "安全意识水平"
        ]
    },
    "学习成长环境": {
        "roles": ["儿童", "教师", "学习伙伴", "家长", "辅导员"],
        "target": "评估儿童的学习兴趣、认知发展和知识掌握能力",
        "desc": "📚 结构化学习成长环境",
        "activities": [
            "趣味学习活动", "知识探索游戏", "创意表达练习",
            "思维训练任务", "阅读分享时间", "科学实验体验"
        ],
        "triggers": [
            "新知识学习", "问题解决挑战", "创意表达机会",
            "思维逻辑训练", "成果展示时刻", "学习困难克服"
        ],
        "observation_points": [
            "学习兴趣动机", "认知理解能力", "创造表达能力",
            "逻辑思维发展", "专注持续性", "学习成就感"
        ]
    },
    "情感社交发展": {
        "roles": ["儿童", "心理老师", "朋友", "家庭成员", "社区伙伴"],
        "target": "专门观察儿童的情感发展、同理心和社交技能成长",
        "desc": "💖 情感社交能力发展环境",
        "activities": [
            "情感表达游戏", "同理心培养活动", "友谊建立互动",
            "冲突解决练习", "助人为乐体验", "团队协作项目"
        ],
        "triggers": [
            "他人情感需求", "友谊冲突处理", "需要表达感受",
            "帮助他人机会", "分享快乐时刻", "面对挫折困难"
        ],
        "observation_points": [
            "情感认知理解", "同理心发展", "友谊建立维护",
            "冲突解决能力", "利他行为表现", "情绪调节技巧"
        ]
    }
}

# 基于发育心理学的儿童年龄段特征配置
CHILDREN_DEVELOPMENT_PROFILES = {
    "2-3岁幼儿期": {
        "language_development": 4,  # 语言发展水平 (1-5, 5为最高)
        "social_skills": 3,         # 社交技能水平
        "cognitive_ability": 3,     # 认知能力水平
        "emotional_regulation": 2,  # 情绪调节能力
        "motor_skills": 4,          # 运动技能发展
        "independence_level": 2,    # 独立性水平
        "typical_interests": "玩具车、积木、简单绘本、音乐游戏",
        "development_focus": "语言爆发期，自我意识萌芽",
        "stage_characteristics": "探索期，模仿学习为主"
    },
    "3-4岁学前期": {
        "language_development": 5,
        "social_skills": 4,
        "cognitive_ability": 4,
        "emotional_regulation": 3,
        "motor_skills": 4,
        "independence_level": 3,
        "typical_interests": "角色扮演、拼图游戏、故事书、绘画创作",
        "development_focus": "想象力丰富，社交意识增强",
        "stage_characteristics": "好奇心旺盛，喜欢提问"
    },
    "4-5岁中班期": {
        "language_development": 5,
        "social_skills": 4,
        "cognitive_ability": 4,
        "emotional_regulation": 4,
        "motor_skills": 5,
        "independence_level": 4,
        "typical_interests": "复杂游戏、科学探索、运动活动、团队游戏",
        "development_focus": "规则意识形成，友谊关系建立",
        "stage_characteristics": "活跃好动，具备初步逻辑思维"
    },
    "5-6岁大班期": {
        "language_development": 5,
        "social_skills": 5,
        "cognitive_ability": 5,
        "emotional_regulation": 4,
        "motor_skills": 5,
        "independence_level": 4,
        "typical_interests": "学习准备活动、复杂建构、数字游戏、合作项目",
        "development_focus": "学习准备度提升，责任感增强",
        "stage_characteristics": "入学准备期，自控能力发展"
    },
    "6-8岁小学低年级": {
        "language_development": 5,
        "social_skills": 5,
        "cognitive_ability": 5,
        "emotional_regulation": 5,
        "motor_skills": 5,
        "independence_level": 5,
        "typical_interests": "学科学习、体育运动、艺术创作、科学实验",
        "development_focus": "学业适应，同伴关系深化",
        "stage_characteristics": "规则感强，求知欲旺盛"
    }
}

# 基于儿童发展心理学的综合评估指标
DEVELOPMENT_EVALUATION_METRICS = {
    "语言沟通发展": {
        "description": "儿童语言表达和理解交流的发展水平",
        "subscales": {
            "表达清晰度": "语言表达的清晰程度和词汇丰富性",
            "理解能力": "对他人语言和指令的理解水平", 
            "交流主动性": "主动发起交流和对话的积极性",
            "语用技能": "在不同情境中恰当使用语言的能力"
        },
        "scoring_criteria": {
            5: "优秀 - 语言发达，表达流畅，交流自然",
            4: "良好 - 语言发展正常，表达较清晰",
            3: "一般 - 基本语言能力，偶有不清晰", 
            2: "需关注 - 语言发展略慢，需要鼓励",
            1: "需支持 - 语言发展明显滞后，需要专业指导"
        }
    },
    "社交互动能力": {
        "description": "与他人建立关系和互动交往的能力发展",
        "subscales": {
            "友谊建立": "主动与他人建立友谊的能力",
            "合作分享": "与他人合作和分享的表现",
            "冲突解决": "处理人际冲突的成熟度",
            "群体适应": "在集体中的适应和融入能力"
        },
        "scoring_criteria": {
            5: "优秀 - 社交能力突出，人际关系和谐",
            4: "良好 - 能很好地与他人相处",
            3: "一般 - 基本社交能力，有待提高",
            2: "需关注 - 社交略显被动，需要引导",
            1: "需支持 - 社交困难，需要重点帮助"
        }
    },
    "认知学习能力": {
        "description": "学习新知识和解决问题的认知能力发展",
        "subscales": {
            "注意专注": "集中注意力和持续专注的能力",
            "记忆学习": "记忆和学习新信息的能力",
            "逻辑思维": "逻辑推理和抽象思维的发展",
            "创造想象": "创造性思维和想象力的表现"
        },
        "scoring_criteria": {
            5: "优秀 - 认知能力强，学习效率高",
            4: "良好 - 学习能力正常，理解力好",
            3: "一般 - 基本认知能力，学习稳定",
            2: "需关注 - 学习略有困难，需要支持",
            1: "需支持 - 认知发展滞后，需要专业帮助"
        }
    },
    "情绪调节发展": {
        "description": "情绪识别、表达和调节的发展水平",
        "subscales": {
            "情绪认知": "识别和理解自己及他人情绪的能力",
            "情绪表达": "恰当表达情绪和感受的能力",
            "自我调节": "管理和调节自己情绪的能力",
            "同理关怀": "理解他人感受并给予关怀的能力"
        },
        "scoring_criteria": {
            5: "优秀 - 情绪管理成熟，同理心强",
            4: "良好 - 情绪调节较好，表达恰当",
            3: "一般 - 基本情绪能力，有成长空间",
            2: "需关注 - 情绪调节需要指导",
            1: "需支持 - 情绪发展需要重点关注"
        }
    },
    "运动技能发展": {
        "description": "大运动和精细运动技能的发展水平",
        "subscales": {
            "大运动": "跑跳、平衡、协调等大肌肉运动能力",
            "精细运动": "手部精细动作和手眼协调能力",
            "运动规划": "运动计划和空间感知能力",
            "运动兴趣": "对体育活动和运动的积极性"
        },
        "scoring_criteria": {
            5: "优秀 - 运动技能发达，协调性好",
            4: "良好 - 运动发展正常，动作协调",
            3: "一般 - 基本运动能力，发展平稳",
            2: "需关注 - 运动发展略慢，需要练习",
            1: "需支持 - 运动技能滞后，需要专业指导"
        }
    },
    "独立自理能力": {
        "description": "日常生活自理和独立行动的能力发展",
        "subscales": {
            "生活自理": "日常生活自理能力的发展水平",
            "任务执行": "独立完成任务的能力",
            "问题解决": "遇到困难时的应对和解决能力",
            "责任意识": "对自己行为负责的意识发展"
        },
        "scoring_criteria": {
            5: "优秀 - 独立性强，自理能力突出",
            4: "良好 - 基本独立，自理能力好",
            3: "一般 - 部分独立，需要适度帮助",
            2: "需关注 - 依赖性较强，需要鼓励",
            1: "需支持 - 独立性不足，需要重点培养"
        }
    }
}

def call_kimi_with_profile(prompt, child_profile, max_retries=3):
    """调用AI API生成对话，带重试机制，基于儿童发展特征描述"""
    headers = {
        "Authorization": f"Bearer {API_KEY}",
        "Content-Type": "application/json"
    }
    
    profile_description = f"""
    正常儿童发展特征配置（基于发育心理学理论）：
    - 年龄发展阶段: {child_profile.get('stage_characteristics', '未指定')}
    - 语言发展水平: {child_profile['language_development']}/5 (5为最高水平)
    - 社交技能水平: {child_profile['social_skills']}/5
    - 认知能力水平: {child_profile['cognitive_ability']}/5
    - 情绪调节能力: {child_profile['emotional_regulation']}/5
    - 运动技能发展: {child_profile['motor_skills']}/5
    - 独立性水平: {child_profile['independence_level']}/5
    - 典型兴趣爱好: {child_profile['typical_interests']}
    - 发展重点: {child_profile.get('development_focus', '未指定')}
    """
    
    system_prompt = (
        "你是一个专业的儿童发展心理学专家，请严格按照以下儿童发展理论来模拟正常健康儿童的行为：\n"
        + profile_description +
        "\n正常儿童行为表现要求："
        "\n1. 语言沟通：根据年龄展现相应的语言能力、词汇量和表达方式"
        "\n2. 社交互动：体现同龄儿童的社交兴趣、友谊建立和合作能力"
        "\n3. 认知学习：展现好奇心、学习兴趣和相应的思维能力"
        "\n4. 情绪发展：表现年龄适宜的情绪表达和调节能力"
        "\n5. 运动发展：体现相应的大运动和精细运动发展水平"
        "\n6. 独立能力：展现符合年龄的自理能力和独立性"
        "\n严格格式：\"角色名:发言内容\"。每句换行，语言生动自然，行为符合正常儿童发展特点。"
        "\n要求：展现儿童的纯真、好奇、活跃和成长特点，避免过于成熟或不符合年龄的表达。"
    )
    
    payload = {
        "model": "moonshot-v1-8k",
        "messages": [
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": prompt}
        ],
        "temperature": 0.7,
        "max_tokens": 2048
    }
    
    for attempt in range(max_retries):
        try:
            response = requests.post(API_URL, headers=headers, json=payload, timeout=30)
            
            if response.status_code == 200:
                return response.json()['choices'][0]['message']['content']
            elif response.status_code == 429:
                wait_time = (attempt + 1) * 5
                print(f"API速率限制，等待{wait_time}秒后重试...")
                if 'st' in globals():
                    st.warning(f"⏱️ API速率限制，等待{wait_time}秒后重试... (尝试 {attempt + 1}/{max_retries})")
                time.sleep(wait_time)
                continue
            else:
                return f"API调用失败: {response.status_code} - {response.text}"
                
        except requests.exceptions.Timeout:
            if attempt == max_retries - 1:
                return "API调用超时，请稍后重试"
            time.sleep(2)
            continue
        except Exception as e:
            if attempt == max_retries - 1:
                return f"网络错误: {str(e)}"
            time.sleep(2)
            continue
    
    return "API调用失败：超过最大重试次数"

def evaluate_child_development(dialogue, child_profile, scene_info):
    """基于发育心理学评估儿童发展水平"""
    lines = dialogue.split('\n')
    child_lines = [line for line in lines if '儿童' in line]
    
    if not child_lines:
        return {metric: 3.0 for metric in DEVELOPMENT_EVALUATION_METRICS.keys()}
    
    evaluation_scores = {}
    
    # 语言沟通发展评估
    language_base = child_profile['language_development']
    language_indicators = 0
    
    # 检查语言表达质量
    expressive_quality = len([line for line in child_lines if any(word in line for word in ['我觉得', '我想要', '我喜欢', '因为'])])
    if expressive_quality > 0:
        language_indicators += 1
    
    # 检查交流主动性
    initiative_communication = len([line for line in child_lines if any(word in line for word in ['你好', '我们', '一起', '可以吗'])])
    if initiative_communication > 0:
        language_indicators += 1
    
    # 检查问题询问
    questioning = len([line for line in child_lines if '?' in line or any(word in line for word in ['为什么', '怎么', '什么'])])
    if questioning > 0:
        language_indicators += 0.5
    
    language_score = min(5, language_base + (language_indicators * 0.3))
    evaluation_scores["语言沟通发展"] = max(1, language_score)
    
    # 社交互动能力评估
    social_base = child_profile['social_skills']
    social_quality = 0
    
    # 检查友善互动
    friendly_interaction = len([line for line in child_lines if any(word in line for word in ['谢谢', '对不起', '分享', '帮助'])])
    if friendly_interaction > 0:
        social_quality += 1
    
    # 检查合作行为
    cooperation = len([line for line in child_lines if any(word in line for word in ['一起', '我们', '大家', '合作'])])
    if cooperation > 0:
        social_quality += 0.5
    
    # 检查情感表达
    emotional_expression = len([line for line in child_lines if any(word in line for word in ['开心', '高兴', '难过', '生气'])])
    if emotional_expression > 0:
        social_quality += 0.5
    
    social_score = min(5, social_base + (social_quality * 0.3))
    evaluation_scores["社交互动能力"] = max(1, social_score)
    
    # 认知学习能力评估
    cognitive_base = child_profile['cognitive_ability']
    cognitive_indicators = 0
    
    # 检查好奇探索
    curiosity = len([line for line in child_lines if any(word in line for word in ['为什么', '怎么', '想知道', '有趣'])])
    if curiosity > 0:
        cognitive_indicators += 1
    
    # 检查学习兴趣
    learning_interest = len([line for line in child_lines if any(word in line for word in ['学会', '明白', '知道了', '我懂了'])])
    if learning_interest > 0:
        cognitive_indicators += 0.5
    
    # 检查创造想象
    creativity = len([line for line in child_lines if any(word in line for word in ['想象', '创造', '发明', '设计'])])
    if creativity > 0:
        cognitive_indicators += 0.5
    
    cognitive_score = min(5, cognitive_base + (cognitive_indicators * 0.3))
    evaluation_scores["认知学习能力"] = max(1, cognitive_score)
    
    # 情绪调节发展评估  
    emotion_base = child_profile['emotional_regulation']
    emotion_regulation = 0
    
    # 检查情绪表达
    emotion_words = ['开心', '难过', '生气', '害怕', '兴奋', '紧张']
    emotion_expressions = len([line for line in child_lines 
                              if any(word in line for word in emotion_words)])
    if emotion_expressions > 0:
        emotion_regulation += 0.5
    
    # 检查同理心表现
    empathy_words = ['你还好吗', '没关系', '安慰', '关心']
    empathy_attempts = len([line for line in child_lines 
                           if any(word in line for word in empathy_words)])
    if empathy_attempts > 0:
        emotion_regulation += 0.5
    
    # 检查问题解决
    problem_solving = len([line for line in child_lines 
                          if any(word in line for word in ['试试', '想办法', '解决', '努力'])])
    if problem_solving > 0:
        emotion_regulation += 0.5
    
    emotion_score = min(5, emotion_base + (emotion_regulation * 0.3))
    evaluation_scores["情绪调节发展"] = max(1, emotion_score)
    
    # 运动技能发展评估
    motor_base = child_profile['motor_skills']
    motor_performance = 0
    
    # 检查运动相关活动
    motor_activities = len([line for line in child_lines 
                           if any(word in line for word in ['跑', '跳', '爬', '画', '拿', '做'])])
    if motor_activities > 0:
        motor_performance += 0.5
    
    # 检查精细动作
    fine_motor = len([line for line in child_lines 
                     if any(word in line for word in ['画画', '写字', '剪纸', '拼图'])])
    if fine_motor > 0:
        motor_performance += 0.5
    
    motor_score = min(5, motor_base + (motor_performance * 0.3))
    evaluation_scores["运动技能发展"] = max(1, motor_score)
    
    # 独立自理能力评估
    independence_base = child_profile['independence_level']
    independence_quality = 0
    
    # 检查自主行为
    autonomous_behavior = len([line for line in child_lines 
                              if any(word in line for word in ['我自己', '我来', '我会', '我能'])])
    if autonomous_behavior > 0:
        independence_quality += 0.5
    
    # 检查责任意识
    responsibility = len([line for line in child_lines 
                         if any(word in line for word in ['我的', '应该', '负责', '任务'])])
    if responsibility > 0:
        independence_quality += 0.5
    
    independence_score = min(5, independence_base + (independence_quality * 0.3))
    evaluation_scores["独立自理能力"] = max(1, independence_score)
    
    # 添加随机变异模拟真实发展的不确定性
    for metric in evaluation_scores:
        variation = np.random.normal(0, 0.15)  # 小幅随机变化
        evaluation_scores[metric] = max(1, min(5, evaluation_scores[metric] + variation))
        evaluation_scores[metric] = round(evaluation_scores[metric], 2)
    
    return evaluation_scores

def generate_observation_batch(templates, scenes, num_observations_per_combo=3):
    """生成批量观察配置"""
    observations = []
    observation_counter = 0
    
    for template_name, profile in templates.items():
        for scene_name, scene_data in scenes.items():
            for activity in scene_data['activities'][:2]:
                for trigger in scene_data['triggers'][:2]:
                    for i in range(num_observations_per_combo):
                        observation_counter += 1
                        
                        # 添加轻微发展变异
                        varied_profile = profile.copy()
                        for key in ['language_development', 'social_skills', 'cognitive_ability']:
                            if key in varied_profile:
                                variation = np.random.randint(-1, 2)
                                varied_profile[key] = max(1, min(5, varied_profile[key] + variation))
                        
                        timestamp = datetime.datetime.now().strftime('%Y%m%d_%H%M%S_%f')[:-3]
                        unique_id = f"DEV_{observation_counter:03d}_{template_name[:4]}_{scene_name[:4]}_{timestamp}"
                        
                        observations.append({
                            'template': template_name,
                            'scene': scene_name,
                            'activity': activity,
                            'trigger': trigger,
                            'child_profile': varied_profile,
                            'observation_id': unique_id,
                            'batch_index': observation_counter
                        })
    
    return observations

def run_single_observation(observation_config):
    """运行单个观察"""
    try:
        scene_data = DEVELOPMENT_SCENE_CONFIG[observation_config['scene']]
        
        # 构建基于儿童发展的prompt
        prompt = (
            f"儿童成长观察情境：{observation_config['scene']} - {observation_config['activity']}\n"
            f"观察要点：{', '.join(scene_data['observation_points'][:3])}\n"
            f"情境触发：{observation_config['trigger']}\n"
            f"参与角色：儿童、{scene_data['roles'][1]}、{scene_data['roles'][2]}\n"
            f"请基于儿童发展心理学理论，模拟该年龄段儿童在此情境下的自然行为表现。\n"
            f"要求：15-20轮对话，体现儿童的天真好奇、活泼可爱，"
            f"语言和行为符合该年龄段的正常发展特征。\n"
            f"格式：'角色名:内容'，每句换行。体现儿童的纯真和成长活力。"
        )
        
        dialogue = call_kimi_with_profile(prompt, observation_config['child_profile'])
        
        # 使用发展标准评估
        evaluation_scores = evaluate_child_development(
            dialogue, 
            observation_config['child_profile'], 
            DEVELOPMENT_SCENE_CONFIG[observation_config['scene']]
        )
        
        record = {
            'observation_id': observation_config['observation_id'],
            'timestamp': datetime.datetime.now(),
            'template': observation_config['template'],
            'scene': observation_config['scene'],
            'activity': observation_config['activity'],
            'trigger': observation_config['trigger'],
            'child_profile': observation_config['child_profile'],
            'dialogue': dialogue,
            'evaluation_scores': evaluation_scores,
            'developmental_observations': extract_developmental_observations(dialogue),
            'notes': f"发展观察 - {observation_config['template']}"
        }
        
        return record
        
    except Exception as e:
        return {
            'observation_id': observation_config['observation_id'],
            'timestamp': datetime.datetime.now(),
            'error': str(e),
            'template': observation_config.get('template', 'unknown'),
            'scene': observation_config.get('scene', 'unknown')
        }

def extract_developmental_observations(dialogue):
    """从对话中提取发展观察要点"""
    lines = dialogue.split('\n')
    child_lines = [line for line in lines if '儿童' in line]
    
    observations = {
        "语言表达特点": [],
        "社交互动表现": [],
        "认知学习行为": [],
        "情绪情感反应": [],
        "运动活动参与": []
    }
    
    for line in child_lines:
        # 语言表达识别
        if any(word in line for word in ['我想', '我觉得', '因为', '如果']):
            observations["语言表达特点"].append("复杂句式表达")
        elif any(word in line for word in ['为什么', '怎么', '什么']):
            observations["语言表达特点"].append("主动提问询问")
        
        # 社交互动识别  
        if any(word in line for word in ['一起', '我们', '分享', '帮助']):
            observations["社交互动表现"].append("合作分享意识")
        elif any(word in line for word in ['谢谢', '对不起', '请']):
            observations["社交互动表现"].append("礼貌社交行为")
        
        # 认知学习识别
        if any(word in line for word in ['学会', '明白', '知道', '理解']):
            observations["认知学习行为"].append("学习理解能力")
        elif any(word in line for word in ['想象', '创造', '发明']):
            observations["认知学习行为"].append("创造想象表现")
        
        # 情绪反应识别
        if any(word in line for word in ['开心', '高兴', '兴奋']):
            observations["情绪情感反应"].append("积极情绪表达")
        elif any(word in line for word in ['难过', '生气', '害怕']):
            observations["情绪情感反应"].append("情绪认知表达")
        
        # 运动活动识别
        if any(word in line for word in ['跑', '跳', '爬', '玩']):
            observations["运动活动参与"].append("大运动活动")
        elif any(word in line for word in ['画', '写', '做', '拼']):
            observations["运动活动参与"].append("精细动作表现")
    
    # 清理空列表
    observations = {k: v for k, v in observations.items() if v}
    
    return observations

def run_batch_observations(observations, progress_callback=None):
    """运行批量观察"""
    results = []
    delay_between_requests = 25
    
    for i, observation in enumerate(observations):
        if progress_callback:
            progress_callback(i + 1, len(observations))
        
        if 'st' in globals():
            remaining_observations = len(observations) - i - 1
            estimated_time = remaining_observations * delay_between_requests / 60
            st.info(f"⏳ 正在处理第 {i+1}/{len(observations)} 个发展观察，预计还需 {estimated_time:.1f} 分钟")
        
        result = run_single_observation(observation)
        results.append(result)
        
        if i < len(observations) - 1:
            print(f"等待{delay_between_requests}秒避免API限制...")
            if 'st' in globals():
                progress_bar = st.progress(0)
                for wait_second in range(delay_between_requests):
                    progress_bar.progress((wait_second + 1) / delay_between_requests)
                    time.sleep(1)
                progress_bar.empty()
            else:
                time.sleep(delay_between_requests)
    
    return results

def generate_development_analysis(records):
    """生成儿童发展的统计分析报告"""
    if not records:
        return {}
    
    analysis = {}
    
    # 基础发展统计
    analysis['发展观察概况'] = {
        '观察总数': len(records),
        '观察时间跨度': f"{min(r['timestamp'] for r in records).strftime('%Y-%m-%d')} 至 {max(r['timestamp'] for r in records).strftime('%Y-%m-%d')}",
        '涉及情境数': len(set(r['scene'] for r in records)),
        '涉及年龄段数': len(set(r.get('template', '自定义') for r in records))
    }
    
    # 按年龄段分析
    age_stats = {}
    for record in records:
        age_group = record.get('template', '自定义')
        if age_group not in age_stats:
            age_stats[age_group] = {
                '观察次数': 0,
                '语言发展得分': [],
                '社交能力得分': [],
                '认知学习得分': [],
                '情绪调节得分': [],
                '运动技能得分': [],
                '独立能力得分': []
            }
        age_stats[age_group]['观察次数'] += 1
        age_stats[age_group]['语言发展得分'].append(record['evaluation_scores']['语言沟通发展'])
        age_stats[age_group]['社交能力得分'].append(record['evaluation_scores']['社交互动能力'])
        age_stats[age_group]['认知学习得分'].append(record['evaluation_scores']['认知学习能力'])
        age_stats[age_group]['情绪调节得分'].append(record['evaluation_scores']['情绪调节发展'])
        age_stats[age_group]['运动技能得分'].append(record['evaluation_scores']['运动技能发展'])
        age_stats[age_group]['独立能力得分'].append(record['evaluation_scores']['独立自理能力'])
    
    # 计算统计值
    for age_group, stats in age_stats.items():
        for metric in ['语言发展得分', '社交能力得分', '认知学习得分', '情绪调节得分', '运动技能得分', '独立能力得分']:
            scores = stats[metric]
            stats[f'{metric}_均值'] = np.mean(scores)
            stats[f'{metric}_标准差'] = np.std(scores)
            stats[f'{metric}_范围'] = f"{np.min(scores):.1f}-{np.max(scores):.1f}"
            del stats[metric]
    
    analysis['年龄段分析'] = age_stats
    
    # 按观察情境分析
    context_stats = {}
    for record in records:
        context = record['scene']
        if context not in context_stats:
            context_stats[context] = {
                '观察次数': 0,
                '语言表现': [],
                '社交表现': [],
                '学习表现': []
            }
        context_stats[context]['观察次数'] += 1
        context_stats[context]['语言表现'].append(record['evaluation_scores']['语言沟通发展'])
        context_stats[context]['社交表现'].append(record['evaluation_scores']['社交互动能力'])
        context_stats[context]['学习表现'].append(record['evaluation_scores']['认知学习能力'])
    
    for context, stats in context_stats.items():
        for metric in ['语言表现', '社交表现', '学习表现']:
            scores = stats[metric]
            stats[f'{metric}_均值'] = np.mean(scores)
            del stats[metric]
    
    analysis['情境分析'] = context_stats
    
    # 整体发展表现
    all_language = [r['evaluation_scores']['语言沟通发展'] for r in records]
    all_social = [r['evaluation_scores']['社交互动能力'] for r in records]
    all_cognitive = [r['evaluation_scores']['认知学习能力'] for r in records]
    all_emotional = [r['evaluation_scores']['情绪调节发展'] for r in records]
    all_motor = [r['evaluation_scores']['运动技能发展'] for r in records]
    all_independence = [r['evaluation_scores']['独立自理能力'] for r in records]
    
    analysis['整体发展表现'] = {
        '语言沟通发展水平': f"{np.mean(all_language):.2f} ± {np.std(all_language):.2f}",
        '社交互动能力水平': f"{np.mean(all_social):.2f} ± {np.std(all_social):.2f}",
        '认知学习能力水平': f"{np.mean(all_cognitive):.2f} ± {np.std(all_cognitive):.2f}",
        '情绪调节发展水平': f"{np.mean(all_emotional):.2f} ± {np.std(all_emotional):.2f}",
        '运动技能发展水平': f"{np.mean(all_motor):.2f} ± {np.std(all_motor):.2f}",
        '独立自理能力水平': f"{np.mean(all_independence):.2f} ± {np.std(all_independence):.2f}",
        '综合发展指数': f"{(np.mean(all_language) + np.mean(all_social) + np.mean(all_cognitive))/3:.2f}"
    }
    
    # 发展建议和指导
    recommendations = []
    
    # 分析整体发展水平
    overall_avg = (np.mean(all_language) + np.mean(all_social) + np.mean(all_cognitive) + 
                   np.mean(all_emotional) + np.mean(all_motor) + np.mean(all_independence)) / 6
    
    if overall_avg >= 4.5:
        recommendations.append("整体发展优秀，建议继续保持良好的成长环境")
    elif overall_avg >= 4.0:
        recommendations.append("发展水平良好，可适当增加挑战性活动")
    elif overall_avg >= 3.5:
        recommendations.append("发展基本正常，建议多样化成长体验")
    else:
        recommendations.append("某些领域需要重点关注，建议增加针对性活动")
    
    # 分析强弱项
    domains = {
        '语言': np.mean(all_language),
        '社交': np.mean(all_social),
        '认知': np.mean(all_cognitive),
        '情绪': np.mean(all_emotional),
        '运动': np.mean(all_motor),
        '独立': np.mean(all_independence)
    }
    
    strongest = max(domains.keys(), key=lambda x: domains[x])
    weakest = min(domains.keys(), key=lambda x: domains[x])
    
    recommendations.append(f"{strongest}发展是优势领域，可以作为其他能力发展的支撑")
    
    if domains[weakest] < 3.5:
        recommendations.append(f"{weakest}发展需要特别关注，建议增加相关活动")
    
    # 分析最佳发展情境
    if context_stats:
        best_context = max(context_stats.keys(), 
                          key=lambda x: (context_stats[x]['语言表现_均值'] + 
                                       context_stats[x]['社交表现_均值'] + 
                                       context_stats[x]['学习表现_均值']) / 3)
        recommendations.append(f"在{best_context}中表现最佳，可多安排类似活动")
    
    analysis['发展建议与指导'] = recommendations
    
    return analysis

def create_development_excel_report(records, analysis):
    """创建儿童发展的Excel报告"""
    if not EXCEL_AVAILABLE:
        return None
    
    output = io.BytesIO()
    workbook = Workbook()
    workbook.remove(workbook.active)
    
    # 1. 发展观察概览
    overview_sheet = workbook.create_sheet("儿童发展观察概览")
    overview_sheet.append(["正常儿童成长发育观察报告（基于发育心理学理论）"])
    overview_sheet.append([])
    overview_sheet.append(["报告生成时间", datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')])
    overview_sheet.append(["评估理论基础", "儿童发展心理学 + 多元智能理论"])
    overview_sheet.append([])
    
    overview_sheet.append(["观察概况"])
    for key, value in analysis.get('发展观察概况', {}).items():
        overview_sheet.append([key, value])
    
    overview_sheet.append([])
    overview_sheet.append(["整体发展表现"])
    for key, value in analysis.get('整体发展表现', {}).items():
        overview_sheet.append([key, value])
    
    overview_sheet.append([])
    overview_sheet.append(["发展建议与指导"])
    for recommendation in analysis.get('发展建议与指导', []):
        overview_sheet.append([recommendation])
    
    # 2. 详细观察数据
    data_sheet = workbook.create_sheet("详细观察数据")
    headers = ["观察ID", "时间", "年龄段", "观察情境", "观察活动", "情境触发",
              "语言沟通发展", "社交互动能力", "认知学习能力", "情绪调节发展", 
              "运动技能发展", "独立自理能力", "综合发展指数",
              "发展阶段", "发展重点", "典型兴趣", "备注"]
    data_sheet.append(headers)
    
    for record in records:
        profile = record.get('child_profile', {})
        scores = record['evaluation_scores']
        development_index = sum(scores.values()) / len(scores)
        
        row = [
            record['observation_id'],
            record['timestamp'].strftime('%Y-%m-%d %H:%M:%S'),
            record.get('template', '自定义'),
            record['scene'],
            record.get('activity', ''),
            record.get('trigger', ''),
            scores['语言沟通发展'],
            scores['社交互动能力'],
            scores['认知学习能力'],
            scores['情绪调节发展'],
            scores['运动技能发展'],
            scores['独立自理能力'],
            f"{development_index:.2f}",
            profile.get('stage_characteristics', ''),
            profile.get('development_focus', ''),
            profile.get('typical_interests', ''),
            record.get('notes', '')
        ]
        data_sheet.append(row)
    
    # 3. 年龄段对比分析
    if analysis.get('年龄段分析'):
        age_sheet = workbook.create_sheet("年龄段分析")
        age_headers = ["年龄段", "观察次数", "语言发展均值", "社交能力均值", 
                      "认知学习均值", "情绪调节均值", "运动技能均值", "独立能力均值",
                      "综合发展"]
        age_sheet.append(age_headers)
        
        for age_group, stats in analysis['年龄段分析'].items():
            development_avg = (stats['语言发展得分_均值'] + stats['社交能力得分_均值'] + 
                             stats['认知学习得分_均值'] + stats['情绪调节得分_均值'] + 
                             stats['运动技能得分_均值'] + stats['独立能力得分_均值']) / 6
            row = [
                age_group,
                stats['观察次数'],
                f"{stats['语言发展得分_均值']:.2f}",
                f"{stats['社交能力得分_均值']:.2f}",
                f"{stats['认知学习得分_均值']:.2f}",
                f"{stats['情绪调节得分_均值']:.2f}",
                f"{stats['运动技能得分_均值']:.2f}",
                f"{stats['独立能力得分_均值']:.2f}",
                f"{development_avg:.2f}"
            ]
            age_sheet.append(row)
    
    # 4. 发展观察记录
    if any('developmental_observations' in record for record in records):
        obs_sheet = workbook.create_sheet("发展观察记录")
        obs_sheet.append(["观察ID", "语言表达特点", "社交互动表现", "认知学习行为", "情绪情感反应", "运动活动参与"])
        
        for record in records:
            if 'developmental_observations' in record:
                obs = record['developmental_observations']
                row = [
                    record['observation_id'],
                    '; '.join(obs.get('语言表达特点', [])),
                    '; '.join(obs.get('社交互动表现', [])),
                    '; '.join(obs.get('认知学习行为', [])),
                    '; '.join(obs.get('情绪情感反应', [])),
                    '; '.join(obs.get('运动活动参与', []))
                ]
                obs_sheet.append(row)
    
    # 5. 对话记录（用于质性分析）
    dialogue_sheet = workbook.create_sheet("行为对话记录")
    dialogue_sheet.append(["观察ID", "年龄段", "观察情境", "对话内容"])
    
    for record in records:
        dialogue_sheet.append([
            record['observation_id'],
            record.get('template', '自定义'),
            record['scene'],
            record['dialogue']
        ])
    
    # 应用专业样式
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if cell.row == 1:
                    cell.font = Font(bold=True, size=12)
                    cell.fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
                    cell.font = Font(bold=True, color="FFFFFF")
                elif any(keyword in str(cell.value) for keyword in ['优秀', '良好', '发展']) if cell.value else False:
                    cell.fill = PatternFill(start_color="E8F5E8", end_color="E8F5E8", fill_type="solid")
                cell.alignment = Alignment(wrap_text=True, vertical="top")
    
    workbook.save(output)
    output.seek(0)
    return output.getvalue()

# 主页面
st.title("🌟 正常儿童成长发育AI观察平台 - 专业版")
st.markdown("**基于儿童发展心理学理论和多元智能评估框架**")

# 侧边栏导航
st.sidebar.title("🧭 导航")
page = st.sidebar.selectbox("选择功能页面", [
    "快速发育观察", "批量发展研究", "个性化观察设计", 
    "发展数据分析", "观察记录管理", "📊 发展报告中心"
])

# 页面路由
if page == "快速发育观察":
    st.header("👶 快速发育观察")
    st.markdown("使用标准化年龄段分级进行快速儿童发展观察")
    
    col1, col2 = st.columns(2)
    
    with col1:
        st.subheader("👦 选择观察对象")
        selected_age_group = st.selectbox("年龄发展阶段", list(CHILDREN_DEVELOPMENT_PROFILES.keys()))
        
        profile = CHILDREN_DEVELOPMENT_PROFILES[selected_age_group]
        
        # 显示发展特征
        with st.expander("查看发展特征配置", expanded=True):
            st.write(f"**发展阶段特征**: {profile['stage_characteristics']}")
            st.write(f"**语言发展水平**: {profile['language_development']}/5")
            st.write(f"**社交技能水平**: {profile['social_skills']}/5")
            st.write(f"**认知能力水平**: {profile['cognitive_ability']}/5")
            st.write(f"**情绪调节能力**: {profile['emotional_regulation']}/5")
            st.write(f"**运动技能发展**: {profile['motor_skills']}/5")
            st.write(f"**独立性水平**: {profile['independence_level']}/5")
            st.write(f"**典型兴趣**: {profile['typical_interests']}")
            st.write(f"**发展重点**: {profile['development_focus']}")
        
        selected_scene = st.selectbox("选择观察情境", list(DEVELOPMENT_SCENE_CONFIG.keys()))
        
        scene_data = DEVELOPMENT_SCENE_CONFIG[selected_scene]
        
        # 显示场景信息
        with st.expander("观察情境详情"):
            st.write(f"**目标**: {scene_data['target']}")
            st.write(f"**观察要点**: {', '.join(scene_data['observation_points'])}")
        
        selected_activity = st.selectbox("选择观察活动", scene_data['activities'])
        selected_trigger = st.selectbox("选择情境触发", scene_data['triggers'])
    
    with col2:
        st.subheader("🔍 执行观察")
        
        if st.button("🌟 开始发育观察", type="primary", use_container_width=True):
            timestamp = datetime.datetime.now().strftime('%Y%m%d_%H%M%S_%f')[:-3]
            observation_config = {
                'template': selected_age_group,
                'scene': selected_scene,
                'activity': selected_activity,
                'trigger': selected_trigger,
                'child_profile': profile.copy(),
                'observation_id': f"DEV_{selected_age_group[:4]}_{timestamp}"
            }
            
            with st.spinner("🤖 正在生成发育观察对话..."):
                result = run_single_observation(observation_config)
            
            if 'error' not in result:
                st.session_state.observation_records.append(result)
                
                st.success(f"✅ 发育观察完成！ID: {result['observation_id']}")
                
                # 显示观察结果
                st.subheader("📊 发育观察结果")
                
                col_result1, col_result2 = st.columns(2)
                
                with col_result1:
                    st.write("**发展能力评估得分** (5分为最高水平):")
                    for metric, score in result['evaluation_scores'].items():
                        # 根据得分显示不同颜色
                        if score >= 4.5:
                            st.success(f"{metric}: {score}/5.0 (优秀)")
                        elif score >= 4.0:
                            st.info(f"{metric}: {score}/5.0 (良好)")
                        elif score >= 3.0:
                            st.warning(f"{metric}: {score}/5.0 (一般)")
                        else:
                            st.error(f"{metric}: {score}/5.0 (需关注)")
                
                with col_result2:
                    st.write("**发展观察要点**:")
                    if 'developmental_observations' in result:
                        for category, observations in result['developmental_observations'].items():
                            if observations:
                                st.write(f"**{category}**: {', '.join(observations)}")
                    
                    st.write("**对话预览**:")
                    dialogue_lines = result['dialogue'].split('\n')[:8]
                    for line in dialogue_lines:
                        if ':' in line and line.strip():
                            if '儿童' in line:
                                st.markdown(f"👶 {line}")
                            else:
                                st.markdown(f"👤 {line}")
                    
                    if len(result['dialogue'].split('\n')) > 8:
                        st.markdown("*...查看完整记录请前往'观察记录管理'*")
                
                # 显示发展建议
                st.subheader("💡 发展建议")
                development_avg = sum(result['evaluation_scores'].values()) / len(result['evaluation_scores'])
                
                if development_avg >= 4.5:
                    st.success("🌟 发展表现优秀！建议继续保持多样化的成长环境")
                elif development_avg >= 4.0:
                    st.info("👍 发展表现良好！可适当增加挑战性活动")
                elif development_avg >= 3.0:
                    st.warning("📈 发展基本正常，建议增加针对性的成长活动")
                else:
                    st.error("🔍 某些方面需要重点关注，建议咨询专业儿童发展指导")
                    
            else:
                st.error(f"❌ 观察失败: {result['error']}")

elif page == "批量发展研究":
    st.header("📊 批量发展研究")
    st.markdown("进行多组发展对照研究，获取统计学有效的发展数据")
    
    col1, col2 = st.columns([2, 1])
    
    with col1:
        st.subheader("🎯 研究设计")
        
        research_scale = st.radio(
            "选择研究规模",
            ["试点研究（推荐）", "标准研究", "大样本研究"],
            help="根据研究需要选择合适的样本规模"
        )
        
        if research_scale == "试点研究（推荐）":
            default_ages = list(CHILDREN_DEVELOPMENT_PROFILES.keys())[:2]
            default_contexts = list(DEVELOPMENT_SCENE_CONFIG.keys())[:2]
            default_repeats = 1
            st.info("🚀 试点研究：验证观察工具效果，约需5-8分钟")
        elif research_scale == "标准研究":
            default_ages = list(CHILDREN_DEVELOPMENT_PROFILES.keys())[:3]
            default_contexts = list(DEVELOPMENT_SCENE_CONFIG.keys())[:3]
            default_repeats = 2
            st.info("⏳ 标准研究：获得可靠发展数据，约需20-30分钟")
        else:
            default_ages = list(CHILDREN_DEVELOPMENT_PROFILES.keys())
            default_contexts = list(DEVELOPMENT_SCENE_CONFIG.keys())
            default_repeats = 2
            st.warning("⏰ 大样本研究：完整发展研究数据，约需60-90分钟")
        
        selected_ages = st.multiselect(
            "选择年龄段组", 
            list(CHILDREN_DEVELOPMENT_PROFILES.keys()),
            default=default_ages
        )
        
        selected_contexts = st.multiselect(
            "选择观察情境",
            list(DEVELOPMENT_SCENE_CONFIG.keys()),
            default=default_contexts
        )
        
        repeats_per_combo = st.slider(
            "每组合重复次数", 
            1, 3, 
            default_repeats,
            help="增加重复次数提高统计可靠性"
        )
        
        if selected_ages and selected_contexts:
            age_dict = {k: CHILDREN_DEVELOPMENT_PROFILES[k] for k in selected_ages}
            context_dict = {k: DEVELOPMENT_SCENE_CONFIG[k] for k in selected_contexts}
            
            observations = generate_observation_batch(
                age_dict, 
                context_dict, 
                repeats_per_combo
            )
            
            st.info(f"📊 将生成 {len(observations)} 个发展观察")
            
            # 研究设计预览
            with st.expander("研究设计预览", expanded=False):
                preview_df = pd.DataFrame([
                    {
                        '年龄段': obs['template'],
                        '观察情境': obs['scene'],
                        '观察活动': obs['activity'],
                        '情境触发': obs['trigger']
                    } for obs in observations[:10]
                ])
                st.dataframe(preview_df)
                if len(observations) > 10:
                    st.write(f"*...还有 {len(observations) - 10} 个观察*")
    
    with col2:
        st.subheader("🚀 执行研究")
        
        if 'development_batch_ready' not in st.session_state:
            st.session_state.development_batch_ready = False
        if 'development_batch_running' not in st.session_state:
            st.session_state.development_batch_running = False
        
        if selected_ages and selected_contexts:
            estimated_minutes = len(observations) * 25 / 60
            st.info(f"📊 观察数量: {len(observations)}")
            st.info(f"⏰ 预计时间: {estimated_minutes:.1f} 分钟")
            
            if not st.session_state.development_batch_ready and not st.session_state.development_batch_running:
                if st.button("⚡ 准备开始研究", use_container_width=True):
                    st.session_state.development_batch_ready = True
                    st.rerun()
            
            elif st.session_state.development_batch_ready and not st.session_state.development_batch_running:
                st.warning("⏰ **重要**: 由于API限制，批量研究需要较长时间")
                st.info("💡 确认后将立即开始，请保持网络稳定")
                
                col_btn1, col_btn2 = st.columns(2)
                
                with col_btn1:
                    if st.button("❌ 取消", use_container_width=True):
                        st.session_state.development_batch_ready = False
                        st.rerun()
                
                with col_btn2:
                    if st.button("✅ 开始研究", type="primary", use_container_width=True):
                        st.session_state.development_batch_running = True
                        st.session_state.development_batch_ready = False
                        st.rerun()
            
            elif st.session_state.development_batch_running:
                st.success("🔬 发展研究进行中...")
                
                progress_container = st.container()
                with progress_container:
                    progress_bar = st.progress(0)
                    status_text = st.empty()
                    result_container = st.empty()
                
                def update_progress(current, total):
                    progress = current / total
                    progress_bar.progress(progress)
                    remaining_time = (total - current) * 25 / 60
                    status_text.text(f"研究进度: {current}/{total} ({progress:.1%}) - 预计还需 {remaining_time:.1f} 分钟")
                
                try:
                    results = run_batch_observations(observations, update_progress)
                    
                    successful_results = [r for r in results if 'error' not in r]
                    failed_results = [r for r in results if 'error' in r]
                    
                    st.session_state.observation_records.extend(successful_results)
                    st.session_state.current_batch_results = successful_results
                    
                    with result_container:
                        st.success(f"✅ 发展研究完成！")
                        st.write(f"**成功观察**: {len(successful_results)} 个")
                        
                        if failed_results:
                            st.error(f"**失败观察**: {len(failed_results)} 个")
                        
                        if successful_results:
                            st.subheader("📈 研究结果概览")
                            
                            # 按年龄段统计
                            age_stats = {}
                            for result in successful_results:
                                age_group = result['template']
                                if age_group not in age_stats:
                                    age_stats[age_group] = []
                                
                                # 计算综合发展得分
                                development_score = sum(result['evaluation_scores'].values()) / len(result['evaluation_scores'])
                                age_stats[age_group].append(development_score)
                            
                            stats_df = pd.DataFrame([
                                {
                                    '年龄段': age_group,
                                    '样本数': len(scores),
                                    '发展指数均值': f"{np.mean(scores):.2f}",
                                    '标准差': f"{np.std(scores):.2f}",
                                    '95%置信区间': f"{np.mean(scores) - 1.96*np.std(scores)/np.sqrt(len(scores)):.2f}-{np.mean(scores) + 1.96*np.std(scores)/np.sqrt(len(scores)):.2f}"
                                } for age_group, scores in age_stats.items()
                            ])
                            
                            st.dataframe(stats_df, use_container_width=True)
                    
                    st.session_state.development_batch_running = False
                    
                    if st.button("🔄 进行新研究", use_container_width=True):
                        st.session_state.development_batch_ready = False
                        st.session_state.development_batch_running = False
                        st.rerun()
                        
                except Exception as e:
                    st.error(f"❌ 研究出错: {str(e)}")
                    st.session_state.development_batch_running = False
                    if st.button("🔄 重新尝试", use_container_width=True):
                        st.rerun()
        
        else:
            st.error("请先选择年龄段和观察情境")

elif page == "个性化观察设计":
    st.header("⚙️ 个性化观察设计")
    st.markdown("基于儿童发展理论自定义个体化观察参数")
    
    col1, col2 = st.columns(2)
    
    with col1:
        st.subheader("🎭 观察情境设置")
        selected_scene = st.selectbox("选择观察情境", list(DEVELOPMENT_SCENE_CONFIG.keys()))
        scene_data = DEVELOPMENT_SCENE_CONFIG[selected_scene]
        
        st.info(f"**观察目标**: {scene_data['target']}")
        
        # 显示观察要点
        with st.expander("发展观察要点"):
            for point in scene_data['observation_points']:
                st.write(f"• {point}")
        
        selected_activity = st.selectbox("选择观察活动", scene_data['activities'])
        selected_trigger = st.selectbox("选择情境触发", scene_data['triggers'])
    
    with col2:
        st.subheader("👶 儿童发展配置")
        
        template_base = st.selectbox("基于年龄段模板", ["自定义"] + list(CHILDREN_DEVELOPMENT_PROFILES.keys()))
        
        if template_base != "自定义":
            base_profile = CHILDREN_DEVELOPMENT_PROFILES[template_base]
            st.info(f"基于: {base_profile['stage_characteristics']}")
        else:
            base_profile = {
                'language_development': 3,
                'social_skills': 3,
                'cognitive_ability': 3,
                'emotional_regulation': 3,
                'motor_skills': 3,
                'independence_level': 3,
                'typical_interests': "各种游戏、探索活动",
                'development_focus': "全面发展",
                'stage_characteristics': "自定义配置"
            }
        
        st.write("**发展能力配置** (基于儿童发展心理学)")
        
        language_dev = st.slider(
            "语言发展水平", 1, 5, base_profile['language_development'],
            help="1=语言发展滞后，5=语言发展超前"
        )
        
        social_skills = st.slider(
            "社交技能水平", 1, 5, base_profile['social_skills'],
            help="1=社交技能需支持，5=社交技能优秀"
        )
        
        st.write("**其他发展配置**")
        
        cognitive_ability = st.slider(
            "认知能力水平", 1, 5, base_profile['cognitive_ability'],
            help="1=认知发展滞后，5=认知发展超前"
        )
        
        emotional_regulation = st.slider(
            "情绪调节能力", 1, 5, base_profile['emotional_regulation'],
            help="1=情绪调节困难，5=情绪调节成熟"
        )
        
        motor_skills = st.slider(
            "运动技能发展", 1, 5, base_profile['motor_skills'],
            help="1=运动发展滞后，5=运动技能优秀"
        )
        
        independence_level = st.slider(
            "独立性水平", 1, 5, base_profile['independence_level'],
            help="1=依赖性强，5=独立性强"
        )
        
        typical_interests = st.text_input(
            "典型兴趣爱好", 
            base_profile['typical_interests'],
            help="描述该儿童的主要兴趣和爱好"
        )
        
        development_focus = st.text_input(
            "发展重点",
            base_profile['development_focus'],
            help="当前阶段的主要发展任务"
        )
        
        # 根据配置自动推断发展特征
        total_development = language_dev + social_skills + cognitive_ability + emotional_regulation + motor_skills + independence_level
        avg_development = total_development / 6
        
        if avg_development >= 4.5:
            stage_desc = "发展优秀，各项能力超前"
        elif avg_development >= 4.0:
            stage_desc = "发展良好，能力均衡"
        elif avg_development >= 3.0:
            stage_desc = "发展正常，稳步成长"
        else:
            stage_desc = "发展需关注，需要更多支持"
        
        st.info(f"**推断的发展特征**: {stage_desc}")
        
        child_profile = {
            'language_development': language_dev,
            'social_skills': social_skills,
            'cognitive_ability': cognitive_ability,
            'emotional_regulation': emotional_regulation,
            'motor_skills': motor_skills,
            'independence_level': independence_level,
            'typical_interests': typical_interests,
            'development_focus': development_focus,
            'stage_characteristics': stage_desc
        }
    
    # 执行个性化观察
    st.subheader("🔍 执行个性化观察")
    
    if st.button("🌟 开始个性化观察", type="primary", use_container_width=True):
        timestamp = datetime.datetime.now().strftime('%Y%m%d_%H%M%S_%f')[:-3]
        observation_config = {
            'template': template_base if template_base != "自定义" else "个性化配置",
            'scene': selected_scene,
            'activity': selected_activity,
            'trigger': selected_trigger,
            'child_profile': child_profile,
            'observation_id': f"CUSTOM_{timestamp}"
        }
        
        with st.spinner("🤖 正在生成个性化观察..."):
            result = run_single_observation(observation_config)
        
        if 'error' not in result:
            st.session_state.observation_records.append(result)
            
            st.success(f"✅ 个性化观察完成！ID: {result['observation_id']}")
            
            # 显示详细观察结果
            st.subheader("📊 个性化观察结果")
            
            col_res1, col_res2, col_res3 = st.columns(3)
            
            with col_res1:
                st.write("**核心发展评估**:")
                st.metric("语言沟通发展", f"{result['evaluation_scores']['语言沟通发展']:.2f}/5")
                st.metric("社交互动能力", f"{result['evaluation_scores']['社交互动能力']:.2f}/5")
                st.metric("认知学习能力", f"{result['evaluation_scores']['认知学习能力']:.2f}/5")
                
                core_avg = (result['evaluation_scores']['语言沟通发展'] + 
                           result['evaluation_scores']['社交互动能力'] + 
                           result['evaluation_scores']['认知学习能力']) / 3
                st.metric("核心发展指数", f"{core_avg:.2f}/5")
            
            with col_res2:
                st.write("**其他发展评估**:")
                st.metric("情绪调节发展", f"{result['evaluation_scores']['情绪调节发展']:.2f}/5")
                st.metric("运动技能发展", f"{result['evaluation_scores']['运动技能发展']:.2f}/5")
                st.metric("独立自理能力", f"{result['evaluation_scores']['独立自理能力']:.2f}/5")
                
                all_avg = sum(result['evaluation_scores'].values()) / len(result['evaluation_scores'])
                st.metric("综合发展指数", f"{all_avg:.2f}/5")
            
            with col_res3:
                st.write("**发展观察**:")
                if 'developmental_observations' in result:
                    for category, observations in result['developmental_observations'].items():
                        if observations:
                            st.write(f"**{category}**:")
                            for obs in observations:
                                st.write(f"• {obs}")
                
            # 个性化建议
            st.subheader("💡 个性化发展建议")
            
            suggestions = []
            
            if result['evaluation_scores']['语言沟通发展'] >= 4.5:
                suggestions.append("🗣️ 语言发展优秀：可以尝试更复杂的表达活动和双语学习")
            elif result['evaluation_scores']['语言沟通发展'] < 3.0:
                suggestions.append("📚 语言发展需关注：建议增加阅读、故事分享和对话练习")
            
            if result['evaluation_scores']['社交互动能力'] >= 4.5:
                suggestions.append("👥 社交能力突出：可以承担小组领导角色，帮助其他孩子")
            elif result['evaluation_scores']['社交互动能力'] < 3.0:
                suggestions.append("🤝 社交能力需提升：建议多参与团体活动和角色扮演游戏")
            
            if result['evaluation_scores']['认知学习能力'] >= 4.5:
                suggestions.append("🧠 认知能力优秀：可以提供更有挑战性的学习任务")
            elif result['evaluation_scores']['认知学习能力'] < 3.0:
                suggestions.append("📖 认知发展需支持：建议通过游戏化学习增强学习兴趣")
            
            if result['evaluation_scores']['情绪调节发展'] < 3.0:
                suggestions.append("😌 情绪调节需指导：建议情绪识别训练和放松技巧练习")
            
            if result['evaluation_scores']['运动技能发展'] >= 4.5:
                suggestions.append("🏃 运动技能优秀：可以尝试更复杂的体育活动和精细动作训练")
            elif result['evaluation_scores']['运动技能发展'] < 3.0:
                suggestions.append("🤸 运动发展需加强：建议增加户外活动和手工制作")
            
            if not suggestions:
                suggestions.append("✅ 整体发展均衡良好，建议继续多样化的成长体验")
            
            for suggestion in suggestions:
                st.success(suggestion)
                
        else:
            st.error(f"❌ 观察失败: {result['error']}")
    
    # 保存配置
    if st.button("💾 保存观察配置", use_container_width=True):
        st.session_state.custom_child_profile = child_profile
        st.session_state.custom_scene_config = {
            'scene': selected_scene,
            'activity': selected_activity,
            'trigger': selected_trigger
        }
        st.success("✅ 个性化配置已保存！")

elif page == "发展数据分析":
    st.header("📈 发展数据分析")
    
    records = st.session_state.observation_records
    
    if not records:
        st.warning("📊 暂无观察数据，请先进行发展观察")
        st.stop()
    
    # 生成发展分析
    analysis = generate_development_analysis(records)
    
    # 发展概况
    st.subheader("🌟 发展观察概况")
    
    col1, col2, col3, col4 = st.columns(4)
    with col1:
        st.metric("观察总数", len(records))
    with col2:
        age_groups = [r.get('template', '自定义') for r in records]
        most_common = max(set(age_groups), key=age_groups.count) if age_groups else "无"
        st.metric("主要年龄段", most_common.split('期')[0])
    with col3:
        contexts = [r['scene'] for r in records]
        most_used_context = max(set(contexts), key=contexts.count)
        st.metric("主要观察情境", most_used_context.replace('日常', ''))
    with col4:
        all_development_scores = []
        for r in records:
            development_score = sum(r['evaluation_scores'].values()) / len(r['evaluation_scores'])
            all_development_scores.append(development_score)
        avg_development = np.mean(all_development_scores)
        st.metric("平均发展指数", f"{avg_development:.2f}/5")
    
    # 发展能力雷达图
    st.subheader("🎯 发展能力分析")
    
    # 发展能力雷达图
    avg_scores = {}
    for metric in DEVELOPMENT_EVALUATION_METRICS.keys():
        scores = [r['evaluation_scores'][metric] for r in records]
        avg_scores[metric] = np.mean(scores)
    
    fig_radar = go.Figure()
    fig_radar.add_trace(go.Scatterpolar(
        r=list(avg_scores.values()),
        theta=list(avg_scores.keys()),
        fill='toself',
        name='平均发展水平',
        line_color='rgb(100, 200, 100)'
    ))
    fig_radar.update_layout(
        polar=dict(
            radialaxis=dict(
                visible=True,
                range=[1, 5],
                tickvals=[1, 2, 3, 4, 5],
                ticktext=['需支持', '需关注', '一般', '良好', '优秀']
            )),
        showlegend=True,
        title="儿童发展能力雷达图",
        height=500
    )
    st.plotly_chart(fig_radar, use_container_width=True)
    
    # 年龄段对比分析
    st.subheader("📊 年龄段发展对比")
    
    if len(set([r.get('template', '自定义') for r in records])) > 1:
        age_data = {}
        for record in records:
            age_group = record.get('template', '自定义')
            if age_group not in age_data:
                age_data[age_group] = {
                    '语言沟通': [],
                    '社交互动': [],
                    '认知学习': [],
                    '情绪调节': [],
                    '运动技能': [],
                    '独立能力': []
                }
            
            age_data[age_group]['语言沟通'].append(record['evaluation_scores']['语言沟通发展'])
            age_data[age_group]['社交互动'].append(record['evaluation_scores']['社交互动能力'])
            age_data[age_group]['认知学习'].append(record['evaluation_scores']['认知学习能力'])
            age_data[age_group]['情绪调节'].append(record['evaluation_scores']['情绪调节发展'])
            age_data[age_group]['运动技能'].append(record['evaluation_scores']['运动技能发展'])
            age_data[age_group]['独立能力'].append(record['evaluation_scores']['独立自理能力'])
        
        # 创建对比图表
        comparison_data = []
        for age_group, abilities in age_data.items():
            for ability, scores in abilities.items():
                comparison_data.append({
                    '年龄段': age_group,
                    '发展领域': ability,
                    '平均得分': np.mean(scores),
                    '标准差': np.std(scores)
                })
        
        df_comparison = pd.DataFrame(comparison_data)
        
        fig_comparison = px.bar(
            df_comparison, 
            x='年龄段', 
            y='平均得分', 
            color='发展领域',
            title="不同年龄段的发展领域对比",
            labels={'平均得分': '发展水平 (1-5分)'},
            height=400
        )
        fig_comparison.update_layout(yaxis_range=[1, 5])
        st.plotly_chart(fig_comparison, use_container_width=True)
    
    # 观察情境效应分析
    st.subheader("🎭 观察情境效应分析")
    
    context_data = {}
    for record in records:
        context = record['scene']
        if context not in context_data:
            context_data[context] = []
        
        # 计算综合表现得分
        comprehensive_score = sum(record['evaluation_scores'].values()) / len(record['evaluation_scores'])
        context_data[context].append(comprehensive_score)
    
    if len(context_data) > 1:
        context_comparison = []
        for context, scores in context_data.items():
            context_comparison.append({
                '观察情境': context,
                '样本数': len(scores),
                '平均表现': np.mean(scores),
                '标准差': np.std(scores),
                '表现水平': '优秀' if np.mean(scores) >= 4.5 else '良好' if np.mean(scores) >= 4.0 else '一般'
            })
        
        df_context = pd.DataFrame(context_comparison)
        
        fig_context = px.bar(
            df_context,
            x='观察情境',
            y='平均表现',
            color='表现水平',
            title="不同观察情境下的表现对比",
            labels={'平均表现': '平均发展水平 (1-5分)'},
            height=400
        )
        st.plotly_chart(fig_context, use_container_width=True)
        
        # 显示情境分析表格
        st.dataframe(df_context.drop('表现水平', axis=1), use_container_width=True)
    
    # 发展建议和指导
    st.subheader("🔍 发展发现与成长指导")
    
    if analysis.get('发展建议与指导'):
        col_finding1, col_finding2 = st.columns(2)
        
        with col_finding1:
            st.write("### 📋 主要发展发现")
            for i, finding in enumerate(analysis['发展建议与指导'], 1):
                if '优秀' in finding or '良好' in finding:
                    st.success(f"{i}. {finding}")
                elif '需要' in finding or '关注' in finding:
                    st.warning(f"{i}. {finding}")
                else:
                    st.info(f"{i}. {finding}")
        
        with col_finding2:
            st.write("### 💡 成长支持建议")
            
            # 基于观察结果提供具体建议
            language_avg = np.mean([r['evaluation_scores']['语言沟通发展'] for r in records])
            social_avg = np.mean([r['evaluation_scores']['社交互动能力'] for r in records])
            cognitive_avg = np.mean([r['evaluation_scores']['认知学习能力'] for r in records])
            
            st.write("**基于发展心理学的成长建议**:")
            
            if language_avg >= 4.5:
                st.write("• 🗣️ **语言发展超前**")
                st.write("  - 可以尝试第二语言学习")
                st.write("  - 增加复杂表达和创作活动")
                st.write("  - 参与讲故事和演讲活动")
            elif language_avg < 3.0:
                st.write("• 📚 **语言发展需支持**")
                st.write("  - 增加亲子阅读时间")
                st.write("  - 多进行对话练习")
                st.write("  - 使用游戏化语言学习")
            
            if social_avg >= 4.5:
                st.write("• 👥 **社交能力突出**")
                st.write("  - 可以承担小组长角色")
                st.write("  - 参与更多团队合作项目")
                st.write("  - 培养领导力和同理心")
            elif social_avg < 3.0:
                st.write("• 🤝 **社交技能需培养**")
                st.write("  - 多参与集体活动")
                st.write("  - 练习分享和合作")
                st.write("  - 学习社交礼仪和情感表达")
            
            if cognitive_avg >= 4.5:
                st.write("• 🧠 **认知能力优秀**")
                st.write("  - 提供更有挑战性的任务")
                st.write("  - 培养创造性思维")
                st.write("  - 探索STEM和艺术活动")
            elif cognitive_avg < 3.0:
                st.write("• 📖 **认知发展需引导**")
                st.write("  - 通过游戏化学习")
                st.write("  - 培养观察和思考习惯")
                st.write("  - 循序渐进建立学习兴趣")
    
    # 统计显著性检验（如果有多组数据）
    age_groups = [r.get('template', '自定义') for r in records]
    if len(set(age_groups)) > 1:
        st.subheader("📐 统计学分析")
        
        try:
            from scipy import stats
            
            # 进行方差分析
            groups = {}
            for record in records:
                age_group = record.get('template', '自定义')
                if age_group not in groups:
                    groups[age_group] = []
                
                development_score = sum(record['evaluation_scores'].values()) / len(record['evaluation_scores'])
                groups[age_group].append(development_score)
            
            if len(groups) >= 2:
                group_values = list(groups.values())
                f_stat, p_value = stats.f_oneway(*group_values)
                
                st.write(f"**单因素方差分析结果**:")
                st.write(f"- F统计量: {f_stat:.3f}")
                st.write(f"- p值: {p_value:.3f}")
                
                if p_value < 0.05:
                    st.success("✅ 不同年龄段间发展差异具有统计学意义 (p < 0.05)")
                else:
                    st.info("ℹ️ 不同年龄段间发展差异无统计学意义 (p ≥ 0.05)")
        
        except ImportError:
            st.info("💡 安装scipy模块可启用统计学分析功能")

elif page == "观察记录管理":
    st.header("📚 观察记录管理")
    
    records = st.session_state.observation_records
    
    if not records:
        st.info("📝 暂无观察记录")
        st.stop()
    
    st.subheader(f"📊 共有 {len(records)} 条发展观察记录")
    
    # 高级筛选选项
    col_filter1, col_filter2, col_filter3, col_filter4 = st.columns(4)
    
    with col_filter1:
        age_filter = st.selectbox(
            "按年龄段筛选", 
            ["全部"] + list(set([r.get('template', '自定义') for r in records]))
        )
    
    with col_filter2:
        context_filter = st.selectbox(
            "按观察情境筛选",
            ["全部"] + list(set([r['scene'] for r in records]))
        )
    
    with col_filter3:
        development_filter = st.selectbox(
            "按发展水平筛选",
            ["全部", "优秀 (4.5-5分)", "良好 (4-4.5分)", "一般 (3-4分)", "需关注 (1-3分)"]
        )
    
    with col_filter4:
        sort_by = st.selectbox(
            "排序方式", 
            ["时间倒序", "时间正序", "综合发展指数", "语言发展水平", "社交能力水平"]
        )
    
    # 应用筛选
    filtered_records = records
    
    if age_filter != "全部":
        filtered_records = [r for r in filtered_records if r.get('template', '自定义') == age_filter]
    
    if context_filter != "全部":
        filtered_records = [r for r in filtered_records if r['scene'] == context_filter]
    
    if development_filter != "全部":
        def get_development_score(record):
            return sum(record['evaluation_scores'].values()) / len(record['evaluation_scores'])
        
        if development_filter == "优秀 (4.5-5分)":
            filtered_records = [r for r in filtered_records if get_development_score(r) >= 4.5]
        elif development_filter == "良好 (4-4.5分)":
            filtered_records = [r for r in filtered_records if 4.0 <= get_development_score(r) < 4.5]
        elif development_filter == "一般 (3-4分)":
            filtered_records = [r for r in filtered_records if 3.0 <= get_development_score(r) < 4.0]
        elif development_filter == "需关注 (1-3分)":
            filtered_records = [r for r in filtered_records if get_development_score(r) < 3.0]
    
    # 应用排序
    if sort_by == "时间正序":
        filtered_records = sorted(filtered_records, key=lambda x: x['timestamp'])
    elif sort_by == "综合发展指数":
        filtered_records = sorted(filtered_records, 
            key=lambda x: sum(x['evaluation_scores'].values()) / len(x['evaluation_scores']), 
            reverse=True)
    elif sort_by == "语言发展水平":
        filtered_records = sorted(filtered_records, 
            key=lambda x: x['evaluation_scores']['语言沟通发展'], reverse=True)
    elif sort_by == "社交能力水平":
        filtered_records = sorted(filtered_records, 
            key=lambda x: x['evaluation_scores']['社交互动能力'], reverse=True)
    else:  # 时间倒序
        filtered_records = sorted(filtered_records, key=lambda x: x['timestamp'], reverse=True)
    
    st.write(f"筛选后记录数: {len(filtered_records)}")
    
    # 记录列表显示
    for i, record in enumerate(filtered_records):
        
        # 计算综合发展指数
        development_index = sum(record['evaluation_scores'].values()) / len(record['evaluation_scores'])
        
        development_label = ""
        if development_index >= 4.5:
            development_label = "🌟 优秀"
        elif development_index >= 4.0:
            development_label = "👍 良好"
        elif development_index >= 3.0:
            development_label = "📈 一般"
        else:
            development_label = "🔍 需关注"
        
        template_info = f" - {record.get('template', '自定义')}" if record.get('template') else ""
        
        with st.expander(f"🌟 {record['observation_id']}{template_info} - {record['scene']} - {development_label} ({record['timestamp'].strftime('%Y-%m-%d %H:%M')})"):
            
            col1, col2, col3 = st.columns(3)
            
            with col1:
                st.write("**📋 观察基本信息:**")
                if record.get('template'):
                    st.write(f"• 年龄发展阶段: {record['template']}")
                st.write(f"• 观察情境: {record['scene']}")
                st.write(f"• 观察活动: {record.get('activity', '未指定')}")
                st.write(f"• 情境触发: {record.get('trigger', '未指定')}")
                st.write(f"• 观察时间: {record['timestamp'].strftime('%Y-%m-%d %H:%M:%S')}")
                
                if record.get('child_profile'):
                    st.write("**👶 儿童发展配置:**")
                    profile = record['child_profile']
                    st.write(f"• 发展阶段: {profile.get('stage_characteristics', 'N/A')}")
                    st.write(f"• 语言发展: {profile.get('language_development', 'N/A')}/5")
                    st.write(f"• 社交技能: {profile.get('social_skills', 'N/A')}/5")
                    st.write(f"• 认知能力: {profile.get('cognitive_ability', 'N/A')}/5")
                    st.write(f"• 典型兴趣: {profile.get('typical_interests', 'N/A')}")
            
            with col2:
                st.write("**📊 发展评估得分:**")
                
                scores = record['evaluation_scores']
                
                # 核心发展能力
                st.write("*核心发展能力:*")
                language_score = scores['语言沟通发展']
                social_score = scores['社交互动能力']
                cognitive_score = scores['认知学习能力']
                
                if language_score >= 4.5:
                    st.success(f"• 语言沟通发展: {language_score}/5 (优秀)")
                elif language_score >= 4.0:
                    st.info(f"• 语言沟通发展: {language_score}/5 (良好)")
                elif language_score >= 3.0:
                    st.warning(f"• 语言沟通发展: {language_score}/5 (一般)")
                else:
                    st.error(f"• 语言沟通发展: {language_score}/5 (需关注)")
                
                if social_score >= 4.5:
                    st.success(f"• 社交互动能力: {social_score}/5 (优秀)")
                elif social_score >= 4.0:
                    st.info(f"• 社交互动能力: {social_score}/5 (良好)")
                elif social_score >= 3.0:
                    st.warning(f"• 社交互动能力: {social_score}/5 (一般)")
                else:
                    st.error(f"• 社交互动能力: {social_score}/5 (需关注)")
                
                if cognitive_score >= 4.5:
                    st.success(f"• 认知学习能力: {cognitive_score}/5 (优秀)")
                elif cognitive_score >= 4.0:
                    st.info(f"• 认知学习能力: {cognitive_score}/5 (良好)")
                elif cognitive_score >= 3.0:
                    st.warning(f"• 认知学习能力: {cognitive_score}/5 (一般)")
                else:
                    st.error(f"• 认知学习能力: {cognitive_score}/5 (需关注)")
                
                # 其他发展能力
                st.write("*其他发展能力:*")
                st.write(f"• 情绪调节发展: {scores['情绪调节发展']}/5")
                st.write(f"• 运动技能发展: {scores['运动技能发展']}/5")
                st.write(f"• 独立自理能力: {scores['独立自理能力']}/5")
                
                st.write(f"**综合发展指数: {development_index:.2f}/5**")
            
            with col3:
                st.write("**🔍 发展观察记录:**")
                if 'developmental_observations' in record and record['developmental_observations']:
                    for category, observations in record['developmental_observations'].items():
                        if observations:
                            st.write(f"*{category}:*")
                            for obs in observations:
                                st.write(f"• {obs}")
                else:
                    st.write("暂无特殊发展观察记录")
                
                if record.get('notes'):
                    st.write(f"**📝 备注:** {record['notes']}")
            
            # 对话记录
            st.write("**💬 行为观察对话记录:**")
            dialogue_lines = record['dialogue'].split('\n')
            dialogue_text = '\n'.join([line for line in dialogue_lines if line.strip() and ':' in line])
            
            unique_key = f"development_dialogue_{i}_{record['observation_id']}_{record['timestamp'].strftime('%Y%m%d_%H%M%S')}"
            st.text_area("", dialogue_text, height=200, key=unique_key)
            
            # 快速操作按钮
            col_btn1, col_btn2, col_btn3 = st.columns(3)
            
            with col_btn1:
                if st.button(f"📋 生成发展报告", key=f"report_{record['observation_id']}"):
                    st.info("发展报告生成功能开发中...")
            
            with col_btn2:
                if st.button(f"📈 发展趋势", key=f"trend_{record['observation_id']}"):
                    st.info("发展趋势分析功能开发中...")
            
            with col_btn3:
                if st.button(f"🔄 重复观察", key=f"repeat_{record['observation_id']}"):
                    st.info("重复观察功能开发中...")

elif page == "📊 发展报告中心":
    st.header("📊 发展报告中心")
    st.markdown("基于儿童发展理论生成专业的发展评估报告和研究数据")
    
    records = st.session_state.observation_records
    
    if not records:
        st.warning("📊 暂无观察数据，请先进行发展观察")
        st.stop()
    
    st.success(f"📊 当前共有 {len(records)} 条发展观察记录可生成报告")
    
    # 报告类型选择
    st.subheader("📋 选择报告类型")
    
    col1, col2 = st.columns(2)
    
    with col1:
        st.write("### 📄 标准发展报告")
        
        # 基础CSV报告
        if st.button("📊 下载基础发展数据 (CSV)", use_container_width=True):
            df_export = []
            for record in records:
                export_row = {
                    '观察ID': record['observation_id'],
                    '观察时间': record['timestamp'].strftime('%Y-%m-%d %H:%M:%S'),
                    '年龄发展阶段': record.get('template', '自定义'),
                    '观察情境': record['scene'],
                    '观察活动': record.get('activity', ''),
                    '情境触发': record.get('trigger', ''),
                    '语言沟通发展': record['evaluation_scores']['语言沟通发展'],
                    '社交互动能力': record['evaluation_scores']['社交互动能力'],
                    '认知学习能力': record['evaluation_scores']['认知学习能力'],
                    '情绪调节发展': record['evaluation_scores']['情绪调节发展'],
                    '运动技能发展': record['evaluation_scores']['运动技能发展'],
                    '独立自理能力': record['evaluation_scores']['独立自理能力'],
                    '备注': record.get('notes', '')
                }
                
                # 添加儿童发展特征
                if record.get('child_profile'):
                    profile = record['child_profile']
                    export_row.update({
                        '发展阶段特征': profile.get('stage_characteristics', ''),
                        '发展重点': profile.get('development_focus', ''),
                        '语言发展配置': profile.get('language_development', ''),
                        '社交技能配置': profile.get('social_skills', ''),
                        '认知能力配置': profile.get('cognitive_ability', ''),
                        '情绪调节配置': profile.get('emotional_regulation', ''),
                        '运动技能配置': profile.get('motor_skills', ''),
                        '独立性配置': profile.get('independence_level', ''),
                        '典型兴趣描述': profile.get('typical_interests', '')
                    })
                
                # 计算综合发展指数
                development_index = sum(record['evaluation_scores'].values()) / len(record['evaluation_scores'])
                export_row['综合发展指数'] = round(development_index, 2)
                
                df_export.append(export_row)
            
            df = pd.DataFrame(df_export)
            csv = df.to_csv(index=False, encoding='utf-8-sig')
            
            st.download_button(
                label="⬇️ 下载发展评估数据",
                data=csv,
                file_name=f"children_development_assessment_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
                mime='text/csv'
            )
        
        # 对话记录下载
        if st.button("💬 下载行为观察记录 (TXT)", use_container_width=True):
            observation_content = []
            observation_content.append("=" * 70)
            observation_content.append("正常儿童成长发育行为观察记录")
            observation_content.append("基于儿童发展心理学理论 | 多元智能评估框架")
            observation_content.append("=" * 70)
            observation_content.append(f"生成时间: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
            observation_content.append(f"观察记录总数: {len(records)}")
            observation_content.append("=" * 70)
            observation_content.append("")
            
            for i, record in enumerate(records, 1):
                development_index = sum(record['evaluation_scores'].values()) / len(record['evaluation_scores'])
                
                observation_content.append(f"\n【发展观察 {i}】")
                observation_content.append(f"观察ID: {record['observation_id']}")
                observation_content.append(f"观察时间: {record['timestamp'].strftime('%Y-%m-%d %H:%M:%S')}")
                observation_content.append(f"年龄发展阶段: {record.get('template', '自定义')}")
                observation_content.append(f"观察情境: {record['scene']}")
                observation_content.append(f"观察活动: {record.get('activity', '未指定')}")
                observation_content.append(f"情境触发: {record.get('trigger', '未指定')}")
                
                if record.get('child_profile'):
                    profile = record['child_profile']
                    observation_content.append(f"发展阶段特征: {profile.get('stage_characteristics', '')}")
                    observation_content.append(f"发展重点: {profile.get('development_focus', '')}")
                
                observation_content.append(f"综合发展指数: {development_index:.2f}/5.0")
                observation_content.append("-" * 50)
                
                observation_content.append("发展评估得分:")
                for metric, score in record['evaluation_scores'].items():
                    observation_content.append(f"  • {metric}: {score}/5.0")
                
                if 'developmental_observations' in record and record['developmental_observations']:
                    observation_content.append("发展观察要点:")
                    for category, observations in record['developmental_observations'].items():
                        if observations:
                            observation_content.append(f"  {category}: {', '.join(observations)}")
                
                observation_content.append("行为观察对话:")
                observation_content.append(record['dialogue'])
                observation_content.append("-" * 50)
                observation_content.append("")
            
            observation_text = '\n'.join(observation_content)
            
            st.download_button(
                label="⬇️ 下载行为观察记录",
                data=observation_text.encode('utf-8'),
                file_name=f"children_development_observations_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.txt",
                mime='text/plain'
            )
        
        # JSON完整数据
        if st.button("🔧 下载完整发展数据 (JSON)", use_container_width=True):
            json_data = {
                'development_assessment_report': {
                    'report_metadata': {
                        'generation_time': datetime.datetime.now().isoformat(),
                        'assessment_theory': '儿童发展心理学理论',
                        'evaluation_framework': '基于多元智能和发展里程碑评估',
                        'total_observations': len(records),
                        'platform_version': '专业版 v1.0'
                    },
                    'development_summary': generate_development_analysis(records),
                    'detailed_observations': []
                }
            }
            
            for record in records:
                development_record = record.copy()
                development_record['timestamp'] = record['timestamp'].isoformat()
                
                # 添加计算字段
                development_index = sum(record['evaluation_scores'].values()) / len(record['evaluation_scores'])
                development_record['development_index'] = round(development_index, 2)
                
                json_data['development_assessment_report']['detailed_observations'].append(development_record)
            
            json_str = json.dumps(json_data, ensure_ascii=False, indent=2)
            
            st.download_button(
                label="⬇️ 下载完整发展数据",
                data=json_str.encode('utf-8'),
                file_name=f"children_development_complete_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.json",
                mime='application/json'
            )
    
    with col2:
        st.write("### 📈 专业分析报告")
        
        # 生成发展分析报告
        if st.button("📊 生成发展统计分析", use_container_width=True):
            with st.spinner("正在生成发展分析报告..."):
                analysis = generate_development_analysis(records)
            
            st.success("✅ 发展分析报告生成完成！")
            
            # 显示分析预览
            with st.expander("📋 发展分析报告预览", expanded=True):
                if analysis.get('发展观察概况'):
                    st.write("**发展观察概况:**")
                    for key, value in analysis['发展观察概况'].items():
                        st.write(f"- {key}: {value}")
                
                if analysis.get('整体发展表现'):
                    st.write("**整体发展表现:**")
                    for key, value in analysis['整体发展表现'].items():
                        st.write(f"- {key}: {value}")
                
                if analysis.get('发展建议与指导'):
                    st.write("**发展建议与指导:**")
                    for recommendation in analysis['发展建议与指导']:
                        st.write(f"- {recommendation}")
            
            # 提供分析报告下载
            analysis_json = json.dumps(analysis, ensure_ascii=False, indent=2)
            st.download_button(
                label="⬇️ 下载发展分析报告 (JSON)",
                data=analysis_json.encode('utf-8'),
                file_name=f"children_development_analysis_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.json",
                mime='application/json'
            )
        
        # Excel专业报告
        if EXCEL_AVAILABLE:
            if st.button("📋 生成专业Excel报告", use_container_width=True):
                with st.spinner("正在生成专业Excel报告..."):
                    analysis = generate_development_analysis(records)
                    excel_data = create_development_excel_report(records, analysis)
                
                if excel_data:
                    st.success("✅ 专业Excel报告生成完成！")
                    
                    st.download_button(
                        label="⬇️ 下载专业Excel报告",
                        data=excel_data,
                        file_name=f"children_development_professional_report_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                        mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                    )
                else:
                    st.error("❌ Excel报告生成失败，请尝试其他格式")
        else:
            st.info("💡 Excel报告功能需要安装openpyxl模块")
            st.code("pip install openpyxl")
            
            # 替代详细报告
            if st.button("📊 生成详细文本报告", use_container_width=True):
                with st.spinner("正在生成详细报告..."):
                    analysis = generate_development_analysis(records)
                
                # 创建详细文本报告
                detailed_report = []
                detailed_report.append("正常儿童成长发育详细报告")
                detailed_report.append("=" * 50)
                detailed_report.append(f"报告生成时间: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
                detailed_report.append(f"理论基础: 儿童发展心理学")
                detailed_report.append(f"评估框架: 多元智能理论 + 发展里程碑")
                detailed_report.append("")
                
                # 添加发展观察概况
                detailed_report.append("一、发展观察概况")
                detailed_report.append("-" * 30)
                for key, value in analysis.get('发展观察概况', {}).items():
                    detailed_report.append(f"{key}: {value}")
                detailed_report.append("")
                
                # 添加整体表现
                detailed_report.append("二、整体发展表现")
                detailed_report.append("-" * 30)
                for key, value in analysis.get('整体发展表现', {}).items():
                    detailed_report.append(f"{key}: {value}")
                detailed_report.append("")
                
                # 添加年龄段分析
                if analysis.get('年龄段分析'):
                    detailed_report.append("三、年龄段发展分析")
                    detailed_report.append("-" * 30)
                    for age_group, stats in analysis['年龄段分析'].items():
                        detailed_report.append(f"\n{age_group} (n={stats['观察次数']}):")
                        detailed_report.append(f"  - 语言沟通发展: {stats['语言发展得分_均值']:.2f} ± {stats['语言发展得分_标准差']:.2f}")
                        detailed_report.append(f"  - 社交互动能力: {stats['社交能力得分_均值']:.2f} ± {stats['社交能力得分_标准差']:.2f}")
                        detailed_report.append(f"  - 认知学习能力: {stats['认知学习得分_均值']:.2f} ± {stats['认知学习得分_标准差']:.2f}")
                    detailed_report.append("")
                
                # 添加发展建议
                detailed_report.append("四、发展建议与指导")
                detailed_report.append("-" * 30)
                for i, recommendation in enumerate(analysis.get('发展建议与指导', []), 1):
                    detailed_report.append(f"{i}. {recommendation}")
                detailed_report.append("")
                
                # 添加个案明细
                detailed_report.append("五、个案观察明细")
                detailed_report.append("-" * 30)
                for i, record in enumerate(records, 1):
                    development_index = sum(record['evaluation_scores'].values()) / len(record['evaluation_scores'])
                    
                    detailed_report.append(f"\n个案 {i}: {record['observation_id']}")
                    detailed_report.append(f"  观察时间: {record['timestamp'].strftime('%Y-%m-%d %H:%M')}")
                    detailed_report.append(f"  年龄阶段: {record.get('template', '自定义')}")
                    detailed_report.append(f"  观察情境: {record['scene']}")
                    detailed_report.append(f"  综合发展指数: {development_index:.2f}/5.0")
                    
                    development_level = "优秀" if development_index >= 4.5 else "良好" if development_index >= 4.0 else "一般" if development_index >= 3.0 else "需关注"
                    detailed_report.append(f"  发展水平判断: {development_level}")
                
                report_content = '\n'.join(detailed_report)
                
                st.success("✅ 详细文本报告生成完成！")
                
                st.download_button(
                    label="⬇️ 下载详细文本报告",
                    data=report_content.encode('utf-8'),
                    file_name=f"children_development_detailed_report_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.txt",
                    mime='text/plain'
                )
        
        # 研究数据包
        if st.button("📦 生成完整研究数据包", use_container_width=True, type="primary"):
            with st.spinner("正在生成完整研究数据包..."):
                # 生成分析
                analysis = generate_development_analysis(records)
                
                # 创建基础CSV
                df_basic = []
                for record in records:
                    development_index = sum(record['evaluation_scores'].values()) / len(record['evaluation_scores'])
                    
                    row = {
                        '观察ID': record['observation_id'],
                        '观察时间': record['timestamp'].strftime('%Y-%m-%d %H:%M:%S'),
                        '年龄发展阶段': record.get('template', '自定义'),
                        '观察情境': record['scene'],
                        '观察活动': record.get('activity', ''),
                        '情境触发': record.get('trigger', ''),
                        '语言沟通发展': record['evaluation_scores']['语言沟通发展'],
                        '社交互动能力': record['evaluation_scores']['社交互动能力'],
                        '认知学习能力': record['evaluation_scores']['认知学习能力'],
                        '情绪调节发展': record['evaluation_scores']['情绪调节发展'],
                        '运动技能发展': record['evaluation_scores']['运动技能发展'],
                        '独立自理能力': record['evaluation_scores']['独立自理能力'],
                        '综合发展指数': round(development_index, 2),
                        '备注': record.get('notes', '')
                    }
                    
                    # 添加发展配置
                    if record.get('child_profile'):
                        profile = record['child_profile']
                        row.update({
                            '发展阶段特征': profile.get('stage_characteristics', ''),
                            '发展重点': profile.get('development_focus', ''),
                            '语言发展配置': profile.get('language_development', ''),
                            '社交技能配置': profile.get('social_skills', ''),
                            '认知能力配置': profile.get('cognitive_ability', ''),
                            '情绪调节配置': profile.get('emotional_regulation', ''),
                            '运动技能配置': profile.get('motor_skills', ''),
                            '独立性配置': profile.get('independence_level', ''),
                            '典型兴趣': profile.get('typical_interests', '')
                        })
                    
                    df_basic.append(row)
                
                # 创建完整数据JSON
                complete_data = {
                    'research_metadata': {
                        'generation_time': datetime.datetime.now().isoformat(),
                        'assessment_theory': '儿童发展心理学理论',
                        'evaluation_framework': '基于多元智能和发展里程碑评估',
                        'total_observations': len(records),
                        'platform_version': '专业版 v1.0'
                    },
                    'development_analysis': analysis,
                    'raw_observation_data': []
                }
                
                for record in records:
                    research_record = record.copy()
                    research_record['timestamp'] = record['timestamp'].isoformat()
                    complete_data['raw_observation_data'].append(research_record)
            
            st.success("✅ 完整研究数据包生成完成！")
            
            # 创建ZIP文件
            zip_buffer = io.BytesIO()
            with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
                
                # 1. 基础发展数据
                csv_data = pd.DataFrame(df_basic).to_csv(index=False, encoding='utf-8-sig')
                zip_file.writestr("01_基础发展数据.csv", csv_data)
                
                # 2. Excel专业报告（如果可用）
                excel_data = None
                if EXCEL_AVAILABLE:
                    try:
                        excel_data = create_development_excel_report(records, analysis)
                        if excel_data:
                            zip_file.writestr("02_专业分析报告.xlsx", excel_data)
                    except Exception as e:
                        pass  # 如果Excel生成失败，跳过
                
                if not excel_data:
                    # 生成详细文本报告作为替代
                    detailed_report = []
                    detailed_report.append("正常儿童成长发育专业报告")
                    detailed_report.append("基于儿童发展心理学理论")
                    detailed_report.append("=" * 60)
                    detailed_report.append(f"报告生成时间: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
                    detailed_report.append(f"理论基础: 儿童发展心理学 + 多元智能理论")
                    detailed_report.append(f"评估框架: 发展里程碑 + 行为观察")
                    detailed_report.append("")
                    
                    # 发展概况
                    detailed_report.append("一、发展观察概况")
                    detailed_report.append("-" * 40)
                    for key, value in analysis.get('发展观察概况', {}).items():
                        detailed_report.append(f"{key}: {value}")
                    
                    # 整体表现
                    detailed_report.append("\n二、整体发展表现")
                    detailed_report.append("-" * 40)
                    for key, value in analysis.get('整体发展表现', {}).items():
                        detailed_report.append(f"{key}: {value}")
                    
                    # 发展建议
                    detailed_report.append("\n三、发展建议与指导")
                    detailed_report.append("-" * 40)
                    for i, recommendation in enumerate(analysis.get('发展建议与指导', []), 1):
                        detailed_report.append(f"{i}. {recommendation}")
                    
                    report_text = '\n'.join(detailed_report)
                    zip_file.writestr("02_专业分析报告.txt", report_text.encode('utf-8'))
                
                # 3. 完整研究数据
                json_str = json.dumps(complete_data, ensure_ascii=False, indent=2)
                zip_file.writestr("03_完整研究数据.json", json_str.encode('utf-8'))
                
                # 4. 行为观察记录
                observation_content = []
                observation_content.append("正常儿童成长发育行为观察记录集")
                observation_content.append("=" * 50)
                observation_content.append(f"观察理论: 儿童发展心理学")
                observation_content.append(f"记录时间: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
                observation_content.append(f"观察案例数: {len(records)}")
                observation_content.append("")
                
                for i, record in enumerate(records, 1):
                    observation_content.append(f"\n【观察记录 {i}】")
                    observation_content.append(f"观察ID: {record['observation_id']}")
                    observation_content.append(f"时间: {record['timestamp'].strftime('%Y-%m-%d %H:%M:%S')}")
                    observation_content.append(f"年龄阶段: {record.get('template', '自定义')}")
                    observation_content.append(f"情境: {record['scene']}")
                    observation_content.append(f"活动: {record.get('activity', '')}")
                    observation_content.append(f"触发因素: {record.get('trigger', '')}")
                    observation_content.append("-" * 30)
                    observation_content.append("行为观察对话:")
                    observation_content.append(record['dialogue'])
                    observation_content.append("-" * 30)
                    observation_content.append("")
                
                obs_text = '\n'.join(observation_content)
                zip_file.writestr("04_行为观察记录.txt", obs_text.encode('utf-8'))
                
                # 5. 研究说明文档
                readme_content = f"""
正常儿童成长发育AI观察平台 - 专业版
研究数据包说明文档
======================================

生成时间: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
观察记录数: {len(records)}
理论基础: 儿童发展心理学 + 多元智能理论

文件说明:
--------
1. 01_基础发展数据.csv
   - 包含所有观察的核心发展数据
   - 适用于统计分析和数据可视化
   - 包含发展配置和评估得分

"""
                if excel_data:
                    readme_content += """2. 02_专业分析报告.xlsx
   - 多工作表Excel专业报告
   - 包含统计分析、图表和发展解释
   - 适用于教育报告和研究发表
"""
                else:
                    readme_content += """2. 02_专业分析报告.txt
   - 详细的文本格式专业报告
   - 包含统计分析和发展解释
   - 注意: Excel功能不可用，如需Excel格式请安装openpyxl
"""

                readme_content += """
3. 03_完整研究数据.json
   - 包含所有原始数据和元数据
   - 适用于程序处理和深度分析
   - 包含完整的发展观察记录

4. 04_行为观察记录.txt
   - 所有观察的对话记录
   - 用于质性分析和行为模式研究
   - 符合发展观察记录格式

5. README.txt
   - 本说明文档

评估指标说明:
-----------
所有评估得分采用1-5分制，其中:
- 1分: 需要专业支持/发展滞后
- 2分: 需要关注/发展略慢
- 3分: 一般发展/年龄适宜
- 4分: 良好发展/超过平均
- 5分: 优秀发展/发展超前

核心发展评估:
- 语言沟通发展: 表达清晰度、理解能力、交流主动性、语用技能
- 社交互动能力: 友谊建立、合作分享、冲突解决、群体适应  
- 认知学习能力: 注意专注、记忆学习、逻辑思维、创造想象

其他发展评估:
- 情绪调节发展: 情绪认知、情绪表达、自我调节、同理关怀
- 运动技能发展: 大运动、精细运动、运动规划、运动兴趣
- 独立自理能力: 生活自理、任务执行、问题解决、责任意识

使用建议:
--------
1. 教育应用:
   - 使用基础数据进行发展追踪
   - 参考专业报告制定教育计划
   - 结合行为观察进行个案分析

2. 研究应用:
   - 使用完整数据进行发展研究
   - 横向对比和纵向追踪研究
   - 教育干预效果评估

3. 家庭应用:
   - 了解儿童发展状况
   - 制定家庭教育策略
   - 寻找发展优势和需要支持的领域

技术支持:
--------
- 如需Excel功能，请安装: pip install openpyxl
- 数据分析建议使用: pandas, numpy, scipy
- 可视化建议使用: matplotlib, plotly

参考理论:
--------
- 皮亚杰认知发展理论
- 维果茨基社会文化理论
- 加德纳多元智能理论
- 埃里克森心理社会发展理论
- 布朗芬布伦纳生态系统理论

质量保证:
--------
本平台基于最新的儿童发展心理学理论设计，
所有评估指标均参考国际认可的发展里程碑，
确保观察结果的科学性和可靠性。

版权声明:
--------
本数据包仅供教育研究和儿童发展支持使用，
请遵循相关伦理规范和数据保护法规。
"""
                zip_file.writestr("README.txt", readme_content)
            
            zip_buffer.seek(0)
            
            st.download_button(
                label="⬇️ 下载完整研究数据包 (ZIP)",
                data=zip_buffer.getvalue(),
                file_name=f"children_development_research_package_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.zip",
                mime='application/zip'
            )
    
    # 数据统计概览
    st.subheader("📈 数据统计概览")
    
    col_stat1, col_stat2, col_stat3, col_stat4 = st.columns(4)
    
    with col_stat1:
        st.metric("发展观察总数", len(records))
    
    with col_stat2:
        age_groups = [r.get('template', '自定义') for r in records]
        unique_age_groups = len(set(age_groups))
        st.metric("涉及年龄段类型", unique_age_groups)
    
    with col_stat3:
        contexts = [r['scene'] for r in records]
        unique_contexts = len(set(contexts))
        st.metric("涉及观察情境", unique_contexts)
    
    with col_stat4:
        if records:
            time_span = (max(r['timestamp'] for r in records) - min(r['timestamp'] for r in records)).days
            st.metric("观察时间跨度(天)", time_span)
    
    # 发展数据预览
    st.subheader("📊 发展数据预览")
    
    preview_data = []
    for record in records[:10]:
        # 计算综合发展指数
        development_index = sum(record['evaluation_scores'].values()) / len(record['evaluation_scores'])
        
        development_level = "优秀" if development_index >= 4.5 else "良好" if development_index >= 4.0 else "一般" if development_index >= 3.0 else "需关注"
        
        preview_data.append({
            '观察ID': record['observation_id'][:25] + '...' if len(record['observation_id']) > 25 else record['observation_id'],
            '时间': record['timestamp'].strftime('%m-%d %H:%M'),
            '年龄段': record.get('template', '自定义')[:8] + '...' if len(record.get('template', '自定义')) > 8 else record.get('template', '自定义'),
            '观察情境': record['scene'].replace('日常生活', '日常'),
            '发展指数': f"{development_index:.2f}",
            '发展水平': development_level
        })
    
    df_preview = pd.DataFrame(preview_data)
    st.dataframe(df_preview, use_container_width=True)
    
    if len(records) > 10:
        st.info(f"显示前10条记录，共{len(records)}条。完整数据请通过上方下载功能获取。")

# 侧边栏额外信息
st.sidebar.markdown("---")
st.sidebar.markdown("### 📈 观察统计")
if st.session_state.observation_records:
    st.sidebar.metric("观察总数", len(st.session_state.observation_records))
    recent_record = st.session_state.observation_records[-1]
    st.sidebar.write(f"最近观察: {recent_record['timestamp'].strftime('%m-%d %H:%M')}")
    
    # 显示年龄段分布
    age_groups = [r.get('template', '自定义') for r in st.session_state.observation_records]
    age_counts = pd.Series(age_groups).value_counts()
    st.sidebar.write("**年龄段分布:**")
    for age_group, count in age_counts.items():
        short_name = age_group.split('期')[0] if '期' in age_group else age_group
        st.sidebar.write(f"- {short_name}: {count}")
    
    # 发展水平统计
    all_development_scores = []
    for r in st.session_state.observation_records:
        development_score = sum(r['evaluation_scores'].values()) / len(r['evaluation_scores'])
        all_development_scores.append(development_score)
    
    avg_development = np.mean(all_development_scores)
    st.sidebar.metric("平均发展指数", f"{avg_development:.2f}/5")
    
    if avg_development >= 4.5:
        st.sidebar.success("整体发展水平优秀")
    elif avg_development >= 4.0:
        st.sidebar.info("整体发展水平良好")
    elif avg_development >= 3.0:
        st.sidebar.warning("整体发展水平一般")
    else:
        st.sidebar.error("整体发展需要关注")
        
else:
    st.sidebar.write("暂无观察数据")

st.sidebar.markdown("---")
st.sidebar.markdown("### ℹ️ 理论基础说明")
st.sidebar.markdown("""
**理论基础**: 儿童发展心理学

**核心发展领域**:
- 语言沟通发展
- 社交互动能力
- 认知学习能力

**其他发展领域**:
- 情绪调节发展
- 运动技能发展
- 独立自理能力

**参考理论**:
- 皮亚杰认知发展理论
- 维果茨基社会文化理论  
- 加德纳多元智能理论
- 埃里克森心理社会发展理论

**发展水平分级**:
1. 需要支持（1-2分）
2. 需要关注（2-3分）
3. 一般发展（3-4分）
4. 良好发展（4-5分）
5. 优秀发展（5分）
""")

if not EXCEL_AVAILABLE:
    st.sidebar.markdown("---")
    st.sidebar.markdown("### 💡 功能提示")
    st.sidebar.warning("Excel功能不可用")
    st.sidebar.markdown("要启用专业Excel报告功能，请运行：")
    st.sidebar.code("pip install openpyxl")
    st.sidebar.markdown("目前可使用CSV和JSON格式导出发展数据。")

st.sidebar.markdown("---")
st.sidebar.markdown("### ⚠️ API限制说明")
st.sidebar.markdown("""
**当前API限制**: 每分钟3次请求

**对观察影响**:
- 快速发育观察: 立即完成
- 批量发展研究: 每个观察间隔25秒

**建议**:
- 批量研究选择适当规模
- 可分批次进行大样本研究
- 保持网络连接稳定
""")

st.sidebar.markdown("---")
st.sidebar.markdown("### 🌟 教育应用建议")
st.sidebar.markdown("""
**家庭教育**:
- 了解儿童发展特点
- 制定个性化教育方案
- 发现儿童优势和潜能

**幼儿园应用**:
- 观察儿童发展状况
- 设计适龄教育活动
- 与家长沟通发展情况

**研究用途**:
- 儿童发展规律研究
- 教育干预效果评估
- 发展里程碑验证

**注意事项**:
- 本工具仅供发展参考
- 不能替代专业发展评估
- 建议结合多种观察方法
""")

st.sidebar.markdown("---")
st.sidebar.markdown("### 🔄 数据管理")
if st.sidebar.button("🗑️ 清空所有观察数据"):
    if st.sidebar.checkbox("确认清空（此操作不可恢复）"):
        st.session_state.observation_records = []
        st.sidebar.success("✅ 观察数据已清空")
        st.rerun()

# 页脚
st.markdown("---")
st.markdown("""
### 🌟 平台特点

**🎓 科学理论**: 严格遵循儿童发展心理学理论，参考国际发展里程碑  
**🔍 专业观察**: 基于多元智能理论的全面发展评估框架  
**📊 数据驱动**: 生成符合教育研究要求的专业发展报告  
**👨‍👩‍👧‍👦 个性化**: 支持个性化配置和定制化观察设计

**💡 使用提示**: 
- 建议先进行'快速发育观察'熟悉平台功能
- 使用'批量发展研究'获取统计学有效的数据
- 在'📊 发展报告中心'下载完整的专业报告
- 所有观察结果仅供教育参考，建议结合多种评估方法

**⚠️ 重要声明**: 本平台仅供教育研究和儿童发展支持使用，不能替代专业的儿童发展评估。
""")

st.markdown("*基于儿童发展心理学理论 | 参考多元智能和发展里程碑 | 专业版 v1.0*")