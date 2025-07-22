"""Excel导出模块"""
import io
from ..config import EXCEL_AVAILABLE

if EXCEL_AVAILABLE:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment


def create_excel_workbook():
    """创建一个新的Excel工作簿"""
    if not EXCEL_AVAILABLE:
        return None
    
    workbook = Workbook()
    workbook.remove(workbook.active)  # 移除默认工作表
    return workbook


def apply_excel_styles(workbook, header_color="366092", header_font_color="FFFFFF"):
    """
    应用Excel样式
    
    Args:
        workbook: openpyxl工作簿对象
        header_color: 标题行背景色
        header_font_color: 标题行字体颜色
    """
    if not EXCEL_AVAILABLE:
        return
    
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if cell.row == 1:  # 标题行
                    cell.font = Font(bold=True, size=12, color=header_font_color)
                    cell.fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
                cell.alignment = Alignment(wrap_text=True, vertical="top")


def export_to_excel(workbook):
    """
    将工作簿导出为字节流
    
    Args:
        workbook: openpyxl工作簿对象
    
    Returns:
        bytes: Excel文件的字节数据
    """
    if not EXCEL_AVAILABLE or not workbook:
        return None
    
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return output.getvalue()