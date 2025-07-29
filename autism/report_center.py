"""孤独症平台报告中心页面 - 支持DSM-5和ABC双标准"""
import streamlit as st
import pandas as pd
import datetime
import json
import numpy as np

from common.config import EXCEL_AVAILABLE
from common.exporters import (
    export_to_csv, export_to_json, export_to_excel, 
    create_excel_workbook, apply_excel_styles, create_zip_package
)
from common.exporters.text_exporter import export_to_text
from .analyzer import generate_clinical_analysis, prepare_clinical_export_data


def page_report_center():
    """报告中心页面 - 支持双标准"""
    st.header("📊 临床评估报告中心")
    st.markdown("基于ABC量表和DSM-5标准生成专业评估报告和研究数据")
    
    records = st.session_state.experiment_records
    
    if not records:
        st.warning("📊 暂无评估数据，请先进行临床评估")
        st.stop()
    
    # 分析记录中的评估标准分布
    abc_records = [r for r in records if r.get('assessment_standard', 'ABC') == 'ABC']
    dsm5_records = [r for r in records if r.get('assessment_standard', 'ABC') == 'DSM5']
    
    st.success(f"📊 当前共有 {len(records)} 条评估记录可生成报告")
    col_std1, col_std2 = st.columns(2)
    with col_std1:
        st.info(f"📋 ABC量表评估: {len(abc_records)} 条")
    with col_std2:
        st.info(f"📋 DSM-5标准评估: {len(dsm5_records)} 条")
    
    # 报告类型选择
    st.subheader("📋 选择报告类型")
    
    # 如果两种标准都有数据，允许选择
    if abc_records and dsm5_records:
        report_type = st.radio(
            "选择报告标准",
            ["综合报告（包含两种标准）", "仅ABC量表报告", "仅DSM-5标准报告"],
            horizontal=True
        )
        
        if report_type == "仅ABC量表报告":
            selected_records = abc_records
        elif report_type == "仅DSM-5标准报告":
            selected_records = dsm5_records
        else:
            selected_records = records
    else:
        selected_records = records
        report_type = "ABC量表报告" if abc_records else "DSM-5标准报告"
    
    col1, col2 = st.columns(2)
    
    with col1:
        st.write("### 📄 标准报告")
        
        # 基础CSV报告
        if st.button("📊 下载评估数据 (CSV)", use_container_width=True):
            export_data = prepare_export_data_dual(selected_records)
            csv_content = export_to_csv(export_data)
            
            st.download_button(
                label="⬇️ 下载评估数据",
                data=csv_content,
                file_name=f"autism_assessment_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
                mime='text/csv'
            )
        
        # 行为记录下载
        if st.button("💬 下载行为观察记录 (TXT)", use_container_width=True):
            observation_content = create_observation_text_dual(selected_records)
            
            st.download_button(
                label="⬇️ 下载行为观察记录",
                data=export_to_text(observation_content),
                file_name=f"autism_behavior_observations_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.txt",
                mime='text/plain'
            )
        
        # JSON完整数据
        if st.button("🔧 下载完整数据 (JSON)", use_container_width=True):
            json_data = create_complete_json_data_dual(selected_records)
            json_str = export_to_json(json_data)
            
            st.download_button(
                label="⬇️ 下载完整数据",
                data=json_str.encode('utf-8'),
                file_name=f"autism_complete_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.json",
                mime='application/json'
            )
    
    with col2:
        st.write("### 📈 专业分析报告")
        
        # 生成分析报告
        if st.button("📊 生成统计分析", use_container_width=True):
            with st.spinner("正在生成分析报告..."):
                analysis = generate_clinical_analysis_dual(selected_records)
            
            st.success("✅ 分析报告生成完成！")
            
            # 显示分析预览
            with st.expander("📋 分析报告预览", expanded=True):
                display_analysis_preview(analysis, selected_records)
            
            # 提供分析报告下载
            analysis_json = export_to_json(analysis)
            st.download_button(
                label="⬇️ 下载分析报告 (JSON)",
                data=analysis_json.encode('utf-8'),
                file_name=f"autism_analysis_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.json",
                mime='application/json'
            )
        
        # Excel专业报告
        if EXCEL_AVAILABLE:
            if st.button("📋 生成专业Excel报告", use_container_width=True):
                with st.spinner("正在生成专业Excel报告..."):
                    analysis = generate_clinical_analysis_dual(selected_records)
                    excel_data = create_excel_report_dual(selected_records, analysis)
                
                if excel_data:
                    st.success("✅ 专业Excel报告生成完成！")
                    
                    st.download_button(
                        label="⬇️ 下载专业Excel报告",
                        data=excel_data,
                        file_name=f"autism_professional_report_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                        mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                    )
                else:
                    st.error("❌ Excel报告生成失败，请尝试其他格式")
        else:
            st.info("💡 Excel报告功能需要安装openpyxl模块")
            st.code("pip install openpyxl")
            
            # 替代详细报告
            if st.button("📊 生成详细文本报告", use_container_width=True):
                with st.spinner("正在生成详细报告..."):
                    analysis = generate_clinical_analysis_dual(selected_records)
                
                # 创建详细文本报告
                detailed_report = create_detailed_text_report_dual(selected_records, analysis)
                
                st.success("✅ 详细文本报告生成完成！")
                
                st.download_button(
                    label="⬇️ 下载详细文本报告",
                    data=export_to_text(detailed_report),
                    file_name=f"autism_detailed_report_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.txt",
                    mime='text/plain'
                )
        
        # 研究数据包
        if st.button("📦 生成完整研究数据包", use_container_width=True, type="primary"):
            with st.spinner("正在生成完整研究数据包..."):
                zip_data = create_research_package_dual(selected_records)
            
            st.success("✅ 完整研究数据包生成完成！")
            
            st.download_button(
                label="⬇️ 下载完整研究数据包 (ZIP)",
                data=zip_data,
                file_name=f"autism_research_package_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.zip",
                mime='application/zip'
            )
    
    # 数据统计概览
    display_data_overview_dual(selected_records)
    
    # 数据预览
    display_data_preview_dual(selected_records)


def prepare_export_data_dual(records):
    """准备导出数据 - 支持双标准"""
    export_data = []
    
    for record in records:
        assessment_standard = record.get('assessment_standard', 'ABC')
        profile = record.get('autism_profile', {})
        scores = record['evaluation_scores']
        
        # 基础信息
        export_row = {
            '评估ID': record['experiment_id'],
            '评估时间': record['timestamp'].strftime('%Y-%m-%d %H:%M:%S'),
            '评估标准': assessment_standard,
            '配置类型': record.get('template', '自定义'),
            '评估情境': record['scene'],
            '观察活动': record.get('activity', ''),
            '触发因素': record.get('trigger', ''),
            '备注': record.get('notes', '')
        }
        
        if assessment_standard == 'ABC':
            # ABC特定字段
            export_row.update({
                'ABC总分': record.get('abc_total_score', ''),
                'ABC严重程度': record.get('abc_severity', ''),
                '感觉领域得分': scores.get('感觉领域得分', ''),
                '交往领域得分': scores.get('交往领域得分', ''),
                '躯体运动领域得分': scores.get('躯体运动领域得分', ''),
                '语言领域得分': scores.get('语言领域得分', ''),
                '社交与自理领域得分': scores.get('社交与自理领域得分', '')
            })
            
            # ABC配置信息
            if profile:
                export_row.update({
                    '严重程度描述': profile.get('description', ''),
                    '感觉异常程度': f"{profile.get('sensory_abnormal', 0)*100:.0f}%",
                    '交往障碍程度': f"{profile.get('social_impairment', 0)*100:.0f}%",
                    '运动刻板程度': f"{profile.get('motor_stereotypy', 0)*100:.0f}%",
                    '语言缺陷程度': f"{profile.get('language_deficit', 0)*100:.0f}%",
                    '自理缺陷程度': f"{profile.get('self_care_deficit', 0)*100:.0f}%",
                    '行为频率': f"{profile.get('behavior_frequency', 0)*100:.0f}%"
                })
            
            # 识别到的行为
            if 'identified_behaviors' in record:
                all_behaviors = []
                for domain, behaviors in record['identified_behaviors'].items():
                    all_behaviors.extend(behaviors)
                export_row['识别到的行为'] = '; '.join(all_behaviors[:10])
                
        else:  # DSM-5
            # DSM-5特定字段
            core_severity = (scores.get('社交互动质量', 0) + 
                           scores.get('沟通交流能力', 0) + 
                           scores.get('刻板重复行为', 0)) / 3
            
            export_row.update({
                '社交互动缺陷': scores.get('社交互动质量', ''),
                '沟通交流缺陷': scores.get('沟通交流能力', ''),
                '刻板重复行为': scores.get('刻板重复行为', ''),
                '感官处理异常': scores.get('感官处理能力', ''),
                '情绪调节困难': scores.get('情绪行为调节', ''),
                '认知适应缺陷': scores.get('认知适应功能', ''),
                '核心症状综合': round(core_severity, 2)
            })
            
            # DSM-5配置信息
            if profile:
                export_row.update({
                    'DSM5严重程度': profile.get('dsm5_severity', ''),
                    '所需支持水平': profile.get('support_needs', ''),
                    '社交沟通设置': profile.get('social_communication', ''),
                    '刻板行为设置': profile.get('restricted_repetitive', ''),
                    '感官处理设置': profile.get('sensory_processing', ''),
                    '认知功能设置': profile.get('cognitive_function', ''),
                    '适应行为设置': profile.get('adaptive_behavior', ''),
                    '语言水平设置': profile.get('language_level', ''),
                    '特殊兴趣': profile.get('special_interests', '')
                })
        
        export_data.append(export_row)
    
    return export_data


def create_observation_text_dual(records):
    """创建行为观察记录文本 - 支持双标准"""
    observation_content = []
    observation_content.append("=" * 70)
    observation_content.append("孤独症儿童行为观察记录")
    observation_content.append("基于ABC量表和DSM-5诊断标准")
    observation_content.append("=" * 70)
    observation_content.append(f"生成时间: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    observation_content.append(f"评估记录总数: {len(records)}")
    
    # 统计两种标准的数量
    abc_count = len([r for r in records if r.get('assessment_standard', 'ABC') == 'ABC'])
    dsm5_count = len([r for r in records if r.get('assessment_standard', 'ABC') == 'DSM5'])
    observation_content.append(f"ABC量表评估: {abc_count} 条 | DSM-5标准评估: {dsm5_count} 条")
    
    observation_content.append("=" * 70)
    observation_content.append("")
    
    for i, record in enumerate(records, 1):
        assessment_standard = record.get('assessment_standard', 'ABC')
        
        observation_content.append(f"\n【评估记录 {i}】")
        observation_content.append(f"评估ID: {record['experiment_id']}")
        observation_content.append(f"评估时间: {record['timestamp'].strftime('%Y-%m-%d %H:%M:%S')}")
        observation_content.append(f"评估标准: {assessment_standard}")
        observation_content.append(f"配置类型: {record.get('template', '自定义')}")
        observation_content.append(f"评估情境: {record['scene']}")
        observation_content.append(f"观察活动: {record.get('activity', '未指定')}")
        observation_content.append(f"触发因素: {record.get('trigger', '未指定')}")
        
        if assessment_standard == 'ABC':
            observation_content.append(f"ABC总分: {record.get('abc_total_score', 'N/A')}")
            observation_content.append(f"ABC严重程度: {record.get('abc_severity', 'N/A')}")
            observation_content.append("-" * 50)
            
            observation_content.append("各领域得分:")
            for metric, score in record['evaluation_scores'].items():
                observation_content.append(f"  • {metric}: {score}")
            
            if 'identified_behaviors' in record and record['identified_behaviors']:
                observation_content.append("识别到的行为:")
                for domain, behaviors in record['identified_behaviors'].items():
                    if behaviors:
                        observation_content.append(f"  {domain}:")
                        for behavior in behaviors:
                            observation_content.append(f"    - {behavior}")
        else:  # DSM-5
            core_severity = (record['evaluation_scores'].get('社交互动质量', 0) + 
                           record['evaluation_scores'].get('沟通交流能力', 0) + 
                           record['evaluation_scores'].get('刻板重复行为', 0)) / 3
            
            if record.get('autism_profile'):
                profile = record['autism_profile']
                observation_content.append(f"DSM-5严重程度: {profile.get('dsm5_severity', '')}")
                observation_content.append(f"所需支持水平: {profile.get('support_needs', '')}")
            
            observation_content.append(f"核心症状综合严重度: {core_severity:.2f}/5.0")
            observation_content.append("-" * 50)
            
            observation_content.append("临床评估得分:")
            for metric, score in record['evaluation_scores'].items():
                observation_content.append(f"  • {metric}: {score}/5.0")
            
            if 'clinical_observations' in record and record['clinical_observations']:
                observation_content.append("临床观察要点:")
                for category, observations in record['clinical_observations'].items():
                    if observations:
                        observation_content.append(f"  {category}: {', '.join(observations)}")
        
        observation_content.append("行为观察对话:")
        observation_content.append(record['dialogue'])
        observation_content.append("-" * 50)
        observation_content.append("")
    
    return observation_content


def create_complete_json_data_dual(records):
    """创建完整的JSON数据 - 支持双标准"""
    # 分别分析两种标准的数据
    abc_records = [r for r in records if r.get('assessment_standard', 'ABC') == 'ABC']
    dsm5_records = [r for r in records if r.get('assessment_standard', 'ABC') == 'DSM5']
    
    json_data = {
        'assessment_report': {
            'report_metadata': {
                'generation_time': datetime.datetime.now().isoformat(),
                'assessment_tools': 'ABC量表 & DSM-5诊断标准',
                'total_assessments': len(records),
                'abc_assessments': len(abc_records),
                'dsm5_assessments': len(dsm5_records),
                'platform_version': '双标准版 v2.0'
            },
            'assessment_summary': generate_clinical_analysis_dual(records),
            'detailed_assessments': []
        }
    }
    
    for record in records:
        assessment_record = record.copy()
        assessment_record['timestamp'] = record['timestamp'].isoformat()
        
        # 添加计算字段
        if record.get('assessment_standard', 'ABC') == 'DSM5':
            core_severity = (record['evaluation_scores'].get('社交互动质量', 0) + 
                           record['evaluation_scores'].get('沟通交流能力', 0) + 
                           record['evaluation_scores'].get('刻板重复行为', 0)) / 3
            assessment_record['core_symptom_severity'] = round(core_severity, 2)
        
        json_data['assessment_report']['detailed_assessments'].append(assessment_record)
    
    return json_data


def create_excel_report_dual(records, analysis):
    """创建Excel报告 - 支持双标准"""
    if not EXCEL_AVAILABLE:
        return None
    
    workbook = create_excel_workbook()
    
    # 1. 评估概览
    overview_sheet = workbook.create_sheet("评估概览")
    overview_sheet.append(["孤独症儿童行为评估报告"])
    overview_sheet.append(["基于ABC量表和DSM-5诊断标准"])
    overview_sheet.append([])
    overview_sheet.append(["报告生成时间", datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')])
    
    # 分别统计两种标准
    abc_records = [r for r in records if r.get('assessment_standard', 'ABC') == 'ABC']
    dsm5_records = [r for r in records if r.get('assessment_standard', 'ABC') == 'DSM5']
    
    overview_sheet.append(["ABC量表评估数", len(abc_records)])
    overview_sheet.append(["DSM-5标准评估数", len(dsm5_records)])
    overview_sheet.append([])
    
    # 显示分析结果
    if '评估概况' in analysis:
        overview_sheet.append(["评估概况"])
        for key, value in analysis['评估概况'].items():
            overview_sheet.append([key, value])
    
    # 2. ABC评估数据（如果有）
    if abc_records:
        abc_sheet = workbook.create_sheet("ABC评估数据")
        abc_headers = ["评估ID", "时间", "配置类型", "ABC严重程度", "评估情境", 
                      "ABC总分", "感觉领域", "交往领域", "躯体运动", 
                      "语言领域", "社交自理", "主要行为表现"]
        abc_sheet.append(abc_headers)
        
        for record in abc_records:
            scores = record['evaluation_scores']
            
            # 提取主要行为
            main_behaviors = []
            if 'identified_behaviors' in record:
                for behaviors in record['identified_behaviors'].values():
                    main_behaviors.extend(behaviors[:2])
            main_behaviors_str = '; '.join(main_behaviors[:5])
            
            row = [
                record['experiment_id'],
                record['timestamp'].strftime('%Y-%m-%d %H:%M:%S'),
                record.get('template', '自定义'),
                record.get('abc_severity', ''),
                record['scene'],
                record.get('abc_total_score', ''),
                scores.get('感觉领域得分', ''),
                scores.get('交往领域得分', ''),
                scores.get('躯体运动领域得分', ''),
                scores.get('语言领域得分', ''),
                scores.get('社交与自理领域得分', ''),
                main_behaviors_str
            ]
            abc_sheet.append(row)
    
    # 3. DSM-5评估数据（如果有）
    if dsm5_records:
        dsm5_sheet = workbook.create_sheet("DSM5评估数据")
        dsm5_headers = ["评估ID", "时间", "严重程度", "评估情境", 
                       "社交互动缺陷", "沟通交流缺陷", "刻板重复行为", 
                       "感官处理异常", "情绪调节困难", "认知适应缺陷", 
                       "核心症状综合", "DSM5分级", "支持需求"]
        dsm5_sheet.append(dsm5_headers)
        
        for record in dsm5_records:
            scores = record['evaluation_scores']
            profile = record.get('autism_profile', {})
            
            core_severity = (scores.get('社交互动质量', 0) + 
                           scores.get('沟通交流能力', 0) + 
                           scores.get('刻板重复行为', 0)) / 3
            
            row = [
                record['experiment_id'],
                record['timestamp'].strftime('%Y-%m-%d %H:%M:%S'),
                record.get('template', '自定义'),
                record['scene'],
                scores.get('社交互动质量', ''),
                scores.get('沟通交流能力', ''),
                scores.get('刻板重复行为', ''),
                scores.get('感官处理能力', ''),
                scores.get('情绪行为调节', ''),
                scores.get('认知适应功能', ''),
                f"{core_severity:.2f}",
                profile.get('dsm5_severity', ''),
                profile.get('support_needs', '')
            ]
            dsm5_sheet.append(row)
    
    # 4. 统计分析
    if any(key in analysis for key in ['ABC分析', 'DSM5分析', '严重程度分布']):
        stats_sheet = workbook.create_sheet("统计分析")
        stats_sheet.append(["统计分析结果"])
        stats_sheet.append([])
        
        # ABC统计
        if 'ABC分析' in analysis and analysis['ABC分析']:
            stats_sheet.append(["ABC量表分析"])
            for key, value in analysis['ABC分析'].items():
                if isinstance(value, dict):
                    stats_sheet.append([key])
                    for k, v in value.items():
                        stats_sheet.append([f"  {k}", v])
                else:
                    stats_sheet.append([key, value])
            stats_sheet.append([])
        
        # DSM-5统计
        if 'DSM5分析' in analysis and analysis['DSM5分析']:
            stats_sheet.append(["DSM-5标准分析"])
            for key, value in analysis['DSM5分析'].items():
                if isinstance(value, dict):
                    stats_sheet.append([key])
                    for k, v in value.items():
                        stats_sheet.append([f"  {k}", v])
                else:
                    stats_sheet.append([key, value])
    
    # 5. 对话记录
    dialogue_sheet = workbook.create_sheet("对话记录")
    dialogue_sheet.append(["评估ID", "评估标准", "严重程度", "评估情境", "对话内容"])
    
    for record in records[:50]:  # 限制数量避免文件过大
        dialogue_sheet.append([
            record['experiment_id'],
            record.get('assessment_standard', 'ABC'),
            record.get('template', '自定义'),
            record['scene'],
            record['dialogue'][:1000] + '...' if len(record['dialogue']) > 1000 else record['dialogue']
        ])
    
    # 应用样式
    apply_excel_styles(workbook, header_color="4472C4", header_font_color="FFFFFF")
    
    # 特殊样式处理
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value and any(keyword in str(cell.value) for keyword in ['重度', '中度', '需要支持']):
                    from openpyxl.styles import PatternFill
                    cell.fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
    
    return export_to_excel(workbook)


def create_detailed_text_report_dual(records, analysis):
    """创建详细文本报告 - 支持双标准"""
    detailed_report = []
    detailed_report.append("孤独症儿童评估详细报告")
    detailed_report.append("=" * 50)
    detailed_report.append(f"报告生成时间: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    detailed_report.append(f"评估标准: ABC量表 & DSM-5诊断标准")
    
    # 统计两种标准
    abc_count = len([r for r in records if r.get('assessment_standard', 'ABC') == 'ABC'])
    dsm5_count = len([r for r in records if r.get('assessment_standard', 'ABC') == 'DSM5'])
    
    detailed_report.append(f"总评估数: {len(records)} (ABC: {abc_count}, DSM-5: {dsm5_count})")
    detailed_report.append("")
    
    # 添加评估概况
    detailed_report.append("一、评估概况")
    detailed_report.append("-" * 30)
    for key, value in analysis.get('评估概况', {}).items():
        detailed_report.append(f"{key}: {value}")
    detailed_report.append("")
    
    # ABC量表分析（如果有）
    if abc_count > 0 and 'ABC分析' in analysis:
        detailed_report.append("二、ABC量表评估分析")
        detailed_report.append("-" * 30)
        
        if 'ABC总分统计' in analysis['ABC分析']:
            detailed_report.append("ABC总分统计:")
            for key, value in analysis['ABC分析']['ABC总分统计'].items():
                detailed_report.append(f"  - {key}: {value}")
        
        if '严重程度分布' in analysis['ABC分析']:
            detailed_report.append("\n严重程度分布:")
            for key, value in analysis['ABC分析']['严重程度分布'].items():
                detailed_report.append(f"  - {key}: {value}")
        
        detailed_report.append("")
    
    # DSM-5分析（如果有）
    if dsm5_count > 0 and 'DSM5分析' in analysis:
        detailed_report.append("三、DSM-5标准评估分析")
        detailed_report.append("-" * 30)
        
        if '整体表现' in analysis['DSM5分析']:
            detailed_report.append("整体临床表现:")
            for key, value in analysis['DSM5分析']['整体表现'].items():
                detailed_report.append(f"  - {key}: {value}")
        
        detailed_report.append("")
    
    # 临床发现与建议
    if '临床发现与建议' in analysis:
        detailed_report.append("四、临床发现与建议")
        detailed_report.append("-" * 30)
        for i, finding in enumerate(analysis['临床发现与建议'], 1):
            detailed_report.append(f"{i}. {finding}")
        detailed_report.append("")
    
    # 个案明细
    detailed_report.append("五、个案评估明细")
    detailed_report.append("-" * 30)
    
    # 按评估标准分组显示
    for standard in ['ABC', 'DSM5']:
        standard_records = [r for r in records if r.get('assessment_standard', 'ABC') == standard]
        if standard_records:
            detailed_report.append(f"\n{standard}标准评估 ({len(standard_records)}条):")
            
            for i, record in enumerate(standard_records[:10], 1):  # 限制显示数量
                detailed_report.append(f"\n个案 {i}: {record['experiment_id']}")
                detailed_report.append(f"  评估时间: {record['timestamp'].strftime('%Y-%m-%d %H:%M')}")
                detailed_report.append(f"  配置类型: {record.get('template', '自定义')}")
                detailed_report.append(f"  评估情境: {record['scene']}")
                
                if standard == 'ABC':
                    detailed_report.append(f"  ABC总分: {record.get('abc_total_score', 'N/A')}")
                    detailed_report.append(f"  严重程度: {record.get('abc_severity', 'N/A')}")
                else:
                    core_severity = (record['evaluation_scores'].get('社交互动质量', 0) + 
                                   record['evaluation_scores'].get('沟通交流能力', 0) + 
                                   record['evaluation_scores'].get('刻板重复行为', 0)) / 3
                    detailed_report.append(f"  核心症状综合: {core_severity:.2f}/5.0")
                    severity_level = "轻度" if core_severity < 2.5 else "中度" if core_severity < 3.5 else "重度"
                    detailed_report.append(f"  严重程度判断: {severity_level}")
    
    return detailed_report


def create_research_package_dual(records):
    """创建研究数据包 - 支持双标准"""
    # 生成分析
    analysis = generate_clinical_analysis_dual(records)
    
    # 准备各种格式的数据
    files_dict = {}
    
    # 1. 基础数据CSV
    export_data = prepare_export_data_dual(records)
    files_dict["01_评估数据.csv"] = export_to_csv(export_data)
    
    # 2. 专业分析报告
    if EXCEL_AVAILABLE:
        excel_data = create_excel_report_dual(records, analysis)
        if excel_data:
            files_dict["02_专业分析报告.xlsx"] = excel_data
    
    if "02_专业分析报告.xlsx" not in files_dict:
        # Excel不可用时的文本报告
        detailed_report = create_detailed_text_report_dual(records, analysis)
        files_dict["02_专业分析报告.txt"] = '\n'.join(detailed_report)
    
    # 3. 完整研究数据JSON
    complete_data = create_complete_json_data_dual(records)
    files_dict["03_完整研究数据.json"] = export_to_json(complete_data)
    
    # 4. 行为观察记录
    observation_content = create_observation_text_dual(records)
    files_dict["04_行为观察记录.txt"] = '\n'.join(observation_content)
    
    # 5. README文件
    readme_content = create_readme_content_dual(records, EXCEL_AVAILABLE)
    files_dict["README.txt"] = readme_content
    
    return create_zip_package(files_dict)


def create_readme_content_dual(records, excel_available):
    """创建README内容 - 支持双标准"""
    abc_count = len([r for r in records if r.get('assessment_standard', 'ABC') == 'ABC'])
    dsm5_count = len([r for r in records if r.get('assessment_standard', 'ABC') == 'DSM5'])
    
    readme_content = f"""
孤独症儿童AI模拟实验平台 - 双标准版
研究数据包说明文档
======================================

生成时间: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
评估记录数: {len(records)}
- ABC量表评估: {abc_count}条
- DSM-5标准评估: {dsm5_count}条

评估标准说明:
------------
1. ABC量表（孤独症行为量表）
   - 包含57个行为项目，分为5个领域
   - 总分≥67分为孤独症
   - 提供详细的行为识别和量化评分

2. DSM-5诊断标准
   - 基于美国精神疾病诊断与统计手册第五版
   - 评估核心症状：社交沟通缺陷和刻板重复行为
   - 提供支持需求等级评估

文件说明:
--------
1. 01_评估数据.csv
   - 包含所有评估的核心数据
   - 适用于统计分析和数据可视化
   - 包含两种评估标准的数据

"""
    if excel_available:
        readme_content += """2. 02_专业分析报告.xlsx
   - 多工作表Excel专业报告
   - 分别展示ABC和DSM-5评估结果
   - 包含统计分析和对比
"""
    else:
        readme_content += """2. 02_专业分析报告.txt
   - 详细的文本格式专业报告
   - 包含统计分析和临床解释
   - 注意: Excel功能不可用，如需Excel格式请安装openpyxl
"""

    readme_content += """
3. 03_完整研究数据.json
   - 包含所有原始数据和元数据
   - 适用于程序处理和深度分析
   - 包含完整的评估记录

4. 04_行为观察记录.txt
   - 所有评估的对话记录
   - 用于质性分析和行为模式研究
   - 标注了使用的评估标准

5. README.txt
   - 本说明文档

评估指标说明:
-----------
ABC量表评分：
- 总分越高，症状越严重
- 各领域有不同的最高分
- 通过识别具体行为计算得分

DSM-5评分（1-5分制）：
- 1分: 无明显问题/正常范围
- 2分: 轻度问题/需要关注
- 3分: 中度问题/需要支持
- 4分: 明显问题/需要大量支持
- 5分: 严重问题/需要非常大量支持

使用建议:
--------
1. 对比分析:
   - 比较ABC和DSM-5评估结果的一致性
   - 分析不同标准下的严重程度判定
   - 研究两种标准的相关性

2. 临床应用:
   - ABC量表适合行为筛查和定量评估
   - DSM-5标准适合诊断分类和支持需求评估
   - 结合使用可获得更全面的评估

3. 研究应用:
   - 使用完整数据进行深度分析
   - 研究不同评估标准的效度
   - 开发新的评估工具

技术支持:
--------
- 如需Excel功能，请安装: pip install openpyxl
- 数据分析建议使用: pandas, numpy, scipy
- 可视化建议使用: matplotlib, plotly

参考文献:
--------
- Krug, D. A., Arick, J., & Almond, P. (1980). 
  Behavior checklist for identifying severely handicapped 
  individuals with high levels of autistic behavior.
  
- American Psychiatric Association. (2013). 
  Diagnostic and statistical manual of mental disorders (5th ed.).

质量保证:
--------
本平台严格遵循ABC量表和DSM-5的原始设计，
确保评估结果的专业性和可靠性。

版权声明:
--------
本数据包仅供学术研究和临床实践使用，
请遵循相关伦理规范和数据保护法规。
"""
    return readme_content


def display_data_overview_dual(records):
    """显示数据统计概览 - 支持双标准"""
    st.subheader("📈 数据统计概览")
    
    col_stat1, col_stat2, col_stat3, col_stat4 = st.columns(4)
    
    with col_stat1:
        st.metric("评估总数", len(records))
    
    with col_stat2:
        # 统计两种标准
        abc_count = len([r for r in records if r.get('assessment_standard', 'ABC') == 'ABC'])
        dsm5_count = len([r for r in records if r.get('assessment_standard', 'ABC') == 'DSM5'])
        st.metric("ABC/DSM5", f"{abc_count}/{dsm5_count}")
    
    with col_stat3:
        contexts = [r['scene'] for r in records]
        unique_contexts = len(set(contexts))
        st.metric("评估情境数", unique_contexts)
    
    with col_stat4:
        if records:
            time_span = (max(r['timestamp'] for r in records) - min(r['timestamp'] for r in records)).days
            st.metric("时间跨度(天)", time_span)


def display_data_preview_dual(records):
    """显示数据预览 - 支持双标准"""
    st.subheader("📊 数据预览")
    
    preview_data = []
    for record in records[:10]:
        assessment_standard = record.get('assessment_standard', 'ABC')
        
        preview_row = {
            '评估ID': record['experiment_id'][:20] + '...' if len(record['experiment_id']) > 20 else record['experiment_id'],
            '时间': record['timestamp'].strftime('%m-%d %H:%M'),
            '标准': assessment_standard,
            '配置': record.get('template', '自定义')[:8] + '...' if len(record.get('template', '自定义')) > 8 else record.get('template', '自定义'),
            '情境': record['scene'].replace('结构化', '结构')
        }
        
        if assessment_standard == 'ABC':
            preview_row['总分/严重度'] = f"{record.get('abc_total_score', 'N/A')}/{record.get('abc_severity', 'N/A')[:4]}"
            preview_row['行为数'] = sum(len(behaviors) for behaviors in record.get('identified_behaviors', {}).values())
        else:
            core_severity = (record['evaluation_scores'].get('社交互动质量', 0) + 
                           record['evaluation_scores'].get('沟通交流能力', 0) + 
                           record['evaluation_scores'].get('刻板重复行为', 0)) / 3
            severity_level = "轻度" if core_severity < 2.5 else "中度" if core_severity < 3.5 else "重度"
            preview_row['核心症状'] = f"{core_severity:.2f}"
            preview_row['程度'] = severity_level
        
        preview_data.append(preview_row)
    
    df_preview = pd.DataFrame(preview_data)
    st.dataframe(df_preview, use_container_width=True)
    
    if len(records) > 10:
        st.info(f"显示前10条记录，共{len(records)}条。完整数据请通过上方下载功能获取。")


def display_analysis_preview(analysis, records):
    """显示分析预览"""
    if analysis.get('评估概况'):
        st.write("**评估概况:**")
        for key, value in analysis['评估概况'].items():
            st.write(f"- {key}: {value}")
    
    # 分别显示ABC和DSM5的分析结果
    if 'ABC分析' in analysis:
        st.write("\n**ABC量表分析:**")
        if 'ABC总分统计' in analysis['ABC分析']:
            for key, value in analysis['ABC分析']['ABC总分统计'].items():
                st.write(f"- {key}: {value}")
        if '严重程度分布' in analysis['ABC分析']:
            st.write("- 严重程度分布:")
            for key, value in analysis['ABC分析']['严重程度分布'].items():
                st.write(f"  - {key}: {value}")
    
    if 'DSM5分析' in analysis:
        st.write("\n**DSM-5标准分析:**")
        if '整体表现' in analysis['DSM5分析']:
            for key, value in analysis['DSM5分析']['整体表现'].items():
                st.write(f"- {key}: {value}")
    
    if analysis.get('临床发现与建议'):
        st.write("\n**临床发现与建议:**")
        for finding in analysis['临床发现与建议']:
            st.write(f"- {finding}")


def generate_clinical_analysis_dual(records):
    """生成临床分析 - 支持双标准"""
    if not records:
        return {}
    
    analysis = {}
    
    # 基础统计
    analysis['评估概况'] = {
        '评估总数': len(records),
        '评估时间跨度': f"{min(r['timestamp'] for r in records).strftime('%Y-%m-%d')} 至 {max(r['timestamp'] for r in records).strftime('%Y-%m-%d')}",
        '涉及情境数': len(set(r['scene'] for r in records)),
        '涉及配置类型数': len(set(r.get('template', '自定义') for r in records))
    }
    
    # 分别分析ABC和DSM5数据
    abc_records = [r for r in records if r.get('assessment_standard', 'ABC') == 'ABC']
    dsm5_records = [r for r in records if r.get('assessment_standard', 'ABC') == 'DSM5']
    
    # ABC分析
    if abc_records:
        analysis['ABC分析'] = analyze_abc_records(abc_records)
    
    # DSM5分析
    if dsm5_records:
        analysis['DSM5分析'] = analyze_dsm5_records(dsm5_records)
    
    # 综合临床发现
    findings = []
    
    # 基于ABC的发现
    if abc_records:
        avg_abc_total = np.mean([r.get('abc_total_score', 0) for r in abc_records])
        if avg_abc_total >= 67:
            findings.append(f"ABC评估显示明确的孤独症表现（平均总分: {avg_abc_total:.1f}）")
        elif avg_abc_total >= 53:
            findings.append(f"ABC评估显示轻度孤独症表现（平均总分: {avg_abc_total:.1f}）")
    
    # 基于DSM5的发现
    if dsm5_records:
        core_severities = []
        for r in dsm5_records:
            core_severity = (r['evaluation_scores'].get('社交互动质量', 0) + 
                           r['evaluation_scores'].get('沟通交流能力', 0) + 
                           r['evaluation_scores'].get('刻板重复行为', 0)) / 3
            core_severities.append(core_severity)
        
        avg_core = np.mean(core_severities)
        if avg_core >= 4.0:
            findings.append(f"DSM-5评估显示重度核心症状（平均严重度: {avg_core:.2f}/5）")
        elif avg_core >= 3.0:
            findings.append(f"DSM-5评估显示中度核心症状（平均严重度: {avg_core:.2f}/5）")
    
    # 两种标准对比（如果都有数据）
    if abc_records and dsm5_records:
        findings.append("建议综合ABC量表和DSM-5标准制定个体化干预方案")
    
    # 情境相关建议
    context_scores = {}
    for record in records:
        context = record['scene']
        if context not in context_scores:
            context_scores[context] = []
        
        # 根据不同标准计算表现分数
        if record.get('assessment_standard', 'ABC') == 'ABC':
            # ABC总分越低表现越好
            score = record.get('abc_total_score', 0)
        else:
            # DSM5核心症状越低表现越好
            score = (record['evaluation_scores'].get('社交互动质量', 0) + 
                    record['evaluation_scores'].get('沟通交流能力', 0) + 
                    record['evaluation_scores'].get('刻板重复行为', 0)) / 3 * 20  # 转换到类似ABC的尺度
        
        context_scores[context].append(score)
    
    # 找出表现最好的情境
    best_context = min(context_scores.keys(), 
                      key=lambda x: np.mean(context_scores[x]))
    findings.append(f"在{best_context}中表现相对较好，可作为干预起点")
    
    analysis['临床发现与建议'] = findings
    
    return analysis


def analyze_abc_records(abc_records):
    """分析ABC记录"""
    abc_analysis = {}
    
    # ABC总分统计
    total_scores = [r.get('abc_total_score', 0) for r in abc_records]
    abc_analysis['ABC总分统计'] = {
        '平均总分': f"{np.mean(total_scores):.1f}",
        '总分范围': f"{np.min(total_scores):.0f}-{np.max(total_scores):.0f}",
        '标准差': f"{np.std(total_scores):.1f}",
        '中位数': f"{np.median(total_scores):.0f}"
    }
    
    # 严重程度分布
    severity_distribution = {}
    for record in abc_records:
        severity = record.get('abc_severity', '未知')
        if severity not in severity_distribution:
            severity_distribution[severity] = 0
        severity_distribution[severity] += 1
    
    abc_analysis['严重程度分布'] = {
        k: f"{v} ({v/len(abc_records)*100:.1f}%)" 
        for k, v in severity_distribution.items()
    }
    
    # 各领域得分分析
    domain_stats = {}
    domains = ['感觉领域得分', '交往领域得分', '躯体运动领域得分', '语言领域得分', '社交与自理领域得分']
    for domain in domains:
        scores = [r['evaluation_scores'].get(domain, 0) for r in abc_records]
        if scores:
            domain_stats[domain] = {
                '平均分': f"{np.mean(scores):.1f}",
                '最高分': f"{np.max(scores):.0f}",
                '最低分': f"{np.min(scores):.0f}"
            }
    
    abc_analysis['各领域得分'] = domain_stats
    
    # 高频行为分析
    all_behaviors = {}
    for record in abc_records:
        if 'identified_behaviors' in record:
            for domain, behaviors in record['identified_behaviors'].items():
                for behavior in behaviors:
                    if behavior not in all_behaviors:
                        all_behaviors[behavior] = 0
                    all_behaviors[behavior] += 1
    
    # 排序并返回前10个高频行为
    sorted_behaviors = sorted(all_behaviors.items(), key=lambda x: x[1], reverse=True)[:10]
    abc_analysis['高频行为'] = {
        behavior: f"出现{count}次 ({count/len(abc_records)*100:.1f}%)" 
        for behavior, count in sorted_behaviors
    }
    
    return abc_analysis


def analyze_dsm5_records(dsm5_records):
    """分析DSM5记录"""
    dsm5_analysis = {}
    
    # 整体表现
    all_social = [r['evaluation_scores'].get('社交互动质量', 0) for r in dsm5_records]
    all_comm = [r['evaluation_scores'].get('沟通交流能力', 0) for r in dsm5_records]
    all_repetitive = [r['evaluation_scores'].get('刻板重复行为', 0) for r in dsm5_records]
    all_sensory = [r['evaluation_scores'].get('感官处理能力', 0) for r in dsm5_records]
    all_emotion = [r['evaluation_scores'].get('情绪行为调节', 0) for r in dsm5_records]
    all_cognitive = [r['evaluation_scores'].get('认知适应功能', 0) for r in dsm5_records]
    
    dsm5_analysis['整体表现'] = {
        '社交互动缺陷': f"{np.mean(all_social):.2f} ± {np.std(all_social):.2f}",
        '沟通交流缺陷': f"{np.mean(all_comm):.2f} ± {np.std(all_comm):.2f}",
        '刻板重复行为': f"{np.mean(all_repetitive):.2f} ± {np.std(all_repetitive):.2f}",
        '感官处理异常': f"{np.mean(all_sensory):.2f} ± {np.std(all_sensory):.2f}",
        '情绪调节困难': f"{np.mean(all_emotion):.2f} ± {np.std(all_emotion):.2f}",
        '认知适应缺陷': f"{np.mean(all_cognitive):.2f} ± {np.std(all_cognitive):.2f}",
        '核心症状综合': f"{(np.mean(all_social) + np.mean(all_comm) + np.mean(all_repetitive))/3:.2f}"
    }
    
    # 严重程度分析
    severity_stats = {}
    for record in dsm5_records:
        template = record.get('template', '自定义')
        if template not in severity_stats:
            severity_stats[template] = {
                'count': 0,
                'social_scores': [],
                'comm_scores': [],
                'repetitive_scores': []
            }
        
        severity_stats[template]['count'] += 1
        severity_stats[template]['social_scores'].append(record['evaluation_scores'].get('社交互动质量', 0))
        severity_stats[template]['comm_scores'].append(record['evaluation_scores'].get('沟通交流能力', 0))
        severity_stats[template]['repetitive_scores'].append(record['evaluation_scores'].get('刻板重复行为', 0))
    
    dsm5_analysis['严重程度分组'] = {}
    for template, stats in severity_stats.items():
        core_avg = (np.mean(stats['social_scores']) + 
                   np.mean(stats['comm_scores']) + 
                   np.mean(stats['repetitive_scores'])) / 3
        
        dsm5_analysis['严重程度分组'][template] = {
            '样本数': stats['count'],
            '核心症状均值': f"{core_avg:.2f}",
            '社交缺陷': f"{np.mean(stats['social_scores']):.2f}",
            '沟通缺陷': f"{np.mean(stats['comm_scores']):.2f}",
            '刻板行为': f"{np.mean(stats['repetitive_scores']):.2f}"
        }
    
    return dsm5_analysis